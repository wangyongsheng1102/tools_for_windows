# -*- coding: utf-8 -*-
import ast
import csv
import datetime
from datetime import datetime as dt
import math
import os
import random
import subprocess
import sys
import time
from configparser import ConfigParser

import chardet
import openpyxl
import pandas as pd
import psycopg2
from PyQt5.QtCore import Qt, QTimer, QDateTime, QThread, pyqtSignal
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QTabWidget, QLabel, QLineEdit, \
    QPushButton, QTableWidget, QTableWidgetItem, QGroupBox, QMessageBox, QDialogButtonBox, QDialog, QMenu, QAction, \
    QTextEdit, QProgressBar, QCheckBox, QFileDialog, QComboBox, QSizePolicy
from openpyxl import Workbook as openWorkbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, PatternFill, Side, Font
from openpyxl.utils import get_column_letter
from psycopg2 import Error, sql
from xlsxwriter import Workbook

USER = ""
PASSWORD = ""
HOST = ""
PORT = ""
DATABASE = ""

CIS_USER = ""
CIS_PASSWORD = ""
CIS_HOST = ""
CIS_PORT = ""
CIS_DATABASE = ""

ORDER_USER = ""
ORDER_PASSWORD = ""
ORDER_HOST = ""
ORDER_PORT = ""
ORDER_DATABASE = ""

PORTAL_USER = ""
PORTAL_PASSWORD = ""
PORTAL_HOST = ""
PORTAL_PORT = ""
PORTAL_DATABASE = ""

WSL_NAME = ""

CURRENT_DATABASE = ""

LOGIN_SUCCESS = False

TARGET_FOLDER_IMPORT = ""
TARGET_FOLDER_EXPORT = ""

ENABILITY_SYSTEM = [
    'EnabilityCIS',
    'EnabilityOrder',
    'EnabilityPortal'
]


def get_program_path():
    """アプリのパスを取得"""
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def get_config_file_path():
    """コンフィグのパスを取得"""
    return os.path.join(get_program_path(), ".database_config.ini")


def load_config_content(tag):
    """コンフィグをロード"""
    config = ConfigParser()
    config_path = get_config_file_path()
    if os.path.exists(config_path) is False:
        raise
    config.read(config_path, encoding='utf-8')
    return config[tag] if tag in config else {}


def init_login_config_content():
    global USER, PASSWORD, HOST, PORT, DATABASE
    global CIS_USER, CIS_PASSWORD, CIS_HOST, CIS_PORT, CIS_DATABASE
    global ORDER_USER, ORDER_PASSWORD, ORDER_HOST, ORDER_PORT, ORDER_DATABASE
    global PORTAL_USER, PORTAL_PASSWORD, PORTAL_HOST, PORTAL_PORT, PORTAL_DATABASE
    global WSL_NAME
    try:
        login = load_config_content('Login')
    except Exception as e:
        set_message_box("CRITICAL", "コンフィグ", "コンフィグファイルが存在しませんが、チェックしてください。")
        return
    # if inputs['user'] is not None:
    #     USER = inputs['user']
    # if inputs['password'] is not None:
    #     PASSWORD = inputs['password']
    # if inputs['host'] is not None:
    #     HOST = inputs['host']
    # if inputs['port'] is not None:
    #     PORT = inputs['port']
    # if inputs['database'] is not None:
    #     DATABASE = inputs['database']

    if login['cis_user'] is not None:
        CIS_USER = login['cis_user']
    if login['cis_password'] is not None:
        CIS_PASSWORD = login['cis_password']
    if login['cis_host'] is not None:
        CIS_HOST = login['cis_host']
    if login['cis_port'] is not None:
        CIS_PORT = login['cis_port']
    if login['cis_database'] is not None:
        CIS_DATABASE = login['cis_database']

    if login['order_user'] is not None:
        ORDER_USER = login['order_user']
    if login['order_password'] is not None:
        ORDER_PASSWORD = login['order_password']
    if login['order_host'] is not None:
        ORDER_HOST = login['order_host']
    if login['order_port'] is not None:
        ORDER_PORT = login['order_port']
    if login['order_database'] is not None:
        ORDER_DATABASE = login['order_database']

    if login['portal_user'] is not None:
        PORTAL_USER = login['portal_user']
    if login['portal_password'] is not None:
        PORTAL_PASSWORD = login['portal_password']
    if login['portal_host'] is not None:
        PORTAL_HOST = login['portal_host']
    if login['portal_port'] is not None:
        PORTAL_PORT = login['portal_port']
    if login['portal_database'] is not None:
        PORTAL_DATABASE = login['portal_database']

    if login.get('wsl_name') is not None:
        WSL_NAME = login['wsl_name']


def save_file_paths(self):
    """コンフィグにパスインフォを保存"""
    config = ConfigParser()
    config_path = get_config_file_path()
    if os.path.exists(config_path) is False:
        return
    config.read(config_path, encoding='utf-8')
    if CIS_USER is not None and CIS_USER != "":
        config.set('Login', 'cis_user', CIS_USER)
    if CIS_PASSWORD is not None and CIS_PASSWORD != "":
        config.set('Login', 'cis_password', CIS_PASSWORD)
    if CIS_HOST is not None and CIS_HOST != "":
        config.set('Login', 'cis_host', CIS_HOST)
    if CIS_PORT is not None and CIS_PORT != "":
        config.set('Login', 'cis_port', CIS_PORT)
    if CIS_DATABASE is not None and CIS_DATABASE != "":
        config.set('Login', 'cis_database', CIS_DATABASE)
    if ORDER_USER is not None and ORDER_USER != "":
        config.set('Login', 'order_user', ORDER_USER)
    if ORDER_PASSWORD is not None and ORDER_PASSWORD != "":
        config.set('Login', 'order_password', ORDER_PASSWORD)
    if ORDER_HOST is not None and ORDER_HOST != "":
        config.set('Login', 'order_host', ORDER_HOST)
    if ORDER_PORT is not None and ORDER_PORT != "":
        config.set('Login', 'order_port', ORDER_PORT)
    if ORDER_DATABASE is not None and ORDER_DATABASE != "":
        config.set('Login', 'order_database', ORDER_DATABASE)
    if PORTAL_USER is not None and PORTAL_USER != "":
        config.set('Login', 'portal_user', PORTAL_USER)
    if PORTAL_PASSWORD is not None and PORTAL_PASSWORD != "":
        config.set('Login', 'portal_password', PORTAL_PASSWORD)
    if PORTAL_HOST is not None and PORTAL_HOST != "":
        config.set('Login', 'portal_host', PORTAL_HOST)
    if PORTAL_PORT is not None and PORTAL_PORT != "":
        config.set('Login', 'portal_port', PORTAL_PORT)
    if PORTAL_DATABASE is not None and PORTAL_DATABASE != "":
        config.set('Login', 'portal_database', PORTAL_DATABASE)

    if WSL_NAME is not None and WSL_NAME != "":
        config.set('Login', 'wsl_name', WSL_NAME)

    with open(get_config_file_path(), 'w', encoding='utf-8') as configfile:
        config.write(configfile)


def set_message_box(message_type, title, context):
    """メッセージを反映"""
    if message_type == 'WARNING':
        QMessageBox.warning(None, title, context)
    if message_type == 'CRITICAL':
        QMessageBox.critical(None, title, context)
    if message_type == 'INFO':
        QMessageBox.information(None, title, context)
    if message_type == 'QUESTION':
        QMessageBox.question(None, title, context)


def check_conn(user, password, host, port, database):
    try:
        connection = psycopg2.connect(user=user,
                                      password=password,
                                      host=host,
                                      port=int(port),
                                      database=database)

        cursor = connection.cursor()
    except (Exception, Error) as error:
        print("Error while connecting to PostgreSQL", error)
        return False, error
    if connection:
        cursor.close()
        connection.close()
        print("PostgreSQL connection is closed")
        return True, None


def generate_random_color(color):
    if color == "blue":
        blue = random.randint(150, 255)
        red = random.randint(0, blue - 50)
        green = random.randint(0, blue - 50)
    if color == "red":
        red = random.randint(200, 255)
        green = random.randint(0, 100)
        blue = random.randint(0, 100)
    if color == "yellow":
        red = random.randint(200, 255)
        green = random.randint(200, 255)
        blue = random.randint(0, 100)
    return red, green, blue


def get_str_before_first_dot(string, split):
    return string.split(split)[0]


# 按照列表的第一个元素进行排序，使用安全的转换方式
def safe_int(x):
    try:
        return int(x)
    except ValueError:
        return float('inf')  # 返回一个极大值，表示无穷大


def parse_date(item):
    try:
        return dt.strptime(item, '%Y-%m-%d %H:%M:%S.%f').strftime('%Y/%m/%d %H:%M:%S.%f')[:-3]
    except ValueError:
        try:
            return dt.strptime(item, '%Y-%m-%d %H:%M:%S').strftime('%Y/%m/%d %H:%M:%S')
        except ValueError:
            try:
                return dt.strptime(item, '%Y-%m-%d').strftime('%Y/%m/%d')
            except ValueError:
                return item


def number_to_excel_column(n):
    result = ""
    while n > 0:
        remainder = (n - 1) % 26
        result = chr(remainder + ord('A')) + result
        n = (n - 1) // 26
    return result


def read_csv(file_path):
    with open(file_path, 'r', newline='') as file:
        reader = csv.reader(file)
        data = [row for row in reader]
    return data


def write_excel(file_path, data):
    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(row)

    wb.save(file_path)


def compare_csv_to_excel(csv_file1, csv_file2, *args):
    args = args[0]
    deleted_records = []
    added_records = []
    modified_records_before = []
    modified_records_after = []

    with open(csv_file1, 'r', encoding='utf-8') as read_obj_1, open(csv_file2, 'r', encoding='utf-8') as read_obj_2:
        csv_reader = csv.reader(read_obj_1)
        csv_data1_set = list(csv_reader)
        del csv_data1_set[0]
        csv_reader = csv.reader(read_obj_2)
        csv_data2_set = list(csv_reader)
        del csv_data2_set[0]

        if len(csv_data1_set) == len(csv_data2_set) == 0:
            return deleted_records, added_records, modified_records_before, modified_records_after

        # Create dictionaries for quick lookup based on specified indices
        dict1 = {tuple(item[i - 1] for i in args): item for item in csv_data1_set}
        dict2 = {tuple(item[i - 1] for i in args): item for item in csv_data2_set}

        # Check for new items in list2 that are not in list1
        for key in dict2:
            if key not in dict1:
                added_records.append(dict2[key])

        # Check for deleted items in list1 that are not in list2
        for key in dict1:
            if key not in dict2:
                deleted_records.append(dict1[key])

        # Check for modified items
        for key in dict1:
            if key in dict2:
                if dict1[key] != dict2[key]:
                    modified_records_before.append(dict1[key])
                    modified_records_after.append(dict2[key])

        return deleted_records, added_records, modified_records_before, modified_records_after


def dict_to_list(original_list):
    converted_list = []
    for item in original_list:
        converted_item = [value for key, value in item.items()]
        converted_list.append(converted_item)
    return converted_list


def find_tables(name, tables):
    for context in tables:
        if name == context[0]:
            return True
    return False


def compare_dicts(list_dict1, list_dict2):
    if len(list_dict1) != len(list_dict2):
        return [], []

    def filter_nan(d):
        filtered_dict = {}
        for key, value in d.items():
            if isinstance(value, float) and math.isnan(value):
                # continue
                filtered_dict[key] = None
            filtered_dict[key] = value
        return filtered_dict

    differences_before = []
    differences_after = []

    for idx, (dict1, dict2) in enumerate(zip(list_dict1, list_dict2)):
        dict1_filtered = filter_nan(dict1)
        dict2_filtered = filter_nan(dict2)

        if dict1_filtered != dict2_filtered:
            for key in dict1_filtered:
                if dict1_filtered[key] != dict2_filtered[key]:
                    differences_before.append(dict1_filtered)
                    differences_after.append(dict2_filtered)
                    print(f"索引 {idx}: 键 {key} 的值不相等: {dict1_filtered[key]} != {dict2_filtered[key]}")

    if not differences_before and not differences_after:
        # return "两个列表中的字典完全相同"
        return [], []
    else:
        return differences_before, differences_after


def find_matching_dicts(list_of_keys, list_of_dicts_to_search, list_of_dicts_to_match):
    matched_dicts = []

    for search_dict in list_of_dicts_to_search:
        for match_dict in list_of_dicts_to_match:
            match = True
            for key_tuple in list_of_keys:
                key_name, key_value = key_tuple
                if isinstance(match_dict, dict):
                    if match_dict.get(key_tuple) != search_dict.get(key_tuple):
                        match = False
                        break
                else:
                    match = False
                    break

            if match:
                matched_dicts.append(match_dict)
                break  # Once matched, no need to continue searching for this search_dict

    return matched_dicts


def parse_excel(path, tables):
    # 打开Excel文件
    excel_file = path  # 替换为你的Excel文件路径
    wb = openpyxl.load_workbook(excel_file)

    # 选择第一个工作表
    sheet = wb["データ準備"]
    data = []
    # 遍历每个单元格
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            # 在这里添加你的条件判断，假设条件为找到内容为 "条件" 的单元格
            if cell.value is not None and find_tables(cell.value, tables) is True:
                print("※" * 10, cell.value, find_tables(cell.value, tables))
                # data.append(cell.value)
                # 找到符合条件的单元格，获取其所在行号
                start_row = cell.row

                # 找到第二行开始直到空行的所有行内容
                current_row = start_row + 2
                while True:
                    current_cell = sheet.cell(row=current_row, column=cell.column).value  # 假设检查第一列是否为空行
                    if current_cell is None or current_cell == '':
                        break
                    # 获取当前行的所有单元格内容
                    row_data = [sheet.cell(row=current_row, column=col).value for col in range(1, sheet.max_column + 1)]
                    row_data.pop(0)
                    row_data.pop(0)
                    # row_data.insert(0, cell.value)
                    data.append(row_data)
                    current_row += 1

                # 打印获取的数据
                for row_data in data:
                    print(row_data)

                # 可以在此处做进一步处理，例如将数据写入CSV文件等

    # 关闭Excel文件
    wb.close()
    return data


# 自定义的线程类
class WorkerThread(QThread):
    # 定义信号，在线程中处理的任务完成后发射该信号
    thread_signal = pyqtSignal(str)

    def __init__(self, task_func, *args, **kwargs):
        super().__init__()
        self.task_func = task_func
        self.args = args
        self.kwargs = kwargs

    def run(self):
        try:
            # 执行任务函数，并传入参数
            result = self.task_func(*self.args, **self.kwargs)
            self.thread_signal.emit(f"Task completed: {result}")
        except Exception as e:
            self.thread_signal.emit(f"Task error: {str(e)}")


def detect_encoding(byte_data):
    result = chardet.detect(byte_data)
    encoding = result['encoding']
    return encoding


def decode_bytes(byte_data):
    encoding = detect_encoding(byte_data)
    if encoding == 'Windows-1252':
        try:
            return byte_data.decode("utf-16")
        except (UnicodeDecodeError, TypeError):
            return byte_data.decode('utf-8', errors='replace')
    else:
        try:
            return byte_data.decode(encoding)
        except (UnicodeDecodeError, TypeError):
            return byte_data.decode('utf-8', errors='replace')


class EventHandler:
    """EventHandler"""

    def __init__(self, parent):
        self.parent = parent

    def button1_click(self):
        try:
            connection = psycopg2.connect(user=self.parent.line_edit_1.text(),
                                          password=self.parent.line_edit_2.text(),
                                          host=self.parent.line_edit_3.text(),
                                          port=int(self.parent.line_edit_4.text()),
                                          database=self.parent.line_edit_5.text())

            cursor = connection.cursor()
        except (Exception, Error) as error:
            print("Error while connecting to PostgreSQL", error)
            set_message_box("WARNING", "データベース", error)
        finally:
            if connection:
                cursor.close()
                connection.close()
                print("PostgreSQL connection is closed")
                set_message_box("WARNING", "データベース", "finally")

    def app_exit(self):
        """退出ボタン"""
        msg_box = QMessageBox()
        msg_box.setWindowTitle("ツールメッセージ")
        msg_box.setText("ツールを終了したいですか。")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg_box.setDefaultButton(QMessageBox.No)
        msg_box.button(QMessageBox.Yes).setText("はい(&Y)")
        msg_box.button(QMessageBox.No).setText("いいえ(&N)")
        result = msg_box.exec_()
        if result == QMessageBox.Yes:
            save_file_paths(self)
            self.parent.close()

    def whole_database_csv_export(self):
        global USER, PASSWORD, HOST, PORT, DATABASE
        global CIS_USER, CIS_PASSWORD, CIS_HOST, CIS_PORT, CIS_DATABASE
        global ORDER_USER, ORDER_PASSWORD, ORDER_HOST, ORDER_PORT, ORDER_DATABASE
        global PORTAL_USER, PORTAL_PASSWORD, PORTAL_HOST, PORTAL_PORT, PORTAL_DATABASE
        if CURRENT_DATABASE == "EnabilityCIS":
            USER = CIS_USER
            PASSWORD = CIS_PASSWORD
            HOST = CIS_HOST
            PORT = CIS_PORT
            DATABASE = CIS_DATABASE
            schema = "unisys"
        if CURRENT_DATABASE == "EnabilityOrder":
            USER = ORDER_USER
            PASSWORD = ORDER_PASSWORD
            HOST = ORDER_HOST
            PORT = ORDER_PORT
            DATABASE = ORDER_DATABASE
            schema = "public"
        if CURRENT_DATABASE == "EnabilityPortal":
            USER = PORTAL_USER
            PASSWORD = PORTAL_PASSWORD
            HOST = PORTAL_HOST
            PORT = PORTAL_PORT
            DATABASE = PORTAL_DATABASE
            schema = "public"
        conn_params = {
            'user': USER,
            'password': PASSWORD,
            'host': HOST,
            'port': PORT,
            'database': DATABASE
        }
        # self.parent.progress_bar.setValue(0)

        # 连接到数据库
        connection = psycopg2.connect(**conn_params)
        cursor = connection.cursor()

        # 获取所有表名
        cursor.execute(f"""
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = '{schema}'
                AND table_type = 'BASE TABLE';
            """)
        tables = cursor.fetchall()

        os.makedirs(os.path.join(get_program_path(), "export_csv"), exist_ok=True)
        # 对每张表执行导出操作
        for i, table in enumerate(tables):
            table_name = table[0]

            # 构建查询语句，导出表数据
            query = sql.SQL("COPY {} TO STDOUT WITH CSV HEADER").format(
                sql.Identifier(table_name)
            )

            # 执行查询
            csv_file_path = os.path.join(get_program_path(), "export_csv", f"{table_name}.csv")
            with open(csv_file_path, 'w', newline='', encoding='utf-8') as f:
                cursor.copy_expert(query, f)
                self.update_process_bar(int(len(tables) / 100))

        set_message_box("WARNING", "データベース", "導出成功")
        cursor.close()
        connection.close()

    def handleComboBoxClick(self, index):
        global CURRENT_DATABASE
        selected_option = self.parent.switch_system.currentText()
        if CURRENT_DATABASE != selected_option:
            CURRENT_DATABASE = selected_option
            set_message_box("INFO", "データベース", f"データベースを「{CURRENT_DATABASE}」に切り替えが成功しました。")
            self.parent.setWindowTitle(f"BIP-データベースツール-Ver.1.0-Powered by PyQt5 - 「{WSL_NAME}」「{CURRENT_DATABASE}」")

            print("Selected option:", selected_option)

    def update_process_bar(self, add):
        self.parent.progress_bar.setValue(self.parent.progress_bar.value() + add)


def database_start(wsl_name):
    wsl_command = 'wsl -d ' + wsl_name + ' -- su - pg -c "cd ~ && source .bash_profile && pg_ctl start"'

    process = subprocess.Popen(wsl_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate()
    if stdout:
        set_message_box("INFO", "データベース", decode_bytes(stdout))
    if (stdout is None or stdout == "") and stderr:
        set_message_box("INFO", "データベース", decode_bytes(stderr))


class LoginDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("ログイン")
        self.setGeometry(700, 400, 400, 300)
        self.layout = QVBoxLayout(self)
        self.label_width = 70

        # Create tab widget
        self.tab_widget = QTabWidget()

        # Tab 1: EnabilityCIS
        self.tab1 = QWidget()
        self.tab_widget.addTab(self.tab1, "EnabilityCIS")
        self.tab1_layout = QVBoxLayout(self.tab1)
        self.tab1.setLayout(self.tab1_layout)

        # Add input fields to tab 1
        self.hbox_1 = QHBoxLayout()
        self.label_1 = QLabel("ユーザーID:")
        self.label_1.setFixedWidth(self.label_width)
        self.line_edit_1 = QLineEdit()
        self.hbox_1.addWidget(self.label_1)
        self.hbox_1.addWidget(self.line_edit_1)
        self.tab1_layout.addLayout(self.hbox_1)

        self.hbox_2 = QHBoxLayout()
        self.label_2 = QLabel("パスワード:")
        self.label_2.setFixedWidth(self.label_width)
        self.line_edit_2 = QLineEdit()
        self.hbox_2.addWidget(self.label_2)
        self.hbox_2.addWidget(self.line_edit_2)
        self.tab1_layout.addLayout(self.hbox_2)

        self.hbox_3 = QHBoxLayout()
        self.label_3 = QLabel("サーバー名:")
        self.label_3.setFixedWidth(self.label_width)
        self.line_edit_3 = QLineEdit()
        self.hbox_3.addWidget(self.label_3)
        self.hbox_3.addWidget(self.line_edit_3)
        self.tab1_layout.addLayout(self.hbox_3)

        self.hbox_4 = QHBoxLayout()
        self.label_4 = QLabel("ポート:")
        self.label_4.setFixedWidth(self.label_width)
        self.line_edit_4 = QLineEdit()
        self.hbox_4.addWidget(self.label_4)
        self.hbox_4.addWidget(self.line_edit_4)
        self.tab1_layout.addLayout(self.hbox_4)

        self.hbox_5 = QHBoxLayout()
        self.label_5 = QLabel("データベース:")
        self.label_5.setFixedWidth(self.label_width)
        self.line_edit_5 = QLineEdit()
        self.hbox_5.addWidget(self.label_5)
        self.hbox_5.addWidget(self.line_edit_5)
        self.tab1_layout.addLayout(self.hbox_5)

        # Tab 2: EnabilityOrder
        self.tab2 = QWidget()
        self.tab_widget.addTab(self.tab2, "EnabilityOrder")
        self.tab2_layout = QVBoxLayout(self.tab2)
        self.tab2.setLayout(self.tab2_layout)

        # Add input fields to tab 2
        self.hbox_6 = QHBoxLayout()
        self.label_6 = QLabel("ユーザーID:")
        self.label_6.setFixedWidth(self.label_width)
        self.line_edit_6 = QLineEdit()
        self.hbox_6.addWidget(self.label_6)
        self.hbox_6.addWidget(self.line_edit_6)
        self.tab2_layout.addLayout(self.hbox_6)

        self.hbox_7 = QHBoxLayout()
        self.label_7 = QLabel("パスワード:")
        self.label_7.setFixedWidth(self.label_width)
        self.line_edit_7 = QLineEdit()
        self.hbox_7.addWidget(self.label_7)
        self.hbox_7.addWidget(self.line_edit_7)
        self.tab2_layout.addLayout(self.hbox_7)

        self.hbox_8 = QHBoxLayout()
        self.label_8 = QLabel("サーバー名:")
        self.label_8.setFixedWidth(self.label_width)
        self.line_edit_8 = QLineEdit()
        self.hbox_8.addWidget(self.label_8)
        self.hbox_8.addWidget(self.line_edit_8)
        self.tab2_layout.addLayout(self.hbox_8)

        self.hbox_9 = QHBoxLayout()
        self.label_9 = QLabel("ポート:")
        self.label_9.setFixedWidth(self.label_width)
        self.line_edit_9 = QLineEdit()
        self.hbox_9.addWidget(self.label_9)
        self.hbox_9.addWidget(self.line_edit_9)
        self.tab2_layout.addLayout(self.hbox_9)

        self.hbox_10 = QHBoxLayout()
        self.label_10 = QLabel("データベース:")
        self.label_10.setFixedWidth(self.label_width)
        self.line_edit_10 = QLineEdit()
        self.hbox_10.addWidget(self.label_10)
        self.hbox_10.addWidget(self.line_edit_10)
        self.tab2_layout.addLayout(self.hbox_10)

        # Tab 3: EnabilityPortal
        self.tab3 = QWidget()
        self.tab_widget.addTab(self.tab3, "EnabilityPortal")
        self.tab3_layout = QVBoxLayout(self.tab3)
        self.tab3.setLayout(self.tab3_layout)

        # Add input fields to tab 3
        self.hbox_11 = QHBoxLayout()
        self.label_11 = QLabel("ユーザーID:")
        self.label_11.setFixedWidth(self.label_width)
        self.line_edit_11 = QLineEdit()
        self.hbox_11.addWidget(self.label_11)
        self.hbox_11.addWidget(self.line_edit_11)
        self.tab3_layout.addLayout(self.hbox_11)

        self.hbox_12 = QHBoxLayout()
        self.label_12 = QLabel("パスワード:")
        self.label_12.setFixedWidth(self.label_width)
        self.line_edit_12 = QLineEdit()
        self.hbox_12.addWidget(self.label_12)
        self.hbox_12.addWidget(self.line_edit_12)
        self.tab3_layout.addLayout(self.hbox_12)

        self.hbox_13 = QHBoxLayout()
        self.label_13 = QLabel("サーバー名:")
        self.label_13.setFixedWidth(self.label_width)
        self.line_edit_13 = QLineEdit()
        self.hbox_13.addWidget(self.label_13)
        self.hbox_13.addWidget(self.line_edit_13)
        self.tab3_layout.addLayout(self.hbox_13)

        self.hbox_14 = QHBoxLayout()
        self.label_14 = QLabel("ポート:")
        self.label_14.setFixedWidth(self.label_width)
        self.line_edit_14 = QLineEdit()
        self.hbox_14.addWidget(self.label_14)
        self.hbox_14.addWidget(self.line_edit_14)
        self.tab3_layout.addLayout(self.hbox_14)

        self.hbox_15 = QHBoxLayout()
        self.label_15 = QLabel("データベース:")
        self.label_15.setFixedWidth(self.label_width)
        self.line_edit_15 = QLineEdit()
        self.hbox_15.addWidget(self.label_15)
        self.hbox_15.addWidget(self.line_edit_15)
        self.tab3_layout.addLayout(self.hbox_15)

        self.layout.addWidget(self.tab_widget)

        # Add buttons (OK and Cancel)
        self.button_layout = QHBoxLayout()
        self.start_button = QPushButton('データベース起動')
        self.start_button.setStyleSheet("background-color: lightgray")
        self.start_button.clicked.connect(self.start_button_click)
        self.start_edit = QLineEdit()
        self.button_layout.addWidget(self.start_button)
        self.button_layout.addWidget(self.start_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.accepted.connect(self.on_ok_clicked)
        buttons.rejected.connect(self.reject)
        self.button_layout.addWidget(buttons)
        self.layout.addLayout(self.button_layout)
        init_login_config_content()
        if CIS_USER is not None:
            self.line_edit_1.setText(CIS_USER)
        if CIS_PASSWORD is not None:
            self.line_edit_2.setText(CIS_PASSWORD)
        if CIS_HOST is not None:
            self.line_edit_3.setText(CIS_HOST)
        if CIS_PORT is not None:
            self.line_edit_4.setText(CIS_PORT)
        if CIS_DATABASE is not None:
            self.line_edit_5.setText(CIS_DATABASE)
        if ORDER_USER is not None:
            self.line_edit_6.setText(ORDER_USER)
        if ORDER_PASSWORD is not None:
            self.line_edit_7.setText(ORDER_PASSWORD)
        if ORDER_HOST is not None:
            self.line_edit_8.setText(ORDER_HOST)
        if ORDER_PORT is not None:
            self.line_edit_9.setText(ORDER_PORT)
        if ORDER_DATABASE is not None:
            self.line_edit_10.setText(ORDER_DATABASE)
        if PORTAL_USER is not None:
            self.line_edit_11.setText(PORTAL_USER)
        if PORTAL_PASSWORD is not None:
            self.line_edit_12.setText(PORTAL_PASSWORD)
        if PORTAL_HOST is not None:
            self.line_edit_13.setText(PORTAL_HOST)
        if PORTAL_PORT is not None:
            self.line_edit_14.setText(PORTAL_PORT)
        if PORTAL_DATABASE is not None:
            self.line_edit_15.setText(PORTAL_DATABASE)

        if WSL_NAME is not None:
            self.start_edit.setText(WSL_NAME)

    def on_ok_clicked(self):
        global CIS_USER, CIS_PASSWORD, CIS_HOST, CIS_PORT, CIS_DATABASE
        global ORDER_USER, ORDER_PASSWORD, ORDER_HOST, ORDER_PORT, ORDER_DATABASE
        global PORTAL_USER, PORTAL_PASSWORD, PORTAL_HOST, PORTAL_PORT, PORTAL_DATABASE
        global CURRENT_DATABASE, LOGIN_SUCCESS
        global WSL_NAME
        CIS_USER = self.line_edit_1.text()
        CIS_PASSWORD = self.line_edit_2.text()
        CIS_HOST = self.line_edit_3.text()
        CIS_PORT = self.line_edit_4.text()
        CIS_DATABASE = self.line_edit_5.text()

        ORDER_USER = self.line_edit_6.text()
        ORDER_PASSWORD = self.line_edit_7.text()
        ORDER_HOST = self.line_edit_8.text()
        ORDER_PORT = self.line_edit_9.text()
        ORDER_DATABASE = self.line_edit_10.text()

        PORTAL_USER = self.line_edit_11.text()
        PORTAL_PASSWORD = self.line_edit_12.text()
        PORTAL_HOST = self.line_edit_13.text()
        PORTAL_PORT = self.line_edit_14.text()
        PORTAL_DATABASE = self.line_edit_15.text()

        WSL_NAME = self.start_edit.text()

        if WSL_NAME == "" or WSL_NAME is None:
            LOGIN_SUCCESS = False
            set_message_box("CRITICAL", "データベース", "WSLを指定してください。")
            return

        current_index = self.tab_widget.currentIndex()
        if current_index == 0:
            CURRENT_DATABASE = "EnabilityCIS"
            return_flag, context = check_conn(CIS_USER, CIS_PASSWORD, CIS_HOST, CIS_PORT, CIS_DATABASE)
        elif current_index == 1:
            CURRENT_DATABASE = "EnabilityOrder"
            return_flag, context = check_conn(ORDER_USER, ORDER_PASSWORD, ORDER_HOST, ORDER_PORT, ORDER_DATABASE)
        elif current_index == 2:
            CURRENT_DATABASE = "EnabilityPortal"
            return_flag, context = check_conn(PORTAL_USER, PORTAL_PASSWORD, PORTAL_HOST, PORTAL_PORT, PORTAL_DATABASE)
        else:
            CURRENT_DATABASE = None
        if return_flag is False:
            LOGIN_SUCCESS = False
            set_message_box("CRITICAL", "データベース", context.args[0])
        else:
            LOGIN_SUCCESS = True
            set_message_box("WARNING", "データベース", "データベース接続成功")

    def start_button_click(self):
        database_start(self.start_edit.text())


class BlinkingLabel(QLabel):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self._timer = QTimer(self)
        self._timer.timeout.connect(self.toggle_visibility)
        # self._timer.start(1000)
        self.base_text = ""
        self.dot_count = 0
        self.direction = 1

    def start_blinking(self):
        self._timer.start(1000)

    def stop_blinking(self):
        color = QColor(0, 0, 0)
        self.setStyleSheet("QLabel { color: %s }" % color.name())
        self._timer.stop()

    def toggle_visibility(self):
        self.base_text = self.text().replace(".", "")
        self.dot_count += self.direction

        if self.dot_count == 5:
            self.direction = -1
        elif self.dot_count == 0:
            self.direction = 1

        dots = '.' * self.dot_count
        self.setText(self.base_text + dots)

        # color = QColor(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
        red, green, blue = generate_random_color("yellow")
        color = QColor(red, green, blue)
        self.setStyleSheet("QLabel { color: %s }" % color.name())
        # self.setVisible(not self.isVisible())


def row_append(table, context, color, checkbox):
    # tab_table = self.parent.tabs.widget(index).layout().itemAt(0).widget()
    tab_table = table
    current_row_count = tab_table.rowCount()
    tab_table.setRowCount(current_row_count + 1)
    i = 0

    # Fill the new row with data
    for col, line in enumerate(context):
        if checkbox is True:
            layout = QHBoxLayout()
            layout.setAlignment(Qt.AlignCenter)
            checkbox = QCheckBox()
            layout.addWidget(checkbox)
            cellWidget = QWidget()
            cellWidget.setLayout(layout)
            tab_table.setCellWidget(current_row_count, 0, cellWidget)
            i = 1
        if line is None:
            line = ""
        item = QTableWidgetItem(f"{line}")
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        if color is not None:
            item.setBackground(color)
        tab_table.setItem(current_row_count, col + i, item)

    tab_table.resizeColumnsToContents()
    tab_table.resizeRowsToContents()
    QApplication.processEvents()
    # bottomRightItem = tab_table.item(current_row_count, 0)
    # tab_table.scrollToItem(bottomRightItem)
    tab_table.scrollToBottom()


def get_tables_names_from_db():
    global USER, PASSWORD, HOST, PORT, DATABASE
    global CIS_USER, CIS_PASSWORD, CIS_HOST, CIS_PORT, CIS_DATABASE
    global ORDER_USER, ORDER_PASSWORD, ORDER_HOST, ORDER_PORT, ORDER_DATABASE
    global PORTAL_USER, PORTAL_PASSWORD, PORTAL_HOST, PORTAL_PORT, PORTAL_DATABASE
    global TARGET_FOLDER_IMPORT, TARGET_FOLDER_EXPORT
    if CURRENT_DATABASE == "EnabilityCIS":
        USER = CIS_USER
        PASSWORD = CIS_PASSWORD
        HOST = CIS_HOST
        PORT = CIS_PORT
        DATABASE = CIS_DATABASE
        schema = "unisys"
    if CURRENT_DATABASE == "EnabilityOrder":
        USER = ORDER_USER
        PASSWORD = ORDER_PASSWORD
        HOST = ORDER_HOST
        PORT = ORDER_PORT
        DATABASE = ORDER_DATABASE
        schema = "public"
    if CURRENT_DATABASE == "EnabilityPortal":
        USER = PORTAL_USER
        PASSWORD = PORTAL_PASSWORD
        HOST = PORTAL_HOST
        PORT = PORTAL_PORT
        DATABASE = PORTAL_DATABASE
        schema = "public"
    conn_params = {
        'user': USER,
        'password': PASSWORD,
        'host': HOST,
        'port': PORT,
        'database': DATABASE
    }

    # 连接到数据库
    connection = psycopg2.connect(**conn_params)
    cursor = connection.cursor()

    # 获取所有表名
    cursor.execute(f"""
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema = '{schema}'
            AND table_type = 'BASE TABLE';
        """)
    tables = cursor.fetchall()
    cursor.close()
    connection.close()

    return tables


def get_table_name_from_db(table_name, type):
    global USER, PASSWORD, HOST, PORT, DATABASE
    global CIS_USER, CIS_PASSWORD, CIS_HOST, CIS_PORT, CIS_DATABASE
    global ORDER_USER, ORDER_PASSWORD, ORDER_HOST, ORDER_PORT, ORDER_DATABASE
    global PORTAL_USER, PORTAL_PASSWORD, PORTAL_HOST, PORTAL_PORT, PORTAL_DATABASE
    global TARGET_FOLDER_IMPORT, TARGET_FOLDER_EXPORT
    if CURRENT_DATABASE == "EnabilityCIS":
        USER = CIS_USER
        PASSWORD = CIS_PASSWORD
        HOST = CIS_HOST
        PORT = CIS_PORT
        DATABASE = CIS_DATABASE
        schema = "unisys"
    if CURRENT_DATABASE == "EnabilityOrder":
        USER = ORDER_USER
        PASSWORD = ORDER_PASSWORD
        HOST = ORDER_HOST
        PORT = ORDER_PORT
        DATABASE = ORDER_DATABASE
        schema = "public"
    if CURRENT_DATABASE == "EnabilityPortal":
        USER = PORTAL_USER
        PASSWORD = PORTAL_PASSWORD
        HOST = PORTAL_HOST
        PORT = PORTAL_PORT
        DATABASE = PORTAL_DATABASE
        schema = "public"
    conn_params = {
        'user': USER,
        'password': PASSWORD,
        'host': HOST,
        'port': PORT,
        'database': DATABASE
    }
    connection = psycopg2.connect(**conn_params)
    cursor = connection.cursor()
    if type == 'table':
        cursor.execute("""
            SELECT pgd.description
            FROM pg_description pgd
            JOIN pg_class pgc ON pgd.objoid = pgc.oid
            WHERE pgc.relname = %s
            and pgd.objsubid = 0;
        """, (table_name,))
    else:
        cursor.execute("""
                    SELECT pgd.description
                    FROM pg_description pgd
                    JOIN pg_class pgc ON pgd.objoid = pgc.oid
                    WHERE pgc.relname = %s
                    and pgd.objsubid != 0;
                """, (table_name,))
    table_info = cursor.fetchall()

    cursor.close()
    connection.close()

    return table_info


def get_columns_from_db(table_name):
    global USER, PASSWORD, HOST, PORT, DATABASE
    global CIS_USER, CIS_PASSWORD, CIS_HOST, CIS_PORT, CIS_DATABASE
    global ORDER_USER, ORDER_PASSWORD, ORDER_HOST, ORDER_PORT, ORDER_DATABASE
    global PORTAL_USER, PORTAL_PASSWORD, PORTAL_HOST, PORTAL_PORT, PORTAL_DATABASE
    global TARGET_FOLDER_IMPORT, TARGET_FOLDER_EXPORT
    if CURRENT_DATABASE == "EnabilityCIS":
        USER = CIS_USER
        PASSWORD = CIS_PASSWORD
        HOST = CIS_HOST
        PORT = CIS_PORT
        DATABASE = CIS_DATABASE
        schema = "unisys"
    if CURRENT_DATABASE == "EnabilityOrder":
        USER = ORDER_USER
        PASSWORD = ORDER_PASSWORD
        HOST = ORDER_HOST
        PORT = ORDER_PORT
        DATABASE = ORDER_DATABASE
        schema = "public"
    if CURRENT_DATABASE == "EnabilityPortal":
        USER = PORTAL_USER
        PASSWORD = PORTAL_PASSWORD
        HOST = PORTAL_HOST
        PORT = PORTAL_PORT
        DATABASE = PORTAL_DATABASE
        schema = "public"
    conn_params = {
        'user': USER,
        'password': PASSWORD,
        'host': HOST,
        'port': PORT,
        'database': DATABASE
    }
    connection = psycopg2.connect(**conn_params)
    cursor = connection.cursor()
    cursor.execute(f"""
                        SELECT 
                            column_name
                        FROM 
                            information_schema.columns
                        WHERE 
                            table_name = '{table_name}'
                        ORDER BY 
                            ordinal_position;
                    """)

    rows = cursor.fetchall()
    list_of_columns = []

    for row in rows:
        list_of_columns.append(row)
    cursor.close()
    connection.close()

    return list_of_columns


def write_to_excel(work_sheet, row_count, row_context, style_set, start_index):
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    font = Font(bold=True, name='MS Gothic')

    font_for_write = Font(name='MS Gothic')

    fill = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid')
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_bold_font = Font(color="FF0000", bold=True, name='MS Gothic')
    i = start_index
    if style_set == 'border':
        for item in row_context:
            cell = work_sheet.cell(row=row_count + 1, column=i, value=item)
            if item is not None:
                if i > 2:
                    cell.border = border
                cell.number_format = '@'
                cell.font = font_for_write
            i += 1
    elif style_set == 'border-yellow':
        for item in row_context:
            cell = work_sheet.cell(row=row_count + 1, column=i, value=item)
            if item is not None:
                if i > 2:
                    cell.border = border
                    if item != work_sheet.cell(row=row_count, column=i).value:
                        cell.fill = fill_yellow
                        work_sheet.cell(row=row_count, column=i).fill = fill_yellow
                cell.number_format = '@'
                cell.font = font_for_write
            i += 1
    elif style_set == 'title':
        for item in row_context:
            cell = work_sheet.cell(row=row_count + 1, column=i, value=item)
            if item is not None:
                cell.fill = fill
                cell.border = border
                cell.number_format = '@'
                cell.font = font_for_write
            i += 1
    elif style_set == 'red':
        for item in row_context:
            cell = work_sheet.cell(row=row_count + 1, column=i, value=item)
            if item is not None:
                cell.font = red_bold_font
                cell.number_format = '@'
            i += 1
    elif style_set == 'bold':
        for item in row_context:
            cell = work_sheet.cell(row=row_count + 1, column=i, value=item)
            if item is not None:
                cell.font = font
                cell.number_format = '@'
            i += 1
    else:
        for item in row_context:
            cell = work_sheet.cell(row=row_count + 1, column=i, value=item)
            if item is not None:
                # pass
                cell.font = font_for_write
            i += 1


def set_column_autowidth(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)
        if column_letter == 'A' or column_letter == 'B':
            continue
        for cell in column_cells:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        adjusted_width = (max_length + 4)
        sheet.column_dimensions[column_letter].width = adjusted_width


def compare_csv(file1, file2):
    # 读取 CSV 文件为 DataFrame
    df1 = pd.read_csv(file1)
    df2 = pd.read_csv(file2)

    # 使用 equals 方法检查两个 DataFrame 是否完全一致
    return df1.equals(df2)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # self.input_group = QGroupBox("インプット")
        # self.input_layout = QVBoxLayout()
        # self.input_group.setLayout(self.input_layout)
        # self.hbox = QHBoxLayout()
        # self.button = QPushButton("TAB新規")
        # self.button = QPushButton("TAB削除")
        # self.button.clicked.connect(self.add_tab)
        # self.button.clicked.connect(self.removeTab)
        # self.hbox.addWidget(self.button)
        # self.hbox.addWidget(self.button)
        # self.input_layout.addLayout(self.hbox)
        self.tab_widget = QTabWidget()
        self.tab_1 = QWidget()
        self.tab_1.customContextMenuRequested.connect(lambda pos, tab=self.tab_1: self.openTabContextMenu(pos, tab))
        self.tab_widget.addTab(self.tab_1, f"実行前")
        self.tab_1.setDisabled(True)

        self.tab_1_layout = QVBoxLayout(self.tab_1)

        self.button_group_1 = QGroupBox(f"コントロール - 実行前")
        self.button_layout_1 = QVBoxLayout()
        self.button_layout_for_button = QHBoxLayout()
        self.button_layout_for_text = QHBoxLayout()
        self.button_layout_1.addLayout(self.button_layout_for_text)
        self.button_layout_1.addLayout(self.button_layout_for_button)
        self.button_group_1.setLayout(self.button_layout_1)
        self.button_search = QPushButton(f"検索")
        self.button_output = QPushButton(f"全データベース導出")
        self.button_input = QPushButton(f"全データベース導入")
        self.button_input.setDisabled(True)
        self.button_clear = QPushButton(f"全データベースクリア")
        self.button_clear.setDisabled(True)
        self.text_edit = QTextEdit()
        self.text_edit.setFixedHeight(5 * self.text_edit.fontMetrics().lineSpacing())
        self.text_edit.setLineWrapMode(QTextEdit.WidgetWidth)
        self.button_search.clicked.connect(lambda: self.button_clicked(self.button_search))
        self.button_output.clicked.connect(lambda: self.button_clicked(self.button_output))
        self.button_input.clicked.connect(lambda: self.button_clicked(self.button_input))
        self.button_clear.clicked.connect(lambda: self.button_clicked(self.button_clear))
        self.button_layout_for_button.addWidget(self.button_search)
        self.button_layout_for_button.addWidget(self.button_output)
        self.button_layout_for_button.addWidget(self.button_input)
        self.button_layout_for_button.addWidget(self.button_clear)
        self.button_layout_for_text.addWidget(self.text_edit)
        self.tab_1_layout.addWidget(self.button_group_1, 2)

        self.tab_2 = QWidget()
        self.tab_2.customContextMenuRequested.connect(lambda pos, tab=self.tab_2: self.openTabContextMenu(pos, tab))
        self.tab_widget.addTab(self.tab_2, f"実行後")
        self.tab_2.setDisabled(True)

        self.tab_2_layout = QVBoxLayout(self.tab_2)

        self.button_group_2 = QGroupBox(f"コントロール - 実行後")
        self.button_layout_2 = QVBoxLayout()
        self.button_layout_for_button = QHBoxLayout()
        self.button_layout_for_text = QHBoxLayout()
        self.button_layout_2.addLayout(self.button_layout_for_text)
        self.button_layout_2.addLayout(self.button_layout_for_button)
        self.button_group_2.setLayout(self.button_layout_2)
        self.button_search = QPushButton(f"検索")
        self.button_output = QPushButton(f"全データベース導出")
        self.button_input = QPushButton(f"全データベース導入")
        self.button_input.setDisabled(True)
        self.button_clear = QPushButton(f"全データベースクリア")
        self.button_clear.setDisabled(True)
        self.text_edit = QTextEdit()
        self.text_edit.setFixedHeight(5 * self.text_edit.fontMetrics().lineSpacing())
        self.text_edit.setLineWrapMode(QTextEdit.WidgetWidth)
        self.button_search.clicked.connect(lambda: self.button_clicked(self.button_search))
        self.button_output.clicked.connect(lambda: self.button_clicked(self.button_output))
        self.button_input.clicked.connect(lambda: self.button_clicked(self.button_input))
        self.button_clear.clicked.connect(lambda: self.button_clicked(self.button_clear))
        self.button_layout_for_button.addWidget(self.button_search)
        self.button_layout_for_button.addWidget(self.button_output)
        self.button_layout_for_button.addWidget(self.button_input)
        self.button_layout_for_button.addWidget(self.button_clear)
        self.button_layout_for_text.addWidget(self.text_edit)
        self.tab_2_layout.addWidget(self.button_group_2, 2)

        self.tab_a = QWidget()
        self.tab_a.customContextMenuRequested.connect(lambda pos, tab=self.tab_a: self.openTabContextMenu(pos, tab))
        self.tab_widget.addTab(self.tab_a, f"データベース導入導出")

        self.tab_a_layout = QVBoxLayout(self.tab_a)

        self.button_group_a = QGroupBox(f"コントロール - 導入導出")
        self.button_layout_a = QVBoxLayout()
        self.button_layout_for_button = QHBoxLayout()
        self.button_layout_for_text = QHBoxLayout()
        self.button_layout_a.addLayout(self.button_layout_for_text)
        self.button_layout_a.addLayout(self.button_layout_for_button)
        self.button_group_a.setLayout(self.button_layout_a)
        self.button_search = QPushButton(f"検索")
        self.button_output = QPushButton(f"全データベース導出")
        self.button_input = QPushButton(f"全データベース導入")
        # self.button_input.setDisabled(True)
        self.button_clear = QPushButton(f"全データベースクリア")
        # self.button_clear.setDisabled(True)
        self.text_edit = QTextEdit()
        self.text_edit.setFixedHeight(5 * self.text_edit.fontMetrics().lineSpacing())
        self.text_edit.setLineWrapMode(QTextEdit.WidgetWidth)
        self.button_search.clicked.connect(lambda: self.button_clicked(self.button_search))
        self.button_output.clicked.connect(lambda: self.button_clicked(self.button_output))
        self.button_input.clicked.connect(lambda: self.button_clicked(self.button_input))
        self.button_clear.clicked.connect(lambda: self.button_clicked(self.button_clear))
        self.button_layout_for_button.addWidget(self.button_search)
        self.button_layout_for_button.addWidget(self.button_output)
        self.button_layout_for_button.addWidget(self.button_input)
        self.button_layout_for_button.addWidget(self.button_clear)
        self.button_layout_for_text.addWidget(self.text_edit)
        self.tab_a_layout.addWidget(self.button_group_a, 2)

        self.tab_3 = QWidget()
        self.tab_3.customContextMenuRequested.connect(lambda pos, tab=self.tab_3: self.openTabContextMenu(pos, tab))
        self.tab_widget.addTab(self.tab_3, f"比較")

        self.tab_3_layout = QVBoxLayout(self.tab_3)

        self.button_group_3 = QGroupBox(f"コントロール - 比較")
        self.button_layout_3 = QVBoxLayout()
        self.button_layout_for_input_1 = QHBoxLayout()
        self.button_layout_for_input_2 = QHBoxLayout()
        self.button_layout_for_checkbox = QHBoxLayout()
        self.button_layout_for_button = QHBoxLayout()
        self.button_layout_3.addLayout(self.button_layout_for_input_1)
        self.button_layout_3.addLayout(self.button_layout_for_input_2)
        self.button_layout_3.addLayout(self.button_layout_for_checkbox)
        self.button_layout_3.addLayout(self.button_layout_for_button)
        self.button_group_3.setLayout(self.button_layout_3)

        self.label_before = QLabel('実行前CSV')
        self.input_before = QLineEdit()
        self.button_before = QPushButton('開く')
        self.button_before.setDisabled(True)
        self.label_after = QLabel('実行後CSV')
        self.input_after = QLineEdit()
        self.button_after = QPushButton('開く')
        self.button_after.setDisabled(True)
        self.label_checkbox_before = QLabel('データ準備')
        self.label_checkbox_before.setAlignment(Qt.AlignRight)
        self.checkbox_before = QCheckBox()
        self.label_checkbox_after = QLabel('データ比較')
        self.label_checkbox_after.setAlignment(Qt.AlignRight)
        self.checkbox_after = QCheckBox()
        self.button_compare = QPushButton(f"比較")
        self.button_compare_export = QPushButton(f"結果導出")
        self.button_layout_for_input_1.addWidget(self.label_before)
        self.button_layout_for_input_1.addWidget(self.input_before)
        self.button_layout_for_input_1.addWidget(self.button_before)
        self.button_layout_for_input_2.addWidget(self.label_after)
        self.button_layout_for_input_2.addWidget(self.input_after)
        self.button_layout_for_input_2.addWidget(self.button_after)
        self.button_layout_for_checkbox.addWidget(self.label_checkbox_before)
        self.button_layout_for_checkbox.addWidget(self.checkbox_before)
        self.button_layout_for_checkbox.addWidget(self.label_checkbox_after)
        self.button_layout_for_checkbox.addWidget(self.checkbox_after)
        self.button_compare.clicked.connect(lambda: self.button_clicked(self.button_compare))
        self.button_compare_export.clicked.connect(lambda: self.button_clicked(self.button_compare_export))
        self.button_layout_for_button.addWidget(self.button_compare)
        self.button_layout_for_button.addWidget(self.button_compare_export)
        self.tab_3_layout.addWidget(self.button_group_3, 2)

        self.tab_b = QWidget()
        self.tab_b.customContextMenuRequested.connect(lambda pos, tab=self.tab_b: self.openTabContextMenu(pos, tab))
        self.tab_widget.addTab(self.tab_b, f"比較 - 新")

        self.tab_b_layout = QVBoxLayout(self.tab_b)

        self.button_group_b = QGroupBox(f"コントロール - 比較 - 新")
        self.button_layout_b = QVBoxLayout()
        self.button_layout_for_input_1 = QHBoxLayout()
        self.button_layout_for_input_2 = QHBoxLayout()
        self.button_layout_for_button = QHBoxLayout()
        self.button_layout_b.addLayout(self.button_layout_for_input_1)
        self.button_layout_b.addLayout(self.button_layout_for_input_2)
        self.button_layout_b.addLayout(self.button_layout_for_button)
        self.button_group_b.setLayout(self.button_layout_b)

        self.label_before_new = QLabel('現CSV')
        self.input_before_new = QLineEdit()
        self.button_before_new = QPushButton('開く')
        self.button_before_new.setDisabled(True)
        self.label_after_new = QLabel('新CSV')
        self.input_after_new = QLineEdit()
        self.button_after_new = QPushButton('開く')
        self.button_after_new.setDisabled(True)
        self.button_compare_new = QPushButton(f"比較 - 新")
        self.button_compare_export_new = QPushButton(f"結果導出 - 新")
        self.button_layout_for_input_1.addWidget(self.label_before_new)
        self.button_layout_for_input_1.addWidget(self.input_before_new)
        self.button_layout_for_input_1.addWidget(self.button_before_new)
        self.button_layout_for_input_2.addWidget(self.label_after_new)
        self.button_layout_for_input_2.addWidget(self.input_after_new)
        self.button_layout_for_input_2.addWidget(self.button_after_new)
        self.button_layout_for_checkbox.addWidget(self.label_checkbox_before)
        self.button_layout_for_checkbox.addWidget(self.checkbox_before)
        self.button_layout_for_checkbox.addWidget(self.label_checkbox_after)
        self.button_layout_for_checkbox.addWidget(self.checkbox_after)
        self.button_compare_new.clicked.connect(lambda: self.button_clicked(self.button_compare_new))
        self.button_compare_export_new.clicked.connect(lambda: self.button_clicked(self.button_compare_export_new))
        self.button_layout_for_button.addWidget(self.button_compare_new)
        self.button_layout_for_button.addWidget(self.button_compare_export_new)
        self.tab_b_layout.addWidget(self.button_group_b, 2)

        self.tab_4 = QWidget()
        self.tab_4.customContextMenuRequested.connect(lambda pos, tab=self.tab_4: self.openTabContextMenu(pos, tab))
        self.tab_widget.addTab(self.tab_4, f"データインポート")
        # self.tab_4.setDisabled(True)

        self.tab_4_layout = QVBoxLayout(self.tab_4)

        self.button_group_4 = QGroupBox(f"コントロール - インポート")
        self.button_layout_3 = QVBoxLayout()
        self.button_layout_for_input_1 = QHBoxLayout()
        self.button_layout_for_input_2 = QHBoxLayout()
        self.button_layout_for_button = QHBoxLayout()
        self.button_layout_3.addLayout(self.button_layout_for_input_1)
        self.button_layout_3.addLayout(self.button_layout_for_input_2)
        self.button_layout_3.addLayout(self.button_layout_for_button)
        self.button_group_4.setLayout(self.button_layout_3)

        self.label_evidence = QLabel('エビデンス')
        self.input_evidence = QLineEdit()
        self.button_evidence = QPushButton('開く')
        self.button_evidence.setDisabled(True)
        self.button_parse = QPushButton(f"解析")
        self.button_import = QPushButton(f"インポート")
        self.button_layout_for_input_1.addWidget(self.label_evidence)
        self.button_layout_for_input_1.addWidget(self.input_evidence)
        self.button_layout_for_input_1.addWidget(self.button_evidence)
        self.button_parse.clicked.connect(lambda: self.button_clicked(self.button_parse))
        self.button_import.clicked.connect(lambda: self.button_clicked(self.button_import))
        self.button_layout_for_button.addWidget(self.button_parse)
        self.button_layout_for_button.addWidget(self.button_import)
        self.tab_4_layout.addWidget(self.button_group_4, 2)

        self.tab_widget.setCurrentIndex(2)

        # Add table group box
        self.table_group = QGroupBox(f"結果")
        self.table_layout = QVBoxLayout()
        self.table_group.setLayout(self.table_layout)
        self.table = QTableWidget()
        # self.table2.setColumnCount(3)
        # self.table2.setHorizontalHeaderLabels(['Column A', 'Column B', 'Column C'])
        self.table_layout.addWidget(self.table)
        self.table_group.setLayout(self.table_layout)

        # 创建输入组的QGroupBox
        self.button_group = QGroupBox("操作")
        self.button_layout = QHBoxLayout()
        self.button_group.setLayout(self.button_layout)
        self.switch_system = QComboBox()
        self.switch_system.addItems(ENABILITY_SYSTEM)
        # self.switch_system.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.export_button = QPushButton('データベースCSV導出')
        self.export_button.setDisabled(True)
        self.exec_button = QPushButton('EXCEL導出')
        self.exec_button.setDisabled(True)
        self.save_button = QPushButton('CSV導出')
        self.save_button.setDisabled(True)
        self.exit_button = QPushButton('退出')
        self.button_layout.addWidget(self.switch_system)
        self.button_layout.addWidget(self.export_button)
        self.button_layout.addWidget(self.exec_button)
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.exit_button)
        self.button_group.setLayout(self.button_layout)

        self.tips_group = QGroupBox("状態")
        self.tips_layout = QVBoxLayout()
        self.tips_layout_1 = QHBoxLayout()
        self.tips_layout_2 = QHBoxLayout()
        self.status_label = BlinkingLabel('データベース')
        self.tips_label = QLabel('')
        self.progress_bar = QProgressBar()
        self.status_label.start_blinking()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.tips_layout_1.addWidget(self.status_label)
        self.tips_layout_2.addWidget(self.tips_label)
        self.tips_layout_2.addWidget(self.progress_bar)
        self.tips_layout.addLayout(self.tips_layout_1)
        self.tips_layout.addLayout(self.tips_layout_2)
        self.tips_group.setLayout(self.tips_layout)

        self.setWindowTitle(f"BIP-データベースツール-Ver.1.0-Powered by PyQt5 - 「{WSL_NAME}」「{CURRENT_DATABASE}」")
        self.setGeometry(500, 300, 800, 600)

        self.main_widget = QWidget(self)
        self.setCentralWidget(self.main_widget)
        self.main_layout = QVBoxLayout()
        # self.main_layout.addWidget(self.input_group)
        self.main_layout.addWidget(self.tab_widget, 1)
        self.main_layout.addWidget(self.table_group, 5)
        self.main_layout.addWidget(self.button_group, 1)
        self.main_layout.addWidget(self.tips_group, 1)
        self.main_widget.setLayout(self.main_layout)

        self.timer_init()
        self.event_handler = EventHandler(self)
        self.init_ui()
        # init_config_content()
        self.tab_counter = 0

    def init_ui(self):
        """init_ui"""
        self.exit_button.clicked.connect(self.event_handler.app_exit)
        self.export_button.clicked.connect(self.event_handler.whole_database_csv_export)
        self.switch_system.activated.connect(self.event_handler.handleComboBoxClick)

        # self.button1.clicked.connect(self.vevent_handler.button1_click)
        # self.button2.clicked.connect(self.event_handler.app_exit)

    # def add_tab(self):
    #     self.tab_counter += 1
    #
    #     tab2 = QWidget()
    #     tab2.customContextMenuRequested.connect(lambda pos, tab=tab2: self.openTabContextMenu(pos, tab))
    #     self.tab_widget.addTab(tab2, f"Tab - {self.tab_counter}")
    #
    #     # Layout for tab2
    #     tab2_layout = QVBoxLayout(tab2)
    #
    #     # Add button group box
    #     button_group_box2 = QGroupBox(f"Button - {self.tab_counter}")
    #     button_layout = QVBoxLayout()
    #     button_layout2 = QHBoxLayout()
    #     button_layout3 = QHBoxLayout()
    #     button_layout.addLayout(button_layout3)
    #     button_layout.addLayout(button_layout2)
    #     button_group_box2.setLayout(button_layout)
    #     button3 = QPushButton(f"検索")
    #     button4 = QPushButton(f"全データベース導出")
    #     button5 = QPushButton(f"全データベース導入")
    #     button6 = QPushButton(f"全データベースクリア")
    #     text_edit = QTextEdit()
    #     text_edit.setFixedHeight(15 * text_edit.fontMetrics().lineSpacing())
    #     text_edit.setLineWrapMode(QTextEdit.WidgetWidth)
    #     button3.clicked.connect(lambda: self.button_clicked(button3))
    #     button4.clicked.connect(lambda: self.button_clicked(button4))
    #     button5.clicked.connect(lambda: self.button_clicked(button5))
    #     button6.clicked.connect(lambda: self.button_clicked(button6))
    #     button_layout2.addWidget(button3)
    #     button_layout2.addWidget(button4)
    #     button_layout2.addWidget(button5)
    #     button_layout2.addWidget(button6)
    #     button_layout3.addWidget(text_edit)
    #     tab2_layout.addWidget(button_group_box2, 2)
    #
    #     # Add table group box
    #     table_group_box2 = QGroupBox(f"Table - {self.tab_counter}")
    #     table_layout2 = QVBoxLayout()
    #     table_group_box2.setLayout(table_layout2)
    #     table2 = QTableWidget()
    #     table2.setColumnCount(3)
    #     table2.setHorizontalHeaderLabels(['Column A', 'Column B', 'Column C'])
    #     table2.setHorizontalHeaderLabels(['Column A', 'Column B', 'Column C'])
    #     table_layout2.addWidget(table2)
    #     tab2_layout.addWidget(table_group_box2, 5)
    #
    #     self.tab_widget.setCurrentIndex(self.tab_counter - 1)

    def button_clicked(self, button):
        global USER, PASSWORD, HOST, PORT, DATABASE
        global CIS_USER, CIS_PASSWORD, CIS_HOST, CIS_PORT, CIS_DATABASE
        global ORDER_USER, ORDER_PASSWORD, ORDER_HOST, ORDER_PORT, ORDER_DATABASE
        global PORTAL_USER, PORTAL_PASSWORD, PORTAL_HOST, PORTAL_PORT, PORTAL_DATABASE
        global TARGET_FOLDER_IMPORT, TARGET_FOLDER_EXPORT

        print(f"{button.text()} clicked")

        if button.text() == '結果導出':
            if self.table.rowCount() == 0:
                set_message_box("WARNING", "結果導出", f"結果はまだ生成しないで、比較を実行してください。")
                return
            if self.checkbox_before.isChecked() is False and self.checkbox_after.isChecked() is False:
                set_message_box("WARNING", "結果導出", f"「データ準備」または「データ比較」を選んでください。")
                return
            if self.checkbox_before.isChecked():
                tmp_file = os.path.join(get_program_path(), "データ準備" +
                                        str(datetime.datetime.now()).
                                        replace('-', '').replace('.', '').
                                        replace(' ', '').replace('-', '').
                                        replace(':', '') + ".xlsx")
                table_data = []
                for row in range(self.table.rowCount()):
                    row_data = []
                    for column in range(self.table.columnCount()):
                        item = self.table.item(row, column)
                        if item == '更新後':
                            item = '変更'
                        if item is not None:
                            if item.text() == '更新後':
                                row_data.append("変更")
                            else:
                                row_data.append(item.text())
                        else:
                            row_data.append(None)
                    if row_data[len(row_data) - 2] == "更新前":
                        continue

                    table_data.append(row_data)

                tables = get_tables_names_from_db()

                if os.path.exists(tmp_file):
                    wb = load_workbook(tmp_file)
                    ws = wb["データ準備"]
                else:
                    wb = openWorkbook()
                    ws = wb.create_sheet(title='データ準備')

                result_for_excel = []
                write_row_count = 1
                write_to_excel(ws, 0, ["※テスト用DBダンプに基づいて、データ準備を作成する。"], "red", 1)
                for index, row in enumerate(table_data):
                    if row[0] == "比較結果":
                        continue
                    if find_tables(row[0], tables) is True and row[0] != table_data[index - 1][0]:
                        write_to_excel(ws, write_row_count, [None, None, None], "none", 1)
                        write_row_count += 1

                        write_to_excel(ws, write_row_count, [None, None,
                                                             "".join(get_table_name_from_db(row[0], "table")[0])],
                                       "none", 1)
                        write_row_count += 1

                        write_to_excel(ws, write_row_count, [None, None, row[0]], "none", 1)
                        write_row_count += 1

                        columns = get_table_name_from_db(row[0], "column")
                        column_names = []
                        for column in columns:
                            column_names.append("".join(column))
                        write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                        write_row_count += 1

                        columns_new = get_columns_from_db(row[0])
                        column_new_names = []
                        for column in columns_new:
                            column_new_names.append("".join(column))
                        write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                        write_row_count += 1
                        str_to_list = ast.literal_eval(row[1])
                        formatted_list = [parse_date(item) for item in str_to_list]
                        write_to_excel(ws, write_row_count,
                                       [None, row[2]] + formatted_list, "border", 1)
                        write_row_count += 1
                    if find_tables(row[0], tables) is True and row[0] == table_data[index - 1][0]:
                        str_to_list = ast.literal_eval(row[1])
                        formatted_list = [parse_date(item) for item in str_to_list]
                        write_to_excel(ws, write_row_count, [None, row[2]] + formatted_list, "border", 1)
                        write_row_count += 1

                if 'Sheet' in wb.sheetnames:
                    ws_for_remove = wb['Sheet']
                    wb.remove(ws_for_remove)

                font_for_write = Font(name='MS Gothic')
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            cell.font = font_for_write

                set_column_autowidth(ws)

                wb.save(tmp_file)
                wb.close()
                set_message_box("WARNING", "データベース", f"比較結果生成完了しました。\n{tmp_file}")

            if self.checkbox_after.isChecked():
                tmp_file = os.path.join(get_program_path(), "データ比較" +
                                        str(datetime.datetime.now()).
                                        replace('-', '').replace('.', '').
                                        replace(' ', '').replace('-', '').
                                        replace(':', '') + ".xlsx")
                table_data = []
                for row in range(self.table.rowCount()):
                    row_data = []
                    for column in range(self.table.columnCount()):
                        item = self.table.item(row, column)
                        if item is not None:
                            row_data.append(item.text())
                        else:
                            row_data.append(None)
                    table_data.append(row_data)
                if os.path.exists(tmp_file):
                    wb = load_workbook(tmp_file)
                    ws = wb["データ比較"]
                else:
                    wb = openWorkbook()
                    ws = wb.create_sheet(title='データ比較')

                result_for_excel = []
                # 初始化一个字典来存储每个表的第二个字段内容
                table_field2_contents = {}

                # 遍历每个条目，按照表名和操作类型进行集计
                for item in table_data:
                    if item[0] == "比較結果":
                        continue
                    table_name = item[0]
                    operation_type = item[-2]  # 倒数第二个字段，即操作类型
                    field2_content = item[1]  # 第二个字段的具体内容

                    # 如果表名不在字典中，则初始化为一个字典，用来存储操作类型和内容的对应关系
                    if table_name not in table_field2_contents:
                        table_field2_contents[table_name] = {}

                    # 如果操作类型不在表名对应的字典中，则初始化为一个集合来存储内容
                    if operation_type not in table_field2_contents[table_name]:
                        table_field2_contents[table_name][operation_type] = set()

                    # 添加第二个字段的内容到集合中
                    table_field2_contents[table_name][operation_type].add(field2_content)

                write_row_count = 0
                # 打印结果
                write_to_excel(ws, write_row_count, ["[自動採番]、[登録/更新/削除日時]、"
                                                     "[登録/更新/削除者]、[登録/更新/削除機能]"
                                                     "のデータ比較結果が「FALSE」の場合、補足説明が必要がない。"], "red", 1)
                write_row_count += 1
                for table, operations in table_field2_contents.items():
                    write_to_excel(ws, write_row_count, [], "none", 1)
                    write_row_count += 1
                    write_to_excel(ws, write_row_count, [], "none", 1)
                    write_row_count += 1
                    print(f"表名: {table}")
                    table_name = "".join(get_table_name_from_db(table, "table")[0])
                    columns = get_table_name_from_db(table, "column")
                    column_names = []
                    for column in columns:
                        column_names.append("".join(column))
                    columns_new = get_columns_from_db(table)
                    column_new_names = []
                    for column in columns_new:
                        column_new_names.append("".join(column))
                    for operation, contents in operations.items():
                        print(f"操作类型: {operation}")
                        parsed_data = [eval(item) for item in contents]
                        sorted_data = sorted(parsed_data, key=lambda x: safe_int(x[0]))
                        if operation == "新規":
                            compare_row = []
                            add_before_row_1 = []
                            add_after_row_1 = []
                            add_before_row_2 = []
                            add_after_row_2 = []
                            write_to_excel(ws, write_row_count, [None, "現行システム"], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "新規前", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for i in range(len(sorted_data)):
                                write_to_excel(ws, write_row_count, [], "none", 1)
                                add_before_row_1.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "新規後", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for content in sorted_data:
                                # list_data = content.strip('[]').split(', ')
                                # list_data = [item.strip("'") for item in list_data]
                                formatted_list = [parse_date(item) for item in content]
                                write_to_excel(ws, write_row_count, [None, None] + formatted_list, "border", 1)
                                add_after_row_1.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "新システム"], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "新規前", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for i in range(len(sorted_data)):
                                write_to_excel(ws, write_row_count, [], "none", 1)
                                add_before_row_2.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "新規後", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for content in sorted_data:
                                # list_data = content.strip('[]').split(', ')
                                # list_data = [item.strip("'") for item in list_data]
                                write_content = []
                                for i in range(len(content)):
                                    write_content.append("データ")
                                write_to_excel(ws, write_row_count, [None, None] + write_content, "border", 1)
                                add_after_row_2.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "比較結果"], "bold", 1)
                            write_row_count += 1

                            for i in range(len(contents)):
                                write_kansuu = []
                                for y in range(len(column_names)):
                                    write_kansuu.append(f"=EXACT({number_to_excel_column(y + 3)}"
                                                        f"{str(int(add_after_row_1[i]) + 1)},"
                                                        f"{number_to_excel_column(y + 3)}"
                                                        f"{str(int(add_after_row_2[i]) + 1)})")

                                write_to_excel(ws, write_row_count, [None, "新規後"] + write_kansuu, "none", 1)
                                write_row_count += 1

                        if operation == "削除":
                            del_before_row_1 = []
                            del_after_row_1 = []
                            del_before_row_2 = []
                            del_after_row_2 = []
                            write_to_excel(ws, write_row_count, [None, "現行システム"], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "削除前", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for content in sorted_data:
                                # list_data = content.strip('[]').split(', ')
                                # list_data = [item.strip("'") for item in list_data]
                                formatted_list = [parse_date(item) for item in content]
                                write_to_excel(ws, write_row_count, [None, None] + formatted_list, "border", 1)
                                del_before_row_1.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "削除後", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for i in range(len(sorted_data)):
                                write_to_excel(ws, write_row_count, [], "none", 1)
                                del_after_row_1.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "新システム"], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "削除前", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for content in sorted_data:
                                # list_data = content.strip('[]').split(', ')
                                # list_data = [item.strip("'") for item in list_data]
                                write_content = []
                                for i in range(len(content)):
                                    write_content.append("データ")
                                write_to_excel(ws, write_row_count, [None, None] + write_content, "border", 1)
                                del_before_row_2.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "削除後", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for i in range(len(sorted_data)):
                                write_to_excel(ws, write_row_count, [], "none", 1)
                                del_after_row_2.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "比較結果"], "bold", 1)
                            write_row_count += 1

                            for i in range(len(contents)):
                                write_kansuu = []
                                for y in range(len(column_names)):
                                    write_kansuu.append(f"=EXACT({number_to_excel_column(y + 3)}"
                                                        f"{str(int(del_before_row_1[i]) + 1)},"
                                                        f"{number_to_excel_column(y + 3)}"
                                                        f"{str(int(del_before_row_2[i]) + 1)})")

                                write_to_excel(ws, write_row_count, [None, "削除前"] + write_kansuu, "none", 1)
                                write_row_count += 1

                        upd_before_row_1 = []
                        upd_before_row_2 = []
                        upd_after_row_1 = []
                        upd_after_row_2 = []
                        if operation == "更新前":
                            write_to_excel(ws, write_row_count, [None, "現行システム"], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "更新前", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for content in sorted_data:
                                # list_data = content.strip('[]').split(', ')
                                # list_data = [item.strip("'") for item in list_data]
                                formatted_list = [parse_date(item) for item in content]
                                write_to_excel(ws, write_row_count, [None, None] + formatted_list, "border", 1)
                                upd_before_row_1.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "新システム"], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "更新前", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for content in sorted_data:
                                # list_data = content.strip('[]').split(', ')
                                # list_data = [item.strip("'") for item in list_data]
                                write_content = []
                                for i in range(len(content)):
                                    write_content.append("データ")
                                write_to_excel(ws, write_row_count, [None, None] + write_content, "border", 1)
                                upd_before_row_2.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "比較結果"], "bold", 1)
                            write_row_count += 1
                            for i in range(len(contents)):
                                write_kansuu = []
                                for y in range(len(column_names)):
                                    write_kansuu.append(f"=EXACT({number_to_excel_column(y + 3)}"
                                                        f"{str(int(upd_before_row_1[i]) + 1)},"
                                                        f"{number_to_excel_column(y + 3)}"
                                                        f"{str(int(upd_before_row_2[i]) + 1)})")

                                write_to_excel(ws, write_row_count, [None, "更新前"] + write_kansuu, "none", 1)
                                write_row_count += 1

                        if operation == "更新後":
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "現行システム"], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "更新後", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for content in sorted_data:
                                # list_data = content.strip('[]').split(', ')
                                # list_data = [item.strip("'") for item in list_data]
                                formatted_list = [parse_date(item) for item in content]
                                write_to_excel(ws, write_row_count, [None, None] + formatted_list, "border", 1)
                                upd_after_row_1.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "新システム"], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "更新後", table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            for content in sorted_data:
                                # list_data = content.strip('[]').split(', ')
                                # list_data = [item.strip("'") for item in list_data]
                                write_content = []
                                for i in range(len(content)):
                                    write_content.append("データ")
                                write_to_excel(ws, write_row_count, [None, None] + write_content, "border", 1)
                                upd_after_row_2.append(int(write_row_count))
                                write_row_count += 1

                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "比較結果"], "bold", 1)
                            write_row_count += 1
                            for i in range(len(contents)):
                                write_kansuu = []
                                for y in range(len(column_names)):
                                    write_kansuu.append(f"=EXACT({number_to_excel_column(y + 3)}"
                                                        f"{str(int(upd_after_row_1[i]) + 1)},"
                                                        f"{number_to_excel_column(y + 3)}"
                                                        f"{str(int(upd_after_row_2[i]) + 1)})")

                                write_to_excel(ws, write_row_count, [None, "更新後"] + write_kansuu, "none", 1)
                                write_row_count += 1

                        if operation != "更新前":
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1

                        for content in sorted_data:
                            print(f"第二个字段内容: {content}")
                        print()
                    print()

                # for item in result_for_excel:
                #     ws.append(item)

                set_column_autowidth(ws)

                if 'Sheet' in wb.sheetnames:
                    ws_for_remove = wb['Sheet']
                    wb.remove(ws_for_remove)

                font_for_write = Font(name='MS Gothic')
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            cell.font = font_for_write

                wb.save(tmp_file)
                wb.close()
                set_message_box("WARNING", "データベース", f"比較結果生成完了しました。\n{tmp_file}")

        if button.text() == '結果導出 - 新':
            if self.table.rowCount() == 0:
                set_message_box("WARNING", "結果導出", f"結果はまだ生成しないで、比較を実行してください。")
                return

            tmp_file = os.path.join(get_program_path(), "データ比較" +
                                    str(datetime.datetime.now()).
                                    replace('-', '').replace('.', '').
                                    replace(' ', '').replace('-', '').
                                    replace(':', '') + ".xlsx")
            table_data = []
            for row in range(self.table.rowCount()):
                row_data = []
                for column in range(self.table.columnCount()):
                    item = self.table.item(row, column)
                    if item is not None:
                        row_data.append(item.text())
                    else:
                        row_data.append(None)
                table_data.append(row_data)
            if os.path.exists(tmp_file):
                wb = load_workbook(tmp_file)
                ws = wb["テストデータ比較"]
            else:
                wb = openWorkbook()
                ws = wb.create_sheet(title='テストデータ比較')

            result_for_excel = []
            # 初始化一个字典来存储每个表的第二个字段内容
            table_field2_contents = {}

            # 遍历每个条目，按照表名和操作类型进行集计
            for item in table_data:
                if item[0] == "比較結果" or item[0] == "ファイル差異" \
                        or item[0] is None or item[0] == "":
                    continue
                table_name = item[0]
                operation_type = item[-2]  # 倒数第二个字段，即操作类型
                field2_content = item[1]  # 第二个字段的具体内容

                # 如果表名不在字典中，则初始化为一个字典，用来存储操作类型和内容的对应关系
                if table_name not in table_field2_contents:
                    table_field2_contents[table_name] = {}

                # 如果操作类型不在表名对应的字典中，则初始化为一个集合来存储内容
                if operation_type not in table_field2_contents[table_name]:
                    table_field2_contents[table_name][operation_type] = set()

                # 添加第二个字段的内容到集合中
                table_field2_contents[table_name][operation_type].add(field2_content)

            write_row_count = 0
            # 打印结果
            write_to_excel(ws, write_row_count, ["テストデータ参照用"], "red", 1)
            write_row_count += 1
            for table, operations in table_field2_contents.items():
                title_flag = False
                upd_before_list = []
                upd_after_list = []
                write_to_excel(ws, write_row_count, [], "none", 1)
                write_row_count += 1
                write_to_excel(ws, write_row_count, [], "none", 1)
                write_row_count += 1
                print(f"表名: {table}")
                table_name = "".join(get_table_name_from_db(table, "table")[0])
                columns = get_table_name_from_db(table, "column")
                column_names = []
                for column in columns:
                    column_names.append("".join(column))
                columns_new = get_columns_from_db(table)
                column_new_names = []
                for column in columns_new:
                    column_new_names.append("".join(column))
                for operation, contents in operations.items():
                    print(f"操作类型: {operation}")
                    parsed_data = [eval(item) for item in contents]
                    sorted_data = sorted(parsed_data, key=lambda x: safe_int(x[0]))
                    if operation == "新規":
                        if title_flag is False:
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            title_flag = True

                        for content in sorted_data:
                            write_to_excel(ws, write_row_count, ["", "現"], "none", 1)
                            write_row_count += 1
                            formatted_list = [parse_date(item) for item in content]
                            write_to_excel(ws, write_row_count, [None, "新"] + formatted_list, "border", 1)
                            write_row_count += 1
                        write_to_excel(ws, write_row_count, [], "None", 1)
                        write_row_count += 1

                    if operation == "削除":
                        if title_flag is False:
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            title_flag = True

                        for content in sorted_data:
                            formatted_list = [parse_date(item) for item in content]
                            write_to_excel(ws, write_row_count, [None, "現"] + formatted_list, "border", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, "新"], "none", 1)
                            write_row_count += 1
                        write_to_excel(ws, write_row_count, [], "None", 1)
                        write_row_count += 1

                    if operation == "更新前":
                        for content in sorted_data:
                            formatted_list = [parse_date(item) for item in content]
                            upd_before_list.append([None, "現"] + formatted_list)

                    if operation == "更新後":
                        for content in sorted_data:
                            formatted_list = [parse_date(item) for item in content]
                            upd_after_list.append([None, "新"] + formatted_list)

                    if len(upd_before_list) > 0 and len(upd_after_list) > 0:
                        if title_flag is False:
                            write_to_excel(ws, write_row_count, [], "none", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table_name], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None, table], "bold", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_names, "title", 1)
                            write_row_count += 1
                            write_to_excel(ws, write_row_count, [None, None] + column_new_names, "title", 1)
                            write_row_count += 1
                            title_flag = True

                        all_list = sorted(upd_before_list + upd_after_list, key=lambda x: (x[0], x[2]))

                        write_to_excel(ws, write_row_count, [], "None", 1)
                        write_row_count += 1
                        for upd_context in all_list:
                            if upd_context[1] == "現":
                                write_to_excel(ws, write_row_count, upd_context, "border", 1)
                            if upd_context[1] == "新":
                                write_to_excel(ws, write_row_count, upd_context, "border-yellow", 1)
                            write_row_count += 1

                    for content in sorted_data:
                        print(f"第二个字段内容: {content}")
                    print()
                print()

            set_column_autowidth(ws)

            if 'Sheet' in wb.sheetnames:
                ws_for_remove = wb['Sheet']
                wb.remove(ws_for_remove)

            font_for_write = Font(name='MS Gothic')
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        cell.font = font_for_write

            wb.save(tmp_file)
            wb.close()
            set_message_box("WARNING", "データベース", f"比較結果生成完了しました。\n{tmp_file}")

        if button.text() == '解析':
            if self.input_evidence.text() is None or \
                    self.input_evidence.text() == "" or \
                    os.path.exists(self.input_evidence.text()) is False:
                set_message_box("CRITICAL", "ファイル", "エビデンスはありませんので、チェックしてください。")
                return
            if CURRENT_DATABASE == "EnabilityCIS":
                USER = CIS_USER
                PASSWORD = CIS_PASSWORD
                HOST = CIS_HOST
                PORT = CIS_PORT
                DATABASE = CIS_DATABASE
                schema = "unisys"
            if CURRENT_DATABASE == "EnabilityOrder":
                USER = ORDER_USER
                PASSWORD = ORDER_PASSWORD
                HOST = ORDER_HOST
                PORT = ORDER_PORT
                DATABASE = ORDER_DATABASE
                schema = "public"
            if CURRENT_DATABASE == "EnabilityPortal":
                USER = PORTAL_USER
                PASSWORD = PORTAL_PASSWORD
                HOST = PORTAL_HOST
                PORT = PORTAL_PORT
                DATABASE = PORTAL_DATABASE
                schema = "public"
            conn_params = {
                'user': USER,
                'password': PASSWORD,
                'host': HOST,
                'port': PORT,
                'database': DATABASE
            }

            # 连接到数据库
            connection = psycopg2.connect(**conn_params)
            cursor = connection.cursor()

            # 获取所有表名
            cursor.execute(f"""
                    SELECT table_name
                    FROM information_schema.tables
                    WHERE table_schema = '{schema}'
                    AND table_type = 'BASE TABLE';
                """)
            tables = cursor.fetchall()
            cursor.close()
            connection.close()

            self.clearAllTables()
            time.sleep(1)
            self.table.setColumnCount(4)
            self.table.setHorizontalHeaderLabels(["選択", "区分", "ファイル", "備考", "状態"])
            self.table.resizeColumnsToContents()
            self.table.resizeRowsToContents()
            row_data = parse_excel(self.input_evidence.text(), tables)
            for row in row_data:
                row_append(self.table, ["", row, "", ""], Qt.lightGray, True)

        if button.text() == 'インポート':
            # self.clearAllTables()

            for row in range(self.table.rowCount()):
                print(f'Row {row}, Column {0} value: {self.table.cellWidget(row, 0).findChild(QCheckBox).isChecked()}')
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item is not None:
                        print(f'Row {row}, Column {col} value: {item.text()}')

        if button.text() == '比較' or button.text() == '比較 - 新':
            if CURRENT_DATABASE == "EnabilityCIS":
                USER = CIS_USER
                PASSWORD = CIS_PASSWORD
                HOST = CIS_HOST
                PORT = CIS_PORT
                DATABASE = CIS_DATABASE
                schema = "unisys"
            if CURRENT_DATABASE == "EnabilityOrder":
                USER = ORDER_USER
                PASSWORD = ORDER_PASSWORD
                HOST = ORDER_HOST
                PORT = ORDER_PORT
                DATABASE = ORDER_DATABASE
                schema = "public"
            if CURRENT_DATABASE == "EnabilityPortal":
                USER = PORTAL_USER
                PASSWORD = PORTAL_PASSWORD
                HOST = PORTAL_HOST
                PORT = PORTAL_PORT
                DATABASE = PORTAL_DATABASE
                schema = "public"
            conn_params = {
                'user': USER,
                'password': PASSWORD,
                'host': HOST,
                'port': PORT,
                'database': DATABASE
            }
            self.clearAllTables()
            if button.text() == "比較":
                input_before_path = self.input_before.text()
                input_after_path = self.input_after.text()
                if os.path.exists(self.input_before.text()) is False:
                    set_message_box("WARNING", "比較", f"パス「{self.input_before.text()}」は存在しません。")
                    return
                if os.path.exists(self.input_after.text()) is False:
                    set_message_box("WARNING", "比較", f"パス「{self.input_after.text()}」は存在しません。")
                    return
            if button.text() == "比較 - 新":
                input_before_path = self.input_before_new.text()
                input_after_path = self.input_after_new.text()
                if os.path.exists(self.input_before_new.text()) is False:
                    set_message_box("WARNING", "比較 - 新", f"パス「{self.input_before_new.text()}」は存在しません。")
                    return
                if os.path.exists(self.input_after_new.text()) is False:
                    set_message_box("WARNING", "比較 - 新", f"パス「{self.input_after_new.text()}」は存在しません。")
                    return

            files1 = set(os.listdir(input_before_path))
            files2 = set(os.listdir(input_after_path))
            unique_to_folder1 = files1 - files2
            unique_to_folder2 = files2 - files1
            tab_table = self.table
            self.clearAllTables()
            time.sleep(1)
            tab_table.setColumnCount(4)
            tab_table.setHorizontalHeaderLabels(["区分", "ファイル", "備考", "状態"])
            tab_table.resizeColumnsToContents()
            tab_table.resizeRowsToContents()
            if len(unique_to_folder1) > 0 or len(unique_to_folder2) > 0:
                row_append(tab_table, ["ファイル差異", None, None, None], Qt.lightGray, False)
                if unique_to_folder1:
                    for file in unique_to_folder1:
                        row_append(tab_table, ["", file, "実行前だけで存在しています。比較しない決まります。", "✕"], Qt.lightGray, False)
                if unique_to_folder2:
                    for file in unique_to_folder2:
                        row_append(tab_table, ["", file, "実行後だけで存在しています。比較しない決まります。", "✕"], Qt.lightGray, False)
            row_append(tab_table, ["比較結果", None, None, None], Qt.lightGray, False)
            compare_result = []
            common_files = files1.intersection(files2)
            common_files = sorted(common_files)

            for file in common_files:
                print("file:", file)
                csv_file1 = os.path.join(input_before_path, file)
                csv_file2 = os.path.join(input_after_path, file)
                if compare_csv(csv_file1, csv_file2):
                    continue
                # 连接到数据库
                connection = psycopg2.connect(**conn_params)
                cursor = connection.cursor()
                table_name = get_str_before_first_dot(file, ".")
                cursor.execute(f"""
                    SELECT 
                        ordinal_position
                    FROM 
                        information_schema.key_column_usage AS kcu
                    WHERE 
                        constraint_name IN (
                            SELECT constraint_name 
                            FROM information_schema.table_constraints 
                            WHERE table_name = '{table_name}' AND constraint_type = 'PRIMARY KEY'
                        )
                    ORDER BY 
                        ordinal_position;
                """)
                # 获取所有查询结果的行
                rows = cursor.fetchall()

                list_of_indices = []
                # 输出查询结果
                for row in rows:
                    list_of_indices.append(int("".join(map(str, row))))
                print("indices:", list_of_indices)
                if len(list_of_indices) == 0:
                    list_of_indices = [1]

                cursor.close()
                connection.close()
                del_row, add_row, upd_row_before, upd_row_after = compare_csv_to_excel(csv_file1, csv_file2,
                                                                                       list_of_indices)
                print("删除操作：", del_row)
                print("增加操作：", add_row)
                print("修改操作：", upd_row_before)
                print("修改操作：", upd_row_after)
                if len(del_row) > 0:
                    # row_append(tab_table, [table_name, "データ削除", None, None], Qt.lightGray, False)
                    for row in del_row:
                        row_append(tab_table, [table_name, row, "削除", None], Qt.lightGray, False)
                if len(add_row) > 0:
                    # row_append(tab_table, [table_name, "データ新規", None, None], Qt.lightGray, False)
                    for row in add_row:
                        row_append(tab_table, [table_name, row, "新規", None], Qt.lightGray, False)
                if len(upd_row_before) > 0 and len(upd_row_after) > 0:
                    # row_append(tab_table, [table_name, "データ更新", None, None], Qt.lightGray, False)
                    for i, row in enumerate(upd_row_before):
                        row_append(tab_table, [table_name, row, "更新前", None], Qt.lightGray, False)
                        row_append(tab_table, [table_name, upd_row_after[i], "更新後", None], Qt.lightGray, False)
            #     if len(compare_data) > 0:
            #         for compare in compare_data:
            #             compare_result.append(
            #                 [file, compare[0].replace("A: ", "") + "\n" + compare[1].replace("B: ", "")])
            #     else:
            #         pass
            #         # row_append(tab_table, ["", file, "完全に同じ", ""], Qt.lightGray, False)
            #     print("※" * 10, file, "|", compare_data)
            # if len(compare_result) > 0:
            #     for compare in compare_result:
            #         row_append(tab_table, ["", compare[0], compare[1], ""], Qt.red, False)
            # else:
            #     row_append(tab_table, ["", "", "完全に同じ", ""], Qt.lightGray, False)
            set_message_box("WARNING", "データベース", f"比較完了しました。")

        if button.text() == '検索':
            self.clearAllTables()
            if CURRENT_DATABASE == "EnabilityCIS":
                USER = CIS_USER
                PASSWORD = CIS_PASSWORD
                HOST = CIS_HOST
                PORT = CIS_PORT
                DATABASE = CIS_DATABASE
                schema = "unisys"
            if CURRENT_DATABASE == "EnabilityOrder":
                USER = ORDER_USER
                PASSWORD = ORDER_PASSWORD
                HOST = ORDER_HOST
                PORT = ORDER_PORT
                DATABASE = ORDER_DATABASE
                schema = "public"
            if CURRENT_DATABASE == "EnabilityPortal":
                USER = PORTAL_USER
                PASSWORD = PORTAL_PASSWORD
                HOST = PORTAL_HOST
                PORT = PORTAL_PORT
                DATABASE = PORTAL_DATABASE
                schema = "public"
            try:
                connection = psycopg2.connect(user=USER,
                                              password=PASSWORD,
                                              host=HOST,
                                              port=PORT,
                                              database=DATABASE)

                cursor = connection.cursor()
                # self.progress_bar.setValue(0)
                text_info = self.tab_widget.currentWidget().findChild(QTextEdit)
                # 查询数据
                select_query = text_info.toPlainText()
                cursor.execute(select_query)
                total_records = cursor.fetchall()

                table_name = text_info.toPlainText().split("from")[1].split("where")[0] \
                    .replace(" ", "").replace(";", "")
                cursor.execute(f"""
                    SELECT a.attname AS column_name, 
                           pg_catalog.col_description(c.oid, a.attnum) AS column_comment,
                           format_type(a.atttypid, a.atttypmod) AS data_type,
                           a.attnotnull AS is_nullable
                    FROM pg_catalog.pg_attribute a
                    JOIN pg_catalog.pg_class c ON a.attrelid = c.oid
                    LEFT JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
                    WHERE c.relname = '{table_name}'
                      AND a.attnum > 0
                      AND NOT a.attisdropped
                """)
                columns = cursor.fetchall()
                col_title = []
                for column in columns:
                    col_title.append(column[1] + "\n" + column[0])

                tab_table = self.table
                self.clearAllTables()
                time.sleep(1)
                tab_table.setColumnCount(len(col_title))
                tab_table.setHorizontalHeaderLabels(col_title)
                tab_table.resizeColumnsToContents()
                tab_table.resizeRowsToContents()
                for content in total_records:
                    row_append(tab_table, content, Qt.lightGray, False)
                    # self.progress_bar.setValue(self.progress_bar.value() + int(len(total_records) / 100))

            except (Exception, Error) as error:
                print("Error while connecting to PostgreSQL", error)
                set_message_box("CRITICAL", "データベース", error.args[0])
            finally:
                if connection:
                    cursor.close()
                    connection.close()
                    print("PostgreSQL connection is closed")

        if button.text() == '全データベース導出':
            try:
                options = QFileDialog.Options()
                options |= QFileDialog.DontUseNativeDialog
                if TARGET_FOLDER_EXPORT is not None:
                    open_path = TARGET_FOLDER_EXPORT
                folder_path = QFileDialog.getExistingDirectory(self, "導出パス選択",
                                                               directory=open_path,
                                                               options=options)
                if folder_path:
                    print(folder_path)
                    TARGET_FOLDER_EXPORT = folder_path

                    if CURRENT_DATABASE == "EnabilityCIS":
                        USER = CIS_USER
                        PASSWORD = CIS_PASSWORD
                        HOST = CIS_HOST
                        PORT = CIS_PORT
                        DATABASE = CIS_DATABASE
                        schema = "unisys"
                    if CURRENT_DATABASE == "EnabilityOrder":
                        USER = ORDER_USER
                        PASSWORD = ORDER_PASSWORD
                        HOST = ORDER_HOST
                        PORT = ORDER_PORT
                        DATABASE = ORDER_DATABASE
                        schema = "public"
                    if CURRENT_DATABASE == "EnabilityPortal":
                        USER = PORTAL_USER
                        PASSWORD = PORTAL_PASSWORD
                        HOST = PORTAL_HOST
                        PORT = PORTAL_PORT
                        DATABASE = PORTAL_DATABASE
                        schema = "public"
                    conn_params = {
                        'user': USER,
                        'password': PASSWORD,
                        'host': HOST,
                        'port': PORT,
                        'database': DATABASE
                    }

                    # 连接到数据库
                    connection = psycopg2.connect(**conn_params)
                    cursor = connection.cursor()

                    # 获取所有表名
                    cursor.execute(f"""
                            SELECT table_name
                            FROM information_schema.tables
                            WHERE table_schema = '{schema}'
                            AND table_type = 'BASE TABLE';
                        """)
                    tables = cursor.fetchall()

                    # if self.tab_widget.currentIndex() == 0:
                    #     output_folder = "export_csv_before"
                    #     self.input_before.setText(os.path.join(get_program_path(), output_folder))
                    # if self.tab_widget.currentIndex() == 1:
                    #     output_folder = "export_csv_after"
                    #     self.input_after.setText(os.path.join(get_program_path(), output_folder))
                    # os.makedirs(os.path.join(get_program_path(), output_folder), exist_ok=True)
                    tab_table = self.table
                    self.clearAllTables()
                    time.sleep(1)
                    tab_table.setColumnCount(2)
                    tab_table.setHorizontalHeaderLabels(["CSVファイル", "導出状態"])
                    # 对每张表执行导出操作
                    for table in tables:
                        table_name = table[0]

                        # 构建查询语句，导出表数据
                        query = sql.SQL("COPY {} TO STDOUT WITH CSV HEADER").format(
                            sql.Identifier(table_name)
                        )

                        # 执行查询
                        csv_file_path = os.path.join(TARGET_FOLDER_EXPORT, f"{table_name}.csv")
                        with open(csv_file_path, 'w', newline='', encoding='utf-8') as f:
                            cursor.copy_expert(query, f)
                            tab_table.resizeColumnsToContents()
                            tab_table.resizeRowsToContents()
                            row_append(self.table, [csv_file_path, "〇"], Qt.lightGray, False)

                    set_message_box("WARNING", "データベース", "導出成功")
                    cursor.close()
                    connection.close()
            except Exception as e:
                print("An error occurred : ", e)
                raise

        if button.text() == '全データベース導入':
            try:
                options = QFileDialog.Options()
                options |= QFileDialog.DontUseNativeDialog
                if TARGET_FOLDER_IMPORT is not None:
                    open_path = TARGET_FOLDER_IMPORT
                folder_path = QFileDialog.getExistingDirectory(self, "導入パス選択",
                                                               directory=open_path,
                                                               options=options)
                if folder_path:
                    print(folder_path)
                    TARGET_FOLDER_IMPORT = folder_path

                    if CURRENT_DATABASE == "EnabilityCIS":
                        USER = CIS_USER
                        PASSWORD = CIS_PASSWORD
                        HOST = CIS_HOST
                        PORT = CIS_PORT
                        DATABASE = CIS_DATABASE
                        schema = "unisys"
                    if CURRENT_DATABASE == "EnabilityOrder":
                        USER = ORDER_USER
                        PASSWORD = ORDER_PASSWORD
                        HOST = ORDER_HOST
                        PORT = ORDER_PORT
                        DATABASE = ORDER_DATABASE
                        schema = "public"
                    if CURRENT_DATABASE == "EnabilityPortal":
                        USER = PORTAL_USER
                        PASSWORD = PORTAL_PASSWORD
                        HOST = PORTAL_HOST
                        PORT = PORTAL_PORT
                        DATABASE = PORTAL_DATABASE
                        schema = "public"
                    conn_params = {
                        'user': USER,
                        'password': PASSWORD,
                        'host': HOST,
                        'port': PORT,
                        'database': DATABASE
                    }

                    import_flag = True

                    # 连接到数据库
                    connection = psycopg2.connect(**conn_params)
                    cursor = connection.cursor()

                    # 获取所有表名
                    cursor.execute(f"""
                        SELECT table_name
                        FROM information_schema.tables
                        WHERE table_schema = '{schema}'
                        AND table_type = 'BASE TABLE';
                    """)
                    tables = cursor.fetchall()

                    tab_table = self.table
                    self.clearAllTables()
                    time.sleep(1)
                    tab_table.setColumnCount(2)
                    tab_table.setHorizontalHeaderLabels(["CSVファイル", "導入状態"])
                    # 对每张表执行导入操作
                    for table in tables:
                        table_name = table[0]
                        csv_filename = os.path.join(TARGET_FOLDER_IMPORT, f"{table_name}.csv")
                        if os.path.exists(csv_filename):
                            # 打开CSV文件并导入到数据库表中
                            with open(csv_filename, 'r', newline='', encoding='utf-8') as f:
                                try:
                                    print("import csv filename :", csv_filename)
                                    cursor.copy_expert(f"COPY {table_name} FROM STDIN WITH CSV HEADER", f)
                                except (Exception, Error) as error:
                                    set_message_box("CRITICAL", "データベース", error.args[0])
                                    import_flag = False
                                    break
                                tab_table.resizeColumnsToContents()
                                tab_table.resizeRowsToContents()
                                row_append(self.table, [csv_filename, "〇"], Qt.lightGray, False)

                    connection.commit()
                    cursor.close()
                    connection.close()
                    if import_flag is True:
                        set_message_box("WARNING", "データベース", "導入成功")
                    else:
                        set_message_box("WARNING", "データベース", "導入失敗")
            except Exception as e:
                print("An error occurred : ", e)
                raise

        if button.text() == '全データベースクリア':
            reply = QMessageBox.question(self, '確認',
                                         'データベースが削除されますので、データのバックアップを取っていることを確認してください。',
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                pass
            else:
                return
            if CURRENT_DATABASE == "EnabilityCIS":
                USER = CIS_USER
                PASSWORD = CIS_PASSWORD
                HOST = CIS_HOST
                PORT = CIS_PORT
                DATABASE = CIS_DATABASE
                schema = "unisys"
            if CURRENT_DATABASE == "EnabilityOrder":
                USER = ORDER_USER
                PASSWORD = ORDER_PASSWORD
                HOST = ORDER_HOST
                PORT = ORDER_PORT
                DATABASE = ORDER_DATABASE
                schema = "public"
            if CURRENT_DATABASE == "EnabilityPortal":
                USER = PORTAL_USER
                PASSWORD = PORTAL_PASSWORD
                HOST = PORTAL_HOST
                PORT = PORTAL_PORT
                DATABASE = PORTAL_DATABASE
                schema = "public"
            conn_params = {
                'user': USER,
                'password': PASSWORD,
                'host': HOST,
                'port': PORT,
                'database': DATABASE
            }
            connection = psycopg2.connect(**conn_params)
            cursor = connection.cursor()
            tab_table = self.table
            self.clearAllTables()
            time.sleep(1)
            tab_table.setColumnCount(2)
            tab_table.setHorizontalHeaderLabels(["テーブル名", "クリア状態"])
            try:
                # 获取所有表名
                cursor.execute(
                    "SELECT table_name FROM information_schema.tables "
                    f"WHERE table_schema='{schema}' AND table_type='BASE TABLE';")
                tables = cursor.fetchall()

                # 逐个清空表
                for table in tables:
                    table_name = table[0]
                    truncate_query = f"TRUNCATE TABLE {table_name} RESTART IDENTITY CASCADE;"
                    cursor.execute(truncate_query)
                    tab_table.resizeColumnsToContents()
                    tab_table.resizeRowsToContents()
                    row_append(self.table, [table_name, "〇"], Qt.lightGray, False)
                    print(f"Table '{table_name}' has been truncated.")

                # 提交事务
                connection.commit()
                print("All tables have been truncated successfully.")

            except (Exception, psycopg2.DatabaseError) as error:
                print("Error while truncating tables:", error)

            finally:
                # 关闭游标和连接
                if cursor:
                    cursor.close()
                if connection:
                    connection.close()
                    set_message_box("WARNING", "データベース", "クリア成功")

    def simulate_button_click(self):
        self.button.click()

    def openTabContextMenu(self, pos, tab):
        menu = QMenu(self)
        remove_action = QAction("Remove Tab", self)
        remove_action.triggered.connect(lambda: self.removeTab(tab))
        menu.addAction(remove_action)
        menu.exec_(tab.mapToGlobal(pos))

    def removeTab(self):
        self.tab_counter -= 1
        self.tab_widget.removeTab(self.tab_widget.currentIndex())

    def clearAllTables(self):
        tableWidget = self.table
        if tableWidget is not None:
            tableWidget.setRowCount(0)
            for row in range(tableWidget.rowCount()):
                for col in range(tableWidget.columnCount()):
                    item = tableWidget.item(row, col)
                    if item is not None:
                        item.setText("")

    def timer_init(self):
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_datetime)
        self.timer.start(1000)

    def update_datetime(self):
        self.current_datetime = QDateTime.currentDateTime().toString(Qt.ISODate)
        self.tips_label.setText(f'{self.current_datetime}')


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('Windows')  # Windows , windowsvista , Fusion
    login_dialog = LoginDialog()
    result = login_dialog.exec_()

    while result == QDialog.Accepted:
        if LOGIN_SUCCESS:
            # If login is successful, create and show main window
            window = MainWindow()
            window.show()
            # window.simulate_button_click()
            sys.exit(app.exec_())
        else:
            # If login is not successful, show login dialog again
            login_dialog = LoginDialog()
            result = login_dialog.exec_()

    # If user cancels the login dialog or exits, exit the application
    sys.exit()
