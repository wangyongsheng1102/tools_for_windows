import functools
import gc
import glob
import os
import random
import subprocess
import sys
from collections import defaultdict
from configparser import ConfigParser
from datetime import datetime
from zoneinfo import ZoneInfo

import openpyxl
from PyQt5.QtCore import QObject, QEvent, QTimer, QDateTime, Qt, QDate
from PyQt5.QtGui import QColor, QFont, QPixmap, QPainter
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTabWidget, QProgressBar, QLabel, \
    QHBoxLayout, QGroupBox, QMainWindow, QLineEdit, QMessageBox, QAction, QMenu, QTableWidget, QComboBox, QSizePolicy, \
    QHeaderView, QTableWidgetItem, QFileDialog, QCheckBox, QDateEdit, QDialog, QItemDelegate, QListWidget, \
    QListWidgetItem, QSplashScreen, QGraphicsDropShadowEffect
from git import Repo
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from tools_collection.toolkit import neumorphism_helper

CELL_COLOR_RED = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
CELL_COLOR_YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
CELL_COLOR_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

KLEIN_BLUE = QColor(0, 47, 167)
TIFFANY_BLUE = QColor(131, 216, 208)

SCHENBRUNN_YELLOW = QColor(247, 225, 77)
MARS_GREEN = QColor(0, 140, 140)

TITIAN_RED = QColor(176, 89, 35)
CHINA_RED = QColor(230, 0, 0)

HERMES_ORANGE = QColor(232, 88, 39)

WHITE = QColor(255, 255, 255)

# QFontオブジェクトを作成
FONT_STYLE = QFont("Microsoft YaHei", 9)  # フォントファミリ、サイズを指定

HOVER_STYLE = """
    QPushButton:hover {
        background-color: darkgray;
        box-shadow: 0px 0px 20px rgba(0, 0, 0, 0.5);
    }
"""

HOVER_STYLE_WHITE = """
    QPushButton {
        background-color: gray;
    }

    QPushButton:hover {
        background-color: darkgray;
        box-shadow: 0px 0px 10px rgba(0, 0, 0, 0.5);
    }
"""

PROGRESS_BAR_STYLE = """
    QProgressBar {
        border: 1px solid #999999;
        text-align: center;
    }

    QProgressBar::chunk {
        background-color: white;
    }
"""

SHADOW_BUTTON = QGraphicsDropShadowEffect()
SHADOW_BUTTON.setBlurRadius(20)
SHADOW_BUTTON.setOffset(5, 5)
SHADOW_BUTTON.setColor(QColor(160, 160, 160))

INPUT_SUFFIX, \
    INPUT_USER, \
    INPUT_PROJECT, \
    INPUT_DATE = "", "", "", ""

CHECKBOX_1 = False
CHECKBOX_2 = False
CHECKBOX_3 = False
CHECKBOX_4 = False
INPUT_KEYWORD1, \
    INPUT_KEYWORD2, \
    INPUT_KEYWORD3, \
    INPUT_KEYWORD4 = "", "", "", ""

FILE_SUFFIX = [
    "java",
    "jsp",
    "js",
    "css",
    "html"
]

FILE_TYPE = [
    "ソース",
    "コメント"
]

USER_TYPE = [
    '日本側',
    '大連側'
]

JAPAN_USER_LIST = []
JAPAN_USER_LIST_TMP = []

PROJECT_LIST = []


# ユーザー取得コマンド：
# git log --pretty=format:'%an' | sort | uniq　
# Git Bash


def log_and_call(func):
    @functools.wraps(func)
    def wrapper(self):
        set_status_label(self, func.__doc__ + "中")
        return func(self)

    return wrapper


def set_status_label(self, context):
    self.status_label.setText(context)
    QApplication.processEvents()


def set_progress_bar(self, value):
    self.progress_bar.setValue(int(value))
    QApplication.processEvents()


def get_program_path():
    """アプリのパスを取得"""
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def get_config_file_path():
    """コンフィグのパスを取得"""
    return os.path.join(get_program_path(), ".todo_config.ini")


def load_config_content(tag):
    """コンフィグをロード"""
    config = ConfigParser()
    config_path = get_config_file_path()
    if os.path.exists(config_path) is False:
        raise
    config.read(config_path, encoding='utf-8')
    return config[tag] if tag in config else {}


def get_folder(project_path):
    folder = []
    for foldername, subfolders, _ in os.walk(project_path):
        if 'src' in subfolders:
            folder.append(foldername.replace("\\", "/").split("/")[len(foldername.replace("\\", "/").split("/")) - 1])
    return folder


def init_config_content():
    global INPUT_SUFFIX, \
        INPUT_USER, \
        INPUT_DATE, \
        INPUT_PROJECT, \
        CHECKBOX_1, \
        CHECKBOX_2, \
        CHECKBOX_3, \
        CHECKBOX_4, \
        INPUT_KEYWORD1, \
        INPUT_KEYWORD2, \
        INPUT_KEYWORD3, \
        INPUT_KEYWORD4, \
        JAPAN_USER_LIST, \
        PROJECT_LIST
    try:
        ids = load_config_content('Ids')
        paths = load_config_content('Paths')
        users = load_config_content('Users')
    except Exception as e:
        set_message_box("CRITICAL", "コンフィグ", "コンフィグファイルが存在しませんが、チェックしてください。")
        return
    if ids['input_suffix'] is not None:
        INPUT_SUFFIX = ids['input_suffix']
    if ids['input_user'] is not None:
        INPUT_USER = ids['input_user']
    if ids['input_date'] is not None:
        INPUT_DATE = ids['input_date']
    if ids['checkbox_1'] is not None:
        CHECKBOX_1 = eval(ids['checkbox_1'])
    if ids['input_keyword_1'] is not None:
        INPUT_KEYWORD1 = ids['input_keyword_1']
    if ids['checkbox_2'] is not None:
        CHECKBOX_2 = eval(ids['checkbox_2'])
    if ids['input_keyword_2'] is not None:
        INPUT_KEYWORD2 = ids['input_keyword_2']
    if ids['checkbox_3'] is not None:
        CHECKBOX_3 = eval(ids['checkbox_3'])
    if ids['input_keyword_3'] is not None:
        INPUT_KEYWORD3 = ids['input_keyword_3']
    if ids['checkbox_4'] is not None:
        CHECKBOX_4 = eval(ids['checkbox_4'])
    if ids['input_keyword_4'] is not None:
        INPUT_KEYWORD4 = ids['input_keyword_4']
    if paths['input_project'] is not None:
        INPUT_PROJECT = paths['input_project']
    if users['japan_user_list'] is not None:
        JAPAN_USER_LIST = users['japan_user_list'].split(",")
    PROJECT_LIST = get_folder(INPUT_PROJECT)


def save_file_paths(self):
    """コンフィグにパスインフォを保存"""
    config = ConfigParser()
    config_path = get_config_file_path()
    if os.path.exists(config_path) is False:
        return
    config.read(config_path, encoding='utf-8')
    if self.input_suffix.currentText():
        config.set('Ids', 'input_suffix', self.input_suffix.currentText())
    if self.input_user.currentText():
        config.set('Ids', 'input_user', self.input_user.currentText())
    if self.input_date.date():
        config.set('Ids', 'input_date', self.input_date.date().toString("yyyy-MM-dd"))
    if self.input_project.text():
        config.set('Paths', 'input_project', self.input_project.text())

    if self.checkbox_1:
        result = self.checkbox_1.isChecked()
        config.set('Ids', 'checkbox_1', str(result))
    if self.input_keyword_1.text():
        config.set('Ids', 'input_keyword_1', self.input_keyword_1.text())
    if self.checkbox_2:
        result = self.checkbox_2.isChecked()
        config.set('Ids', 'checkbox_2', str(result))
    if self.input_keyword_2.text():
        config.set('Ids', 'input_keyword_2', self.input_keyword_2.text())
    if self.checkbox_3:
        result = self.checkbox_3.isChecked()
        config.set('Ids', 'checkbox_3', str(result))
    if self.input_keyword_3.text():
        config.set('Ids', 'input_keyword_3', self.input_keyword_3.text())
    if self.checkbox_4:
        result = self.checkbox_4.isChecked()
        config.set('Ids', 'checkbox_4', str(result))
    if self.input_keyword_4.text():
        config.set('Ids', 'input_keyword_4', self.input_keyword_4text())
    if JAPAN_USER_LIST:
        config.set('Users', 'japan_user_list', ",".join(map(str, JAPAN_USER_LIST)))

    with open(get_config_file_path(), 'w', encoding='utf-8') as configfile:
        config.write(configfile)


def set_message_box(message_type, title, context):
    """メッセージを反映"""
    if message_type == 'WARNING':
        QMessageBox.warning(None, title, context)
    if message_type == 'CRITICAL':
        QMessageBox.critical(None, title, context)
    if message_type == 'INFO':
        QMessageBox.information(None, title, context)
    if message_type == 'QUESTION':
        QMessageBox.question(None, title, context)


def generate_random_color(color):
    if color == "blue":
        blue = random.randint(150, 255)
        red = random.randint(0, blue - 50)
        green = random.randint(0, blue - 50)
    if color == "red":
        red = random.randint(200, 255)
        green = random.randint(0, 100)
        blue = random.randint(0, 100)
    if color == "yellow":
        red = random.randint(200, 255)
        green = random.randint(200, 255)
        blue = random.randint(0, 100)
    return red, green, blue


def check_git_installed():
    try:
        # Git --version を実行して確認
        result = subprocess.run(
            ['git', '--version'],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=True,
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        print(f"✅ Gitがインストールされています。バージョン: {result.stdout.strip()}")
        set_message_box("INFO", "GIT", f"✅ Gitがインストールされています。バージョン: {result.stdout.strip()}")
        return True
    except FileNotFoundError:
        print("❌ Gitがインストールされていません。Gitをインストールし、システムPATHに追加してください。")
        print("👉 ダウンロードURL: https://git-scm.com/downloads")
        set_message_box("CRITICAL", "GIT", "❌ Gitがインストールされていません。Gitをインストールし、システムPATHに追加してください。")
        return False
    except subprocess.CalledProcessError as e:
        print(f"⚠️ Gitコマンドの実行に失敗しました: {e.stderr.strip()}")
        set_message_box("CRITICAL", "GIT", f"⚠️ Gitコマンドの実行に失敗しました: {e.stderr.strip()}")
        return False


class CharacterSplashScreen(QSplashScreen):
    def __init__(self, animation_chars, parent=None):
        super().__init__(parent)

        pixmap = QPixmap(300, 150)
        pixmap.fill(Qt.transparent)

        self.setPixmap(pixmap)
        self.animation_chars = animation_chars
        self.char_index = 0
        self.message = "アプリ起動中"

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_animation)
        self.timer.start(150)

    def update_animation(self):
        self.char_index = (self.char_index + 1) % len(self.animation_chars)
        self.repaint()

    def drawContents(self, painter: QPainter):
        current_char = self.animation_chars[self.char_index]
        display_text = f"{self.message}{current_char}"

        font = QFont("Microsoft YaHei", 14)
        painter.setFont(font)
        painter.setPen(QColor(200, 200, 200))
        painter.drawText(self.rect(), Qt.AlignCenter, display_text)

    def set_loading_message(self, message):
        self.message = message
        self.repaint()


class BlinkingLabel(QLabel):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self._timer = QTimer(self)
        self._timer.timeout.connect(self.toggle_visibility)
        # self._timer.start(1000)
        self.base_text = ""
        self.dot_count = 0
        self.direction = 1

    def start_blinking(self):
        self._timer.start(1000)

    def stop_blinking(self):
        color = QColor(0, 0, 0)
        self.setStyleSheet("QLabel { color: %s }" % color.name())
        self._timer.stop()

    def toggle_visibility(self):
        self.base_text = self.text().replace(".", "")
        self.dot_count += self.direction

        if self.dot_count == 5:
            self.direction = -1
        elif self.dot_count == 0:
            self.direction = 1

        dots = '.' * self.dot_count
        self.setText(self.base_text + dots)

        # color = QColor(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
        red, green, blue = generate_random_color("blue")
        color = QColor(red, green, blue)
        self.setStyleSheet("QLabel { color: %s }" % color.name())
        # self.setVisible(not self.isVisible())


class ComboCheckBox(QComboBox):
    def __init__(self, parent):
        """
        initial function
        """
        super(ComboCheckBox, self).__init__(parent)

        self.box_list = []  # selected items
        self.text = QLineEdit()  # use to selected items
        self.state = 0  # use to record state
        # self.setStyleSheet("width: 300px; height: 50px; font-size: 21px; font-weight: bold")
        self.text.setReadOnly(True)
        self.setLineEdit(self.text)

    def myadditems(self, items):
        """

        :param items: 传入下拉选项
        :return:
        """
        self.items = ["全て"] + items  # items list
        q = QListWidget()
        for i in range(len(self.items)):
            self.box_list.append(QCheckBox())
            self.box_list[i].setText(self.items[i])
            item = QListWidgetItem(q)
            q.setItemWidget(item, self.box_list[i])
            if i == 0:
                self.box_list[i].stateChanged.connect(self.all_selected)
            else:
                self.box_list[i].stateChanged.connect(self.show_selected)

        # q.setStyleSheet("font-size: 20px; font-weight: bold; height: 40px; margin-left: 5px")
        self.setModel(q.model())
        self.setView(q)

    def all_selected(self):
        """
        decide whether to check all
        :return:
        """
        # change state
        if self.state == 0:
            self.state = 1
            for i in range(1, len(self.items)):
                self.box_list[i].setChecked(True)
        else:
            self.state = 0
            for i in range(1, len(self.items)):
                self.box_list[i].setChecked(False)
        self.show_selected()

    def get_selected(self) -> list:
        """
        get selected items
        :return:
        """
        ret = []
        for i in range(1, len(self.items)):
            if self.box_list[i].isChecked():
                ret.append(self.box_list[i].text())
        return ret

    def set_checked(self, items_to_check):
        """
        动态设置某几项为选中状态
        :param items_to_check: 要选中的项（字符串列表）
        """
        for idx, cb in enumerate(self.box_list):
            if idx > 0 and cb.text() in items_to_check:
                cb.setChecked(True)

        # 更新显示
        self.show_selected()

    def show_selected(self):
        """
        show selected items
        :return:
        """
        self.text.clear()
        ret = '; '.join(self.get_selected())
        self.text.setText(ret)


class TableWidget(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setColumnCount(7)

    def contextMenuEvent(self, event):
        menu = QMenu(self)

        # 复制当前选中的单元格内容
        copy_action = QAction("コピー", self)
        copy_action.triggered.connect(self.copy_selected_items)
        menu.addAction(copy_action)

        # 显示上下文菜单
        menu.exec(event.globalPos())

    def copy_selected_items(self):
        clipboard = QApplication.clipboard()
        selected_ranges = self.selectedRanges()

        if not selected_ranges:
            return  # 没有选中任何区域

        copied_text = []

        for sel_range in selected_ranges:
            top_row = sel_range.topRow()
            bottom_row = sel_range.bottomRow()
            left_col = sel_range.leftColumn()
            right_col = sel_range.rightColumn()

            # 遍历每一行
            for row in range(top_row, bottom_row + 1):  # ✅ 使用正确的 range 函数
                row_data = []
                for col in range(left_col, right_col + 1):  # ✅ 使用正确的 range 函数
                    item = self.item(row, col)
                    row_data.append(item.text() if item else "")
                copied_text.append("\t".join(row_data))  # 同一行用制表符分隔

        clipboard.setText("\n".join(copied_text))  # 不同行用换行符分隔


class SubWindow(QDialog):

    def __init__(self, main_window):
        super().__init__()

        self.helper = neumorphism_helper.NeumorphicHelper()

        self.setWindowTitle("ユーザー確認")
        self.setFixedSize(400, 400)

        # # 设置无边框
        # self.setWindowFlags(Qt.FramelessWindowHint)
        #
        # # 启用透明背景
        # self.setAttribute(Qt.WA_TranslucentBackground)
        #
        # # 设置透明度
        # self.setWindowOpacity(1)

        self.main_window = main_window  # 保存主窗体引用
        self.init_ui()

    def init_ui(self):
        global JAPAN_USER_LIST_TMP

        JAPAN_USER_LIST_TMP = JAPAN_USER_LIST.copy()
        # 主布局
        main_layout = QVBoxLayout(self)

        # 创建表格
        self.table = QTableWidget()
        self.helper.apply_neumorphism(self.table)
        self.table.setColumnCount(3)  # 3 columns: 数据列 + 操作按钮列
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)

        # 设置表头
        self.table.setHorizontalHeaderLabels(["ユーザー", "状態", "操作"])
        table_font = self.table.horizontalHeader().font()
        table_font.setBold(True)
        table_font.setUnderline(True)
        self.table.horizontalHeader().setFont(table_font)

        result_list = self.get_all_user_from_git()

        user_list = [x for x in result_list if x]

        user_list.sort()

        for index, user in enumerate(user_list):
            current_rows = self.table.rowCount()
            self.table.setRowCount(current_rows + 1)
            name_item = QTableWidgetItem(user)
            name_item.setFont(table_font)
            status = "✅" if user in JAPAN_USER_LIST_TMP else "❌"
            status_item = QTableWidgetItem(status)

            self.table.setItem(index, 0, name_item)
            self.table.setItem(index, 1, status_item)

            button = QPushButton("削除" if user in JAPAN_USER_LIST_TMP else "追加")
            self.helper.apply_neumorphism(button)
            if user in JAPAN_USER_LIST_TMP:
                button.setStyleSheet(HOVER_STYLE_WHITE)
            else:
                button.setStyleSheet(HOVER_STYLE)
            button.clicked.connect(lambda checked, r=index: self.on_button_click(r))
            self.table.setCellWidget(index, 2, button)

            self.table.resizeColumnsToContents()
            self.table.resizeRowsToContents()

        # 添加表格到布局
        main_layout.addWidget(self.table)

        # 底部按钮布局
        btn_layout = QHBoxLayout()

        confirm_btn = QPushButton("確認")
        confirm_btn.setStyleSheet(HOVER_STYLE)
        confirm_btn.clicked.connect(self.on_confirm_clicked)
        self.helper.apply_neumorphism(confirm_btn)

        cancel_btn = QPushButton("キャンセル")
        cancel_btn.setStyleSheet(HOVER_STYLE_WHITE)
        cancel_btn.clicked.connect(self.reject)
        self.helper.apply_neumorphism(cancel_btn)

        btn_layout.addWidget(confirm_btn)
        btn_layout.addWidget(cancel_btn)

        main_layout.addLayout(btn_layout)

    def on_button_click(self, row):
        """处理行内按钮点击，动态切换按钮文本和状态"""
        user = self.table.item(row, 0).text()  # 获取当前行用户名
        current_button = self.table.cellWidget(row, 2)  # 获取当前行按钮

        # 判断当前按钮状态并执行对应操作
        if current_button.text() == "削除":
            # 执行删除操作
            if user in JAPAN_USER_LIST_TMP:
                JAPAN_USER_LIST_TMP.remove(user)
            self.table.setItem(row, 1, QTableWidgetItem("❌"))
            current_button.setText("追加")
            current_button.setStyleSheet(HOVER_STYLE)
        else:
            # 执行添加操作
            if user not in JAPAN_USER_LIST_TMP:
                JAPAN_USER_LIST_TMP.append(user)
            self.table.setItem(row, 1, QTableWidgetItem("✅"))
            current_button.setText("削除")
            current_button.setStyleSheet(HOVER_STYLE_WHITE)

        QApplication.processEvents()

    def get_all_user_from_git(self):
        """git log --pretty=format:'%an' | sort | uniq　"""
        result = subprocess.run(
            ['git', 'log', '--format=%aN', '--use-mailmap'],
            cwd=self.main_window.input_project.text(),
            capture_output=True,
            text=True,
            creationflags=subprocess.CREATE_NO_WINDOW
        )

        return list(set(result.stdout.split("\n")))

    def on_confirm_clicked(self):
        global JAPAN_USER_LIST_TMP, JAPAN_USER_LIST

        JAPAN_USER_LIST = JAPAN_USER_LIST_TMP.copy()
        self.accept()


class EventFilter(QObject):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window

    def eventFilter(self, obj, event):
        if isinstance(obj, QLineEdit) and \
                event.type() == QEvent.Type.FocusOut:
            if obj.objectName() == "input_user":
                pass
                # get_svn_username(self.main_window.input_svn.text())
                # if obj.text() in ADMIN_NAME:
                #     self.main_window.all_button.setEnabled(True)
                # else:
                #     self.main_window.all_button.setEnabled(False)
            if obj.objectName() == "input_branch":
                pass
                # branch = self.main_window.input_branch.text()
                # clone_dir = os.path.join(get_program_path(), "git_" + branch)
                # if os.path.exists(clone_dir):
                #     os.chdir(clone_dir)
                # repo_url = REPO_URL
                # result = subprocess.run(['git', 'ls-remote', repo_url, branch], capture_output=True, text=True)
                # print(result)
                # if result.returncode != 0 or result.stdout == '':
                #     self.main_window.setDisabled(False)
                #     set_message_box("CRITICAL", "GIT", f"ブランチ「{branch}」が存在しません。\n"
                #                                        f"エラー「{result.returncode}」が発生しました。")
            # local_ip = get_local_ip()
            # print(local_ip)
        return super().eventFilter(obj, event)


def check_substrings(target, *substrings):
    """
    判断非空字符串是否全部存在于目标字符串中

    :param target: 要搜索的长字符串
    :param substrings: 要检查的多个子字符串
    :return: 如果所有非空子字符串都存在于目标字符串中，返回 True；否则返回 False
    """
    for s in substrings:
        if s:  # 仅处理非空字符串
            if s not in target:
                return False
    return True


class MainWindow(QMainWindow):
    """メインウィンドウ"""

    def __init__(self):
        super().__init__()

        self.helper = neumorphism_helper.NeumorphicHelper()

        self.table = None
        self.event_filter = None
        self.top_group = QGroupBox("選択")
        self.top_group.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.top_group)
        self.top_layout_1 = QHBoxLayout()
        self.top_layout_2 = QHBoxLayout()
        self.top_layout_3 = QHBoxLayout()
        self.top_layout_4 = QHBoxLayout()

        self.top_layout = QVBoxLayout()
        self.label_width_long = 100
        self.label_width_short = 50
        self.label_suffix = QLabel('拡張子')
        self.label_suffix.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.label_suffix)
        # self.input_suffix = QComboBox()
        self.input_suffix = ComboCheckBox(self)
        self.input_suffix.setObjectName("input_suffix")
        self.input_suffix.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.input_suffix)
        self.label_user = QLabel('ユーザー')
        self.label_user.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.label_user)
        self.input_user = ComboCheckBox(self)
        # self.input_user.setObjectName("input_user")
        self.input_user.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.input_user)
        self.label_date = QLabel('日付')
        self.label_date.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.label_date)
        self.input_date = QDateEdit()
        self.input_date.setFont(FONT_STYLE)
        self.input_date.setCalendarPopup(True)
        self.input_date.setDate(QDate.currentDate())
        self.helper.apply_neumorphism(self.input_date)
        # self.label_project_folder = QLabel('フォルダー')
        # self.label_project_folder.setFont(FONT_STYLE)
        # self.input_project_folder = ComboCheckBox(self)
        self.label_project = QLabel('プロジェクトパス')
        self.label_project.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.label_project)
        self.input_project = QLineEdit()
        self.input_project.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.input_project)
        self.button_project = QPushButton('開く')
        self.button_project.setFont(FONT_STYLE)
        self.button_project.setStyleSheet(HOVER_STYLE)
        self.helper.apply_neumorphism(self.button_project)

        self.checkbox_1 = QCheckBox("")
        self.helper.apply_neumorphism(self.checkbox_1)
        self.input_keyword_1 = QLineEdit()
        self.input_keyword_1.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.input_keyword_1)
        self.checkbox_2 = QCheckBox("")
        self.helper.apply_neumorphism(self.checkbox_2)
        self.input_keyword_2 = QLineEdit()
        self.input_keyword_2.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.input_keyword_2)
        self.checkbox_3 = QCheckBox("")
        self.helper.apply_neumorphism(self.checkbox_3)
        self.input_keyword_3 = QLineEdit()
        self.input_keyword_3.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.input_keyword_3)
        self.checkbox_4 = QCheckBox("")
        self.helper.apply_neumorphism(self.checkbox_4)
        self.input_keyword_4 = QLineEdit()
        self.input_keyword_4.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.input_keyword_4)

        self.bottom_right_group = QGroupBox('結果')
        self.bottom_right_group.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.bottom_right_group)
        self.bottom_right_layout = QVBoxLayout()
        self.tab = QTabWidget()
        self.helper.apply_neumorphism(self.tab)
        self.tab_layout = QVBoxLayout()

        self.bottom_layout = QHBoxLayout()
        self.button_group = QGroupBox("操作")
        self.button_group.setFont(FONT_STYLE)
        self.helper.apply_neumorphism(self.button_group)
        self.button_layout = QHBoxLayout()
        self.confirm_button = QPushButton('ユーザー確認')
        self.confirm_button.setFont(FONT_STYLE)
        self.confirm_button.setStyleSheet(HOVER_STYLE)
        self.helper.apply_neumorphism(self.confirm_button)
        self.search_button = QPushButton('検索')
        self.search_button.setFont(FONT_STYLE)
        self.search_button.setStyleSheet(HOVER_STYLE)
        # self.search_button.setGraphicsEffect(SHADOW_BUTTON)
        self.helper.apply_neumorphism(self.search_button)
        self.save_button = QPushButton('出力')
        self.save_button.setFont(FONT_STYLE)
        self.save_button.setStyleSheet(HOVER_STYLE)
        self.save_button.setDisabled(True)
        # self.save_button.setGraphicsEffect(SHADOW_BUTTON)
        self.helper.apply_neumorphism(self.save_button)
        self.exit_button = QPushButton('退出')
        self.exit_button.setFont(FONT_STYLE)
        self.exit_button.setStyleSheet(HOVER_STYLE_WHITE)
        # self.exit_button.setGraphicsEffect(SHADOW_BUTTON)
        self.helper.apply_neumorphism(self.exit_button)
        self.tips_group = QGroupBox("状態")
        self.tips_group.setFont(FONT_STYLE)
        self.tips_layout = QVBoxLayout()
        self.tips_layout_1 = QHBoxLayout()
        self.tips_layout_2 = QHBoxLayout()
        self.status_label = BlinkingLabel('画面初期化')
        self.helper.apply_neumorphism(self.status_label)

        # QLabelにフォントを適用
        self.status_label.setFont(FONT_STYLE)
        self.tips_label = QLabel('')
        self.progress_bar = QProgressBar()
        self.progress_bar.setFont(FONT_STYLE)
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.helper.apply_neumorphism(self.progress_bar)
        # self.progress_bar.setStyleSheet(PROGRESS_BAR_STYLE)
        self.main_layout = QVBoxLayout()
        # self.setAttribute(Qt.WA_TranslucentBackground, True)
        # self.setWindowFlags(Qt.FramelessWindowHint)
        self.initUI()

    def initUI(self):
        """initUI"""
        self.label_suffix.setFixedWidth(self.label_width_short)
        self.label_suffix.setToolTip('検索ファイルの拡張子を指定してください。')
        # self.input_suffix.addItems(FILE_SUFFIX)
        self.input_suffix.myadditems(FILE_SUFFIX)
        self.input_suffix.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.label_user.setFixedWidth(self.label_width_short)
        self.label_user.setToolTip('ユーザータイプを指定してください。')
        self.input_user.myadditems(USER_TYPE)
        self.input_user.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.label_date.setFixedWidth(self.label_width_short)
        self.label_date.setToolTip('選択した日付以降のすべての情報を検索します。\n'
                                   '※指定日付が当日である場合、全時間帯の履歴データを取得します。')
        self.input_date.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        # self.input_project_folder.myadditems(PROJECT_LIST)
        # self.input_project_folder.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.label_project.setFixedWidth(self.label_width_long)
        self.label_project.setToolTip('Gitリポジトリのルートディレクトリ（.gitフォルダを含む）を選択してください。')

        self.top_layout_1.addWidget(self.label_suffix)
        self.top_layout_1.addWidget(self.input_suffix)
        self.confirm_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.top_layout_1.addWidget(self.confirm_button)
        self.top_layout_1.addWidget(self.label_user)
        self.top_layout_1.addWidget(self.input_user)
        self.top_layout_1.addWidget(self.label_date)
        self.top_layout_1.addWidget(self.input_date)
        # self.top_layout_2.addWidget(self.label_project_folder)
        # self.top_layout_2.addWidget(self.input_project_folder)
        self.top_layout_2.addWidget(self.label_project)
        self.top_layout_2.addWidget(self.input_project)
        self.top_layout_2.addWidget(self.button_project)

        self.top_layout_3.addWidget(self.checkbox_1)
        self.top_layout_3.addWidget(self.input_keyword_1)
        self.top_layout_3.addWidget(self.checkbox_2)
        self.top_layout_3.addWidget(self.input_keyword_2)
        self.top_layout_3.addWidget(self.checkbox_3)
        self.top_layout_3.addWidget(self.input_keyword_3)
        self.top_layout_3.addWidget(self.checkbox_4)
        self.top_layout_3.addWidget(self.input_keyword_4)

        self.top_layout.addLayout(self.top_layout_1)
        self.top_layout.addLayout(self.top_layout_2)
        self.top_layout.addLayout(self.top_layout_3)
        self.top_layout.addLayout(self.top_layout_4)
        self.top_group.setLayout(self.top_layout)

        self.table = TableWidget(self)
        self.table.setHorizontalHeaderLabels(['ファイルパス', 'GITハッシュ値', 'GIT日付', 'GIT作成者', '桁目', '内容', '状態'])
        table_font = self.table.horizontalHeader().font()
        table_font.setBold(True)
        table_font.setUnderline(True)
        self.table.horizontalHeader().setFont(table_font)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.setSortingEnabled(True)
        self.tab_layout.addWidget(self.table)
        self.tab.setLayout(self.tab_layout)

        self.bottom_right_layout.addWidget(self.tab)
        self.bottom_right_group.setLayout(self.bottom_right_layout)
        self.bottom_layout.addWidget(self.bottom_right_group)
        # self.button_layout.addWidget(self.confirm_button)
        self.button_layout.addWidget(self.search_button)
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.exit_button)
        self.button_group.setLayout(self.button_layout)

        self.status_label.start_blinking()
        self.tips_layout_1.addWidget(self.status_label)
        self.tips_layout_2.addWidget(self.tips_label)
        self.tips_layout_2.addWidget(self.progress_bar)
        self.tips_layout.addLayout(self.tips_layout_1)
        self.tips_layout.addLayout(self.tips_layout_2)
        self.tips_group.setLayout(self.tips_layout)

        self.main_layout.addWidget(self.top_group, 2)
        self.main_layout.addLayout(self.bottom_layout, 5)
        self.main_layout.addWidget(self.button_group, 1)
        self.main_layout.addWidget(self.tips_group, 1)

        central_widget = QWidget()
        central_widget.setLayout(self.main_layout)
        self.setCentralWidget(central_widget)
        self.setLayout(self.main_layout)

        self.setWindowTitle(f'Enability-GIT履歴検索＆報告生成-Powered by PyQt5')
        self.setGeometry(300, 200, 1200, 600)

        self.timer_init()
        self.init_ui()
        init_config_content()

        if INPUT_SUFFIX is not None:
            self.input_suffix.setCurrentText(INPUT_SUFFIX)
            self.input_suffix.set_checked(INPUT_SUFFIX.replace(" ", "").split(";"))
        if INPUT_USER is not None:
            self.input_user.setCurrentText(INPUT_USER)
            self.input_user.set_checked(INPUT_USER.replace(" ", "").split(";"))
        if INPUT_DATE is not None:
            self.input_date.setDate(QDate.fromString(INPUT_DATE, "yyyy-MM-dd"))
        if INPUT_PROJECT is not None:
            self.input_project.setText(INPUT_PROJECT)
        if CHECKBOX_1 is not None:
            self.checkbox_1.setChecked(CHECKBOX_1)
            self.input_keyword_1.setEnabled(CHECKBOX_1)
        if INPUT_KEYWORD1 is not None:
            self.input_keyword_1.setText(INPUT_KEYWORD1)
        if CHECKBOX_2 is not None:
            self.checkbox_2.setChecked(CHECKBOX_2)
            self.input_keyword_2.setEnabled(CHECKBOX_2)
        if INPUT_KEYWORD2 is not None:
            self.input_keyword_2.setText(INPUT_KEYWORD2)
        if CHECKBOX_3 is not None:
            self.checkbox_3.setChecked(CHECKBOX_3)
            self.input_keyword_3.setEnabled(CHECKBOX_3)
        if INPUT_KEYWORD3 is not None:
            self.input_keyword_3.setText(INPUT_KEYWORD3)
        if CHECKBOX_4 is not None:
            self.checkbox_4.setChecked(CHECKBOX_4)
            self.input_keyword_4.setEnabled(CHECKBOX_4)
        if INPUT_KEYWORD4 is not None:
            self.input_keyword_4.setText(INPUT_KEYWORD4)

        # if PROJECT_LIST is not None:
        #     self.input_project_folder.myadditems(PROJECT_LIST)
        #     self.input_project_folder.set_checked(PROJECT_LIST)

    def init_ui(self):
        """init_ui"""
        self.event_filter = EventFilter(self)
        # self.input_user.installEventFilter(self.event_filter)
        # self.input_branch.installEventFilter(self.event_filter)
        self.checkbox_1.stateChanged.connect(self.on_checkbox_1_changed)
        self.checkbox_2.stateChanged.connect(self.on_checkbox_2_changed)
        self.checkbox_3.stateChanged.connect(self.on_checkbox_3_changed)
        self.checkbox_4.stateChanged.connect(self.on_checkbox_4_changed)

        self.button_project.clicked.connect(self.button_project_click)

        self.confirm_button.clicked.connect(self.button_confirm_click)
        self.search_button.clicked.connect(self.button_search_click)
        self.save_button.clicked.connect(self.button_save_click)
        self.exit_button.clicked.connect(self.app_exit)

    def clearAllTables(self):
        """TABSをクリア"""
        tableWidget = self.table
        if tableWidget is not None:
            tableWidget.setRowCount(0)
            for row in range(tableWidget.rowCount()):
                for col in range(tableWidget.columnCount()):
                    item = tableWidget.item(row, col)
                    if item is not None:
                        item.setText("")

    def row_append(self, context, color):
        tab_table = self.table
        current_row_count = tab_table.rowCount()
        tab_table.setRowCount(current_row_count + 1)

        # Fill the new row with data
        for col, line in enumerate(context):
            if line is None:
                line = ""
            item = QTableWidgetItem(f"{line}")
            item.setTextAlignment(Qt.AlignmentFlag.AlignLeft)
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            if color is not None:
                item.setBackground(color)
            if col == 3:
                font = item.font()
                font.setBold(True)
                item.setFont(font)
            tab_table.setItem(current_row_count, col, item)

        tab_table.resizeColumnsToContents()
        tab_table.resizeRowsToContents()
        QApplication.processEvents()
        # bottomRightItem = tab_table.item(current_row_count, 0)
        # tab_table.scrollToItem(bottomRightItem)
        tab_table.scrollToBottom()

    def on_checkbox_1_changed(self, state):
        """コンボボックス１"""
        if state == 2:
            self.input_keyword_1.setEnabled(True)
        else:
            self.input_keyword_1.setEnabled(False)

    def on_checkbox_2_changed(self, state):
        """コンボボックス２"""
        if state == 2:
            self.input_keyword_2.setEnabled(True)
        else:
            self.input_keyword_2.setEnabled(False)

    def on_checkbox_3_changed(self, state):
        """コンボボックス３"""
        if state == 2:
            self.input_keyword_3.setEnabled(True)
        else:
            self.input_keyword_3.setEnabled(False)

    def on_checkbox_4_changed(self, state):
        """コンボボックス４"""
        if state == 2:
            self.input_keyword_4.setEnabled(True)
        else:
            self.input_keyword_4.setEnabled(False)

    @log_and_call
    def button_project_click(self):
        """プロジェクトパス開く"""
        self.clearAllTables()
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            if self.input_project.text() is not None:
                open_path = os.path.dirname(self.input_project.text())
            folder_path = QFileDialog.getExistingDirectory(self, "プロジェクトパス選択",
                                                           directory=open_path,
                                                           options=options)
            if folder_path:
                self.input_project.setText(folder_path)
        except Exception as e:
            self.search_button.setDisabled(False)
            print("An error occurred : ", e)
            raise

    @log_and_call
    def button_confirm_click(self):
        """ユーザー確認"""
        self.setEnabled(False)
        sub_window = SubWindow(self)
        result = sub_window.exec_()  # 显示模态对话框
        self.setEnabled(True)
        # if result == QDialog.Accepted:
        #     print("✅ 用户点击了确认")
        #     self.setEnabled(True)
        # else:
        #     print("❌ 用户点击了取消")
        #     self.setEnabled(True)

    @log_and_call
    def button_search_click(self):
        """検索"""
        self.clearAllTables()
        self.save_button.setDisabled(True)
        self.search_button.setDisabled(True)
        if os.path.exists(self.input_project.text()) is False:
            set_message_box("CRITICAL", "パス", "プロジェクトパスが存在しませんが、チェックしてください。")
            return

        repo = Repo(self.input_project.text())
        tracked_files = repo.git.ls_tree('-r', '--name-only', 'HEAD').splitlines()

        # now_date_for_compare = datetime.now(datetime.timezone.utc)

        local_tz = ZoneInfo("Asia/Tokyo")
        now_local = datetime.now(local_tz)

        todo_rows = []
        summary = defaultdict(lambda: defaultdict(int))  # file -> commit -> count

        for file_types in self.input_suffix.currentText().replace(" ", "").split(";"):
            file_type = "*." + file_types
            suffixes = [file_types]
            total_files = 0
            for dirpath, dirnames, filenames in os.walk(self.input_project.text()):
                # 过滤指定后缀的文件
                filtered_files = [f for f in filenames if any(f.endswith(s) for s in suffixes)]
                total_files += len(filtered_files)
            print("total_files : ", total_files)

            count = 1
            for target_file in glob.iglob(os.path.join(self.input_project.text(), '**', file_type), recursive=True):
                rel_path = os.path.relpath(target_file, self.input_project.text()).replace("\\", "/")
                # todo for test
                # if "ena2_proto_usecase" not in rel_path:
                #     continue
                print("count : ", count)
                print("int(count / total_files)", int((count / total_files) * 100))
                # self.progress_bar.setValue(int)
                set_progress_bar(self, (count / total_files) * 100)
                # QApplication.processEvents()
                count += 1
                print(f'📄 ファイルの処理中 : {rel_path}')
                set_status_label(self, f'📄 ファイルの処理中 : {rel_path}')
                commits = list(repo.iter_commits(paths=rel_path))

                for commit in reversed(commits):  # 从旧到新
                    try:
                        blob = commit.tree / rel_path
                        commit.committed_datetime.isoformat()

                        if self.input_date.date().year() == now_local.year and \
                                self.input_date.date().month() == now_local.month and \
                                self.input_date.date().day() == now_local.day:
                            pass
                        else:
                            naive_dt = self.input_date.dateTime().toPyDateTime()
                            tz = ZoneInfo("Asia/Tokyo")  # 可替換為其他時區
                            aware_dt = naive_dt.replace(tzinfo=tz)

                            if commit.committed_datetime < aware_dt:
                                print(f"古いバージョン：{commit.hexsha}")
                                continue

                        if self.input_user.currentText().strip().split(";") == USER_TYPE:
                            pass
                        if self.input_user.currentText().strip().split(";")[0] == USER_TYPE[0]:
                            if commit.author.name not in JAPAN_USER_LIST:
                                continue
                        if self.input_user.currentText().strip().split(";")[0] == USER_TYPE[1]:
                            if commit.author.name in JAPAN_USER_LIST:
                                continue

                        content = blob.data_stream.read().decode('utf-8', errors='ignore')
                        todos = []
                        if self.checkbox_1.isChecked():
                            key_word_1 = self.input_keyword_1.text()
                        else:
                            key_word_1 = ""

                        if self.checkbox_2.isChecked():
                            key_word_2 = self.input_keyword_2.text()
                        else:
                            key_word_2 = ""

                        if self.checkbox_3.isChecked():
                            key_word_3 = self.input_keyword_3.text()
                        else:
                            key_word_3 = ""

                        if self.checkbox_4.isChecked():
                            key_word_4 = self.input_keyword_4.text()
                        else:
                            key_word_4 = ""
                        for i, line in enumerate(content.splitlines(), 1):
                            if check_substrings(line, key_word_1, key_word_2, key_word_3, key_word_4):
                                todos.append((i, line.strip()))

                        # summary[rel_path][commit.hexsha] = len(todos)

                        for line_num, full_line in todos:
                            self.row_append(
                                [rel_path,
                                 commit.hexsha,
                                 commit.committed_datetime.isoformat(),
                                 commit.author.name,
                                 line_num,
                                 full_line, "〇"],
                                WHITE)
                    except Exception as e:
                        print(f'⚠️ スキップ {rel_path} @ {commit.hexsha[:7]}: {e}')
                        # set_status_label(self, f'⚠️ スキップ : {rel_path} @ {commit.hexsha[:7]}: {e}')
                        # self.row_append(
                        #     [rel_path,
                        #      commit.hexsha,
                        #      "",
                        #      "",
                        #      "",
                        #      "", "✕"],
                        #     HERMES_ORANGE)
                        continue
        set_progress_bar(self, 100)
        self.table.scrollToTop()
        self.save_button.setDisabled(False)
        self.search_button.setDisabled(False)
        set_status_label(self, '検索完了')
        set_message_box("INFO", "検索", "\n\n" + "※※※検索完了。※※※" + "\n\n")

    @log_and_call
    def button_save_click(self):
        """出力"""
        self.window().setDisabled(True)
        table_data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for column in range(self.table.columnCount()):
                item = self.table.item(row, column)
                row_data.append(item.text())
            if row_data[6] == "〇":
                table_data.append(row_data)
        print(table_data)
        if len(table_data) == 0:
            set_message_box("CRITICAL", "生成", "\n\n" + "※※※生成可能なデータがありません。※※※" + "\n\n")
            self.window().setDisabled(False)
            return

        # 创建工作簿
        wb = openpyxl.Workbook()

        # 获取当前时间并格式化
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

        # 删除默认的 "Sheet"
        default_sheet = wb.active
        wb.remove(default_sheet)

        # ws = wb.create_sheet(title=f"結果{timestamp}")

        for row in table_data:
            project_name = row[0].split("/")[1]
            if project_name not in wb.sheetnames:
                ws = wb.create_sheet(title=project_name)
                # cell = ws.cell(row=1, column=1, value="キーワード")
                # cell.alignment = Alignment(horizontal="center", vertical="center")
                # cell.font = Font(bold=True)
                #
                # cell = ws.cell(row=1, column=2, value=self.input_keyword_1.text())
                # cell.alignment = Alignment(horizontal="center", vertical="center")
                # cell.font = Font(bold=True)
                #
                # cell = ws.cell(row=1, column=3, value=self.input_keyword_2.text())
                # cell.alignment = Alignment(horizontal="center", vertical="center")
                # cell.font = Font(bold=True)
                #
                # cell = ws.cell(row=1, column=4, value=self.input_keyword_3.text())
                # cell.alignment = Alignment(horizontal="center", vertical="center")
                # cell.font = Font(bold=True)
                #
                # cell = ws.cell(row=1, column=5, value=self.input_keyword_4.text())
                # cell.alignment = Alignment(horizontal="center", vertical="center")
                # cell.font = Font(bold=True)

                col_count = self.table.columnCount()
                for col in range(col_count):
                    header_item = self.table.horizontalHeaderItem(col)
                    cell = ws.cell(row=1, column=col + 1, value=header_item.text() if header_item else "")
                    cell.fill = PatternFill(start_color="92C9EA", end_color="92C9EA", fill_type="solid")  # 浅蓝色背景
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(bold=True)
            else:
                ws = wb[project_name]

            # 插入数据（这里假设是一行数据）
            ws.append(row)

        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        for sheet in wb.worksheets:
            # 遍历所有单元格
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = border_style

        for sheet in wb.worksheets:
            # 遍历每一列
            for col_idx in range(1, sheet.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)  # 将列索引转为字母（A, B, C...）

                # 遍历该列的所有单元格
                for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=sheet.max_row):
                    cell = row[0]
                    if cell.value is not None:
                        current_length = len(str(cell.value))
                        if current_length > max_length:
                            max_length = current_length

                # 设置列宽（Excel 列宽单位 ≈ 字符长度 × 1.2）
                adjusted_width = max_length + 2  # 添加额外空间
                sheet.column_dimensions[column_letter].width = adjusted_width

        output_file = os.path.join(get_program_path(), f"出力結果_"
                                                       f"{self.input_suffix.currentText()}_"
                                                       f"{self.input_user.currentText()}_"
                                                       f"{self.input_date.date().toString('yyyyMMdd')}_"
                                                       f"{timestamp}.xlsx")
        wb.save(output_file)
        wb.close()
        del wb
        gc.collect()

        self.window().setDisabled(False)
        set_message_box("INFO", "生成", "\n\n" + f"※※※生成完了。※※※\n"
                                                 f"{output_file}" + "\n\n")

    @log_and_call
    def app_exit(self):
        """アプリを退出"""
        msg_box = QMessageBox()
        msg_box.setWindowTitle("ツールメッセージ")
        msg_box.setText("\n\n"
                        "※※※ツールを終了したいですか。※※※"
                        "\n\n")
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg_box.setDefaultButton(QMessageBox.StandardButton.No)
        msg_box.button(QMessageBox.StandardButton.Yes).setText("はい(&Y)")
        msg_box.button(QMessageBox.StandardButton.No).setText("いいえ(&N)")
        result = msg_box.exec()
        if result == QMessageBox.StandardButton.Yes:
            save_file_paths(self)
            self.close()

    def closeEvent(self, event):
        """closeEvent"""
        config = ConfigParser()
        config_path = get_config_file_path()
        if os.path.exists(config_path) is False:
            return
        config.read(config_path, encoding='utf-8')
        if self.input_suffix.currentText():
            config.set('Ids', 'input_suffix', self.input_suffix.currentText())
        if self.input_user.currentText():
            config.set('Ids', 'input_user', self.input_user.currentText())
        if self.input_date.date():
            config.set('Ids', 'input_date', self.input_date.date().toString("yyyy-MM-dd"))
        if self.input_project.text():
            config.set('Paths', 'input_project', self.input_project.text())

        if self.checkbox_1:
            result = self.checkbox_1.isChecked()
            config.set('Ids', 'checkbox_1', str(result))
        if self.input_keyword_1.text():
            config.set('Ids', 'input_keyword_1', self.input_keyword_1.text())
        if self.checkbox_2:
            result = self.checkbox_2.isChecked()
            config.set('Ids', 'checkbox_2', str(result))
        if self.input_keyword_2.text():
            config.set('Ids', 'input_keyword_2', self.input_keyword_2.text())
        if self.checkbox_3:
            result = self.checkbox_3.isChecked()
            config.set('Ids', 'checkbox_3', str(result))
        if self.input_keyword_3.text():
            config.set('Ids', 'input_keyword_3', self.input_keyword_3.text())
        if self.checkbox_4:
            result = self.checkbox_4.isChecked()
            config.set('Ids', 'checkbox_4', str(result))
        if self.input_keyword_4.text():
            config.set('Ids', 'input_keyword_4', self.input_keyword_4.text())
        if JAPAN_USER_LIST:
            config.set('Users', 'japan_user_list', ",".join(map(str, JAPAN_USER_LIST)))
        with open(get_config_file_path(), 'w', encoding='utf-8') as configfile:
            config.write(configfile)

    def timer_init(self):
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_datetime)
        self.timer.start(1000)

    def update_datetime(self):
        self.current_datetime = QDateTime.currentDateTime().toString(Qt.DateFormat.ISODate)
        self.tips_label.setText(f'{self.current_datetime}')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Windows')  # Windows , windowsvista , Fusion

    # spinner_chars = ['|', '/', '-', '\\']
    # splash = CharacterSplashScreen(spinner_chars)
    # splash.setWindowFlags(Qt.SplashScreen | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
    # splash.show()

    if check_git_installed() is False:
        sys.exit(app.exec_())
    window = MainWindow()
    window.show()

    # splash.finish(window)
    sys.exit(app.exec_())
