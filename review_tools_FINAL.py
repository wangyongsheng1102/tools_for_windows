import datetime
import functools
import gc
import glob
import hashlib
import os
import random
import re
import shutil
import subprocess
import sys
import xml.dom.minidom as xmldom
import xml.etree.ElementTree as ET
import zipfile
from configparser import ConfigParser
from datetime import datetime as dt

import cv2
import imutils
import numpy as np
import openpyxl
from PyQt5 import QtWidgets
from PyQt5.QtCore import Qt, QRect, QTimer, QDateTime, pyqtSignal, QObject
from PyQt5.QtGui import QPainter, QColor, QBrush, QFont, QPen, QPixmap
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTabWidget, QTableWidget, QTableWidgetItem, \
    QProgressBar, QLabel, QHBoxLayout, QGroupBox, QMainWindow, QLineEdit, QComboBox, QSizePolicy, QMessageBox, \
    QFormLayout, QFileDialog, QHeaderView, QDialog, QCheckBox, QSplashScreen
from bs4 import BeautifulSoup
from openpyxl.drawing.image import Image as ImageExcel
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.workbook import Workbook
from skimage.metrics import structural_similarity as compare_ssim

"""修正記録"""
"""２０２４１０３０：
    ①：複数シートのキャプチャー比較功能を完全になる
    ②：カバーチェック異常の修正
    ③：手修正のソースとGITのソースとの比較ロジック修正
"""
"""２０２４１０３１：
    ①：手修正には、最大行目を取得する方法を修正
"""
"""２０２４１１０４：
    ①：手修正には、ソース行目数が不一致の場合、ロジックを修正
"""
"""２０２４１１０６：
    ①：手修正のソースとGITのソースとの比較ロジック修正
"""
"""２０２４１１０７：
    ①：成果物一覧に手修正ありのものを、手修正確認結果ファイルにチェックするのを追加しました。
"""
"""２０２４１１０８：
    ①：ＷＢＳに情報を取得方法を修正
"""
"""２０２４１１１２：
    ①：成果物ファイルを探す方法を修正
"""
"""２０２４１１１３：
    ①：エラー発生時、実行ボタン解放することを修正
"""
"""２０２４１１１５：
    ①：手修正ファイルと成果物一覧に間違っている場合、判断を追加
"""
"""２０２４１１１８：
    ①：ソースチェックには、元ソースがない場合、修正しました。
"""
"""２０２４１２０２：
    ①：キャプチャーを比較する際に、スクロールバーの差異を無視する比較ロジックを修正しました。
"""
"""２０２４１２０７：
    ①：キャプチャーを比較する際に、ピクセルが二以下の不比較の判断を削除しました。
"""
"""２０２４１２１１：
    ①：成果物ファイルを探している際に、システムを条件で検索することを修正しました。
"""
"""２０２４１２１４：
    ①：キャプチャー比較ファイルには、差異点数を明記することを追加しました。
"""
"""２０２４１２１９：
    ①：エビデンスチェックをするかどうがのラジオを追加しました。
"""
"""２０２５０２１４：
    ①：書類記入チェック
"""
"""２０２５０２１７：
    ①：レビュー率の判断を修正
    ②：APIのカバーチェックを修正
    ③：手修正確認結果にはJAVAファイル以外の記入内容の判断を修正
"""
"""２０２５０２１８：
    ①：チェック対象のチェックボックス追加
    ②：手修正ファイルからソースを取得方法修正
"""
"""２０２５０２１９：
    ①：SVNチェックによって、成果物一覧のパス異常を修正
"""
"""２０２５０２２７：
    ①：補足書類のチェック
    ②：カバーのソースとGITのソースを比較方法を修正
    ③：結合テストの特別処理
"""
"""２０２５０２２８：
    ①：界面アップデート
    ②：補足ファイルの探すことを修正
    ③：カバー対象のミス、Customの判断を追加
"""
"""２０２５０３２０：
    ①：結合テストの場合、エビデンスチェックだけのため、類型を追加         「却下」
"""

VERSION_INFO = '20250320'

ENABILITY_TEST_TYPE = [
    'ハイ', 'イイエ'
]

ENABILITY_TYPE = [
    '画面', 'バッチ', 'API', '共通部品'
]

ENABILITY_SYSTEM = [
    'EnabilityCIS',
    'EnabilityOrder',
    'EnabilityPortal',
    'EnabilityPortal2'
]

ENABILITY_OBJECT = [
    'API',
    'バッチ',
    'マルチ',
    '課題対応',
    '画面',
    '環境構築',
    '共通部品',
    '結合テスト',
    '差分結合',
    '差分取込',
    '性能テスト'
]

INPUT_ID, \
    INPUT_TYPE, \
    INPUT_SYSTEM, \
    INPUT_SVN, \
    INPUT_MANUAL, \
    INPUT_OLD, \
    INPUT_NEW, \
    INPUT_WBS, \
    INPUT_ITEST, \
    INPUT_BRANCH, \
    INPUT_OBJECT = "", "", "", "", "", "", "", "", "", "", ""

EXCEL_TEST = "単体テスト仕様書"
EXCEL_EVIDENCE = "単体テストエビデンス"
EXCEL_LIST = "成果物一覧"
EXCEL_COMPARE = "手修正確認結果"
EXCEL_COVERAGE = "カバレッジ結果"
EXCEL_REVIEW = "レビュー記録表"
EXCEL_CD_CHECKLIST = "CDチェックリスト"
EXCEL_UT_CHECKLIST = "UTチェックリスト"

variables = {
    'EXCEL_TEST': EXCEL_TEST,
    'EXCEL_EVIDENCE': EXCEL_EVIDENCE,
    'EXCEL_LIST': EXCEL_LIST,
    'EXCEL_COMPARE': EXCEL_COMPARE,
    'EXCEL_COVERAGE': EXCEL_COVERAGE,
    'EXCEL_REVIEW': EXCEL_REVIEW,
    'EXCEL_CD_CHECKLIST': EXCEL_CD_CHECKLIST,
    'EXCEL_UT_CHECKLIST': EXCEL_UT_CHECKLIST
}

EXCEL_TOTAL_MAP = {"EXCEL_TEST": EXCEL_TEST, "EXCEL_EVIDENCE": EXCEL_EVIDENCE, "EXCEL_LIST": EXCEL_LIST,
                   "EXCEL_COMPARE": EXCEL_COMPARE, "EXCEL_COVERAGE": EXCEL_COVERAGE, "EXCEL_REVIEW": EXCEL_REVIEW,
                   "EXCEL_CD_CHECKLIST": EXCEL_CD_CHECKLIST, "EXCEL_UT_CHECKLIST": EXCEL_UT_CHECKLIST}

COMMON_FUNC_ID = ""
COMMON_FUNC_NAME = ""
COMMON_SYSTEM = ""
COMMON_TYPE = ""
COMMON_USER_NAME = ""
COMMON_USER_DATE = ""

WBS_NAME = ""
WBS_DATE = ""

WBS_REVIEW_USER_NAME = ""
WBS_REVIEW_USER_DATE = ""

COL, ROW = 1, 1

# # DHC fengjm 2024/12/20 No.radio ADD START
# RADIO_FLAG = 0
# # DHC fengjm 2024/12/20 No.radio ADD END
#
# # DHC fengjm 2024/12/20 No.itest ADD START
# ITEST_FLAG = False
# # DHC fengjm 2024/12/20 No.itest ADD END

CHECKBOX_DOC = True
CHECKBOX_CONTEXT = False
CHECKBOX_SOURCE = False
CHECKBOX_MODIFY = False
CHECKBOX_COVERAGE = False
CHECKBOX_PIC = False

KLEIN_BLUE = QColor(0, 47, 167)
TIFFANY_BLUE = QColor(131, 216, 208)

SCHENBRUNN_YELLOW = QColor(247, 225, 77)
MARS_GREEN = QColor(0, 140, 140)

TITIAN_RED = QColor(176, 89, 35)
CHINA_RED = QColor(230, 0, 0)

HERMES_ORANGE = QColor(232, 88, 39)


def log_and_call(func):
    @functools.wraps(func)
    def wrapper(self, *args, **kwargs):
        set_status_label(self, func.__doc__ + "中")
        return func(self, *args, **kwargs)

    return wrapper


def get_program_path():
    """アプリのパスを取得"""
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def get_config_file_path():
    """コンフィグのパスを取得"""
    return os.path.join(get_program_path(), ".review_config.ini")


def load_config_content(tag):
    """コンフィグをロード"""
    config = ConfigParser()
    config_path = get_config_file_path()
    if os.path.exists(config_path) is False:
        raise
    config.read(config_path, encoding='utf-8')
    return config[tag] if tag in config else {}


def init_config_content():
    global INPUT_ID, INPUT_TYPE, INPUT_SYSTEM, INPUT_SVN, INPUT_MANUAL, \
        INPUT_OLD, INPUT_NEW, INPUT_WBS, INPUT_ITEST, INPUT_BRANCH, INPUT_OBJECT, \
        CHECKBOX_DOC, CHECKBOX_CONTEXT, CHECKBOX_SOURCE, CHECKBOX_MODIFY, CHECKBOX_COVERAGE, CHECKBOX_PIC
    try:
        ids = load_config_content('Ids')
        paths = load_config_content('Paths')
    except Exception as e:
        set_message_box("CRITICAL", "コンフィグ", "コンフィグファイルが存在しませんが、チェックしてください。")
        return
    if ids['input_id'] is not None:
        INPUT_ID = ids['input_id']
    if ids['input_type'] is not None:
        INPUT_TYPE = ids['input_type']
    if ids['input_system'] is not None:
        INPUT_SYSTEM = ids['input_system']
    if paths['input_svn'] is not None:
        INPUT_SVN = paths['input_svn']
    if paths['input_manual'] is not None:
        INPUT_MANUAL = paths['input_manual']
    if paths['input_old'] is not None:
        INPUT_OLD = paths['input_old']
    if paths['input_new'] is not None:
        INPUT_NEW = paths['input_new']
    if paths['input_wbs'] is not None:
        INPUT_WBS = paths['input_wbs']
    # DHC fengjm 2024/12/20 No.itest ADD START
    if paths['input_itest'] is not None:
        INPUT_ITEST = paths['input_itest']
    # DHC fengjm 2024/12/20 No.itest ADD END
    if paths['input_branch'] is not None:
        INPUT_BRANCH = paths['input_branch']
    if ids['input_object'] is not None:
        INPUT_OBJECT = ids['input_object']
    if ids['checkbox_doc'] is not None:
        CHECKBOX_DOC = eval(ids['checkbox_doc'])
    if ids['checkbox_context'] is not None:
        CHECKBOX_CONTEXT = eval(ids['checkbox_context'])
    if ids['checkbox_source'] is not None:
        CHECKBOX_SOURCE = eval(ids['checkbox_source'])
    if ids['checkbox_modify'] is not None:
        CHECKBOX_MODIFY = eval(ids['checkbox_modify'])
    if ids['checkbox_coverage'] is not None:
        CHECKBOX_COVERAGE = eval(ids['checkbox_coverage'])
    if ids['checkbox_pic'] is not None:
        CHECKBOX_PIC = eval(ids['checkbox_pic'])


def save_file_paths(self):
    """コンフィグにパスインフォを保存"""
    config = ConfigParser()
    config_path = get_config_file_path()
    if os.path.exists(config_path) is False:
        return
    config.read(config_path, encoding='utf-8')
    if self.parent.input_id.text():
        config.set('Ids', 'input_id', self.parent.input_id.text())
    if self.parent.input_type.currentText():
        config.set('Ids', 'input_type', self.parent.input_type.currentText())
    if self.parent.input_system.currentText():
        config.set('Ids', 'input_system', self.parent.input_system.currentText())
    if self.parent.input_svn.text():
        config.set('Paths', 'input_svn', self.parent.input_svn.text())
    if self.parent.input_manual.text():
        config.set('Paths', 'input_manual', self.parent.input_manual.text())
    if self.parent.input_old.text():
        config.set('Paths', 'input_old', self.parent.input_old.text())
    if self.parent.input_new.text():
        config.set('Paths', 'input_new', self.parent.input_new.text())
    if self.parent.input_wbs.text():
        config.set('Paths', 'input_wbs', self.parent.input_wbs.text())
    # DHC fengjm 2024/12/20 No.itest ADD START
    if self.parent.input_itest.text():
        config.set('Paths', 'input_itest', self.parent.input_itest.text())
    # DHC fengjm 2024/12/20 No.itest ADD END
    if self.parent.input_branch.text():
        config.set('Paths', 'input_branch', self.parent.input_branch.text())
    if self.parent.input_object.currentText():
        config.set('Ids', 'input_object', self.parent.input_object.currentText())
    if self.parent.checkbox_doc:
        result = self.parent.checkbox_doc.isChecked()
        config.set('Ids', 'checkbox_doc', str(result))
    if self.parent.checkbox_context:
        result = self.parent.checkbox_context.isChecked()
        config.set('Ids', 'checkbox_context', str(result))
    if self.parent.checkbox_source:
        result = self.parent.checkbox_source.isChecked()
        config.set('Ids', 'checkbox_source', str(result))
    if self.parent.checkbox_modify:
        result = self.parent.checkbox_modify.isChecked()
        config.set('Ids', 'checkbox_modify', str(result))
    if self.parent.checkbox_coverage:
        result = self.parent.checkbox_coverage.isChecked()
        config.set('Ids', 'checkbox_coverage', str(result))
    if self.parent.checkbox_pic:
        result = self.parent.checkbox_pic.isChecked()
        config.set('Ids', 'checkbox_pic', str(result))
    with open(get_config_file_path(), 'w', encoding='utf-8') as configfile:
        config.write(configfile)


def set_message_box(message_type, title, context):
    """メッセージを反映"""
    if message_type == 'WARNING':
        QMessageBox.warning(None, title, context)
    if message_type == 'CRITICAL':
        QMessageBox.critical(None, title, context)
    if message_type == 'INFO':
        QMessageBox.information(None, title, context)
    if message_type == 'QUESTION':
        QMessageBox.question(None, title, context)


def is_null_check(self):
    """非空チェック"""
    check_flag = False
    context = ""
    if self.parent.input_id.text() is None or self.parent.input_id.text() == '':
        check_flag = True
        context = context + "機能ID\n"
    if self.parent.input_type.currentText() is None or self.parent.input_type.currentText() == '':
        check_flag = True
        context = context + "区分\n"
    if self.parent.input_system.currentText() is None or self.parent.input_system.currentText() == '':
        check_flag = True
        context = context + "システム\n"
    if self.parent.input_svn.text() is None or self.parent.input_svn.text() == '':
        check_flag = True
        context = context + "SVNパス\n"
    if self.parent.input_manual.text() is None or self.parent.input_manual.text() == '':
        check_flag = True
        context = context + "マニュアルパス\n"
    if self.parent.input_old.text() is None or self.parent.input_old.text() == '':
        check_flag = True
        context = context + "元ソース\n"
    if self.parent.input_new.text() is None or self.parent.input_new.text() == '':
        check_flag = True
        context = context + "新ソース\n"
    if self.parent.input_wbs.text() is None or self.parent.input_wbs.text() == '':
        check_flag = True
        context = context + "WBSパス\n"
    if self.parent.input_branch.text() is None or self.parent.input_branch.text() == '':
        check_flag = True
        context = context + "GITブランチ\n"
    if self.parent.input_object.currentText() is None or self.parent.input_object.currentText() == '':
        check_flag = True
        context = context + "対象区分\n"
    return check_flag, context


def set_status_label(self, context):
    self.parent.status_label.setText(context)


def clearAllTables(self):
    """TABSをクリア"""
    for index in range(self.parent.tabs.count()):
        tabWidget = self.parent.tabs.widget(index)
        if tabWidget is not None:
            tableWidget = tabWidget.findChild(QTableWidget)
            if tableWidget is not None:
                tableWidget.setRowCount(0)
                for row in range(tableWidget.rowCount()):
                    for col in range(tableWidget.columnCount()):
                        item = tableWidget.item(row, col)
                        if item is not None:
                            item.setText("")


def copy_folder(src_folder, dest_folder):
    try:
        if not os.path.exists(dest_folder):
            os.makedirs(dest_folder)

        for item in os.listdir(src_folder):
            item_path = os.path.join(src_folder, item)
            dest_path = os.path.join(dest_folder, item)

            if os.path.isdir(item_path):
                copy_folder(item_path, dest_path)
            else:
                shutil.copy2(item_path, dest_path)

        print(f"Folder '{src_folder}' copied to '{dest_folder}' successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")


def read_excel_list(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    try:
        ws = wb[f'{sheet_name}']
    except KeyError:
        return
    try:
        source_files = []
        for row in ws.iter_rows():
            if row[1].value is None or row[1].value == "" or row[1].value == "NO":
                continue
            source_files.append([row[2].value, row[3].value])
    finally:
        wb.close()
        del wb
        gc.collect()

    return source_files


def download_source_from_git(file_path, destination_path):
    server_url = 'git@192.168.70.194:root/bip.git'
    branch_name = INPUT_BRANCH
    command = f'git archive --remote={server_url} --format=tar {branch_name} {file_path} | tar -xO'
    command = command.replace("\\", "/")
    process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate()
    if process.returncode != 0:
        print(f"Error downloading file: {stderr.decode('utf-8')}")
        return False
    else:
        try:
            with open(destination_path, 'wb') as f:
                f.write(stdout)
            print(f"File '{file_path}' from branch '{branch_name}' downloaded to '{destination_path}'")
            return True
        except OSError as e:
            print(f"Write '{destination_path}' error：{e.strerror}")
            return False


def svn_version_check():
    print('svn_version_check start...')
    # cmd_update = 'svn update ' + self.parent.input_browse.text()
    # result = os.system(cmd_update)
    # print("svn update result : ", result)
    result = subprocess.run(['svn', '--version'], text=True, capture_output=True, check=False
                            , shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
    if "svn, version" in result.stdout:
        pass
    else:
        set_message_box("WARNING", "SVN",
                        "コンピューターにはまだSVNコマンドラインがインストールされていません。\nインストールしてください。")


def svn_check_file(file_path):
    """SVNチェック"""
    print('svn_check_file start...')
    # "\"" + file_path + "\""
    result = subprocess.run(['svn', 'status', file_path], text=True, capture_output=True, check=False
                            , shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
    if result.returncode == 0:
        version_result = svn_check_file_version(file_path)
        if version_result is True:
            return ""
        else:
            return "本地のバージョンは最新のではありませんので、チェックしてください。"
    else:
        return "SVNにコミットするかどうかのことをチェックしてください。"


def svn_check_file_version(file_path):
    """SVNバージョンチェック"""
    print('svn_check_file_version start...')
    # "\"" + file_path + "\""
    result = subprocess.run(['svn', 'status', '-u', file_path], text=True, capture_output=True, check=False
                            , shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
    if result.returncode == 0:
        return True
    else:
        return False


def svn_operate(self, folder):
    """SVNから更新"""
    print('svn_operate start...')
    # cmd_update = 'svn update ' + self.parent.input_browse.text()
    # result = os.system(cmd_update)
    # print("svn update result : ", result)
    result = subprocess.run(['svn', '--version'], text=True, capture_output=True, check=False
                            , shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
    if "svn, version" in result.stdout:
        with subprocess.Popen(['svn', 'update', folder],
                              stdin=subprocess.PIPE,
                              stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE, text=True) as proc:
            stdout, stderr = proc.communicate()
        if proc.returncode != 0:
            # set_message_box("CRITICAL", "SVN",
            #                 "サンプルフォルダをSVNから最新版に更新することが失敗しました、\n自分で更新してください。")
            print(f"Command '{self.parent.input_browse.text()}' "
                  f"failed with return code {proc.returncode}")
            print("Errors:", stderr)
        else:
            # set_message_box("INFO", "SVN", "「" + folder.split("\\")[len(folder.split("\\")) - 1] + "」" +
            #                 "\nSVNから更新することが成功しました、\n続けてください。")
            print(f"Command '{self.parent.input_browse.text()}' executed successfully")
            print("Output:", stdout)
    else:
        set_message_box("WARNING", "SVN",
                        "コンピューターにはまだSVNコマンドラインがインストールされていません。\nインストールしてください。")


def file_md5(filename):
    """md5のメソッドを実行"""
    hash_md5 = hashlib.md5()
    with open(filename, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()


def files_are_identical(file1, file2):
    """md5でファイルを比較"""
    return file_md5(file1) == file_md5(file2)


def find_files(folder_path_to_find, word, suffix, system):
    """ファイルを探す"""
    print("find_files :" + suffix)
    word = "(" + word + "_"
    if system is None:
        system = ""
    for root, dirs, files in os.walk(folder_path_to_find):
        for file_con in files:
            if file_con.find(word) >= 0 and file_con.find(suffix) >= 0 and (
                    root.find(system + "\\") > 0 or root.find("\\" + system) > 0) \
                    and root.find("bk\\") < 0:
                # if file_con.find("C_") >= 0:
                #     if word[-1] == 'C':
                #         return os.path.join(root, file_con)
                #     else:
                #         continue
                if (file_con.find("Custom_") >= 0) == (word.find("Custom_") >= 0) \
                        and (file_con.find("_SACM") >= 0) == (word.find("_SACM") >= 0) \
                        and (file_con.find("_SUCM") >= 0) == (word.find("_SUCM") >= 0):
                    return os.path.join(root, file_con)
                else:
                    continue

    return None


def extract_zip_files(zip_dir, extract_dir):
    for dir_path, dir_names, file_names in os.walk(zip_dir):
        for file_name in file_names:
            if file_name.endswith('.zip'):
                zip_path = os.path.join(dir_path, file_name)
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_dir)


def fuzzy_search(directory, pattern):
    matched_files = []
    for dirpath, dirnames, filenames in os.walk(directory):
        for file in glob.iglob(os.path.join(dirpath, pattern), recursive=True):
            matched_files.append(file)
    return matched_files


def find_api_source_in_html(line, source_list):
    for java_file in source_list:
        if line.find(os.path.basename(java_file[2]).replace(".java", "")) >= 0 \
                and line.find("Custom_") == java_file[2].find("Custom_") \
                and line.find("_SACM") == java_file[2].find("_SACM"):
            return True
    return False


def api_coverage_search(directory, source_list):
    matched_files = []
    for index_html in glob.iglob(os.path.join(directory, '**', 'index.html'), recursive=True):
        with open(index_html, 'r') as file:
            lines = file.readlines()
        for index, line in enumerate(lines):
            if find_api_source_in_html(line, source_list):
                soup = BeautifulSoup(line, 'html.parser')
                if soup.find('td', class_='name') is not None:
                    a_tag = soup.find('td', class_='name').find('a')
                    if a_tag is not None:
                        href_value = a_tag['href']
                        text_value = a_tag.get_text(strip=True)
                        print("html_line :", href_value, "text_value :", text_value)
                        if os.path.join(os.path.dirname(index_html),
                                        href_value) + "|=" + text_value not in matched_files:
                            matched_files.append(
                                os.path.join(os.path.dirname(index_html), href_value) + "|=" + text_value)

    return matched_files


def find_source_files(folder_path_to_find, word):
    for root, dirs, files in os.walk(folder_path_to_find):
        for file_con in files:
            if file_con == word:
                return os.path.join(root, file_con)
    return None


def find_files_in_svn(svn_path, func_id, suffix, flag, system):
    """SVNにファイルを探す"""
    excel_path = find_files(svn_path, func_id, suffix, system)
    if excel_path is None and flag is True:
        set_message_box("CRITICAL", "ファイル", "機能ID「" + func_id + "」の成果物が見つかりませんので、チェックしてください。")
        return None
    else:
        return excel_path


def find_coverage_in_folder(svn_path, system):
    if system is None:
        system = ""
    for root, dirs, files in os.walk(svn_path):
        for file_con in files:
            if file_con.find(EXCEL_COVERAGE) >= 0 \
                    and root.find(system + "\\") > 0 \
                    and root.find("bk\\") < 0:
                return os.path.join(root, file_con)
            else:
                continue

    return None


def read_name_excel(path, sheet):
    """外部WBSを読む"""
    sw = load_workbook(f'{path}', data_only=True)

    try:
        src_sheet = sw[f'{sheet}']
    except KeyError:
        raise KeyError('シートが存在しません。')
    try:
        source_list = []
        for row in src_sheet.iter_rows():
            source_list_list = []
            # if row[2].value != "共通部品":
            #     continue
            for cell in row:
                source_list_list.append(cell.value)
            source_list.append(source_list_list)
    finally:
        sw.close()
        del sw
        gc.collect()
    return source_list


def column_letter_to_number(column_letter):
    """ディジットに変更する"""
    column_number = 0
    for char in column_letter:
        column_number = column_number * 26 + (ord(char) - ord("A") + 1)
    return column_number


def get_str_before_first_dot(string, split):
    return string.split(split)[0]


def get_str_after_last_dot(string, split):
    return string.split(split)[len(string.split(split)) - 1]


def generate_random_color(color):
    if color == "blue":
        blue = random.randint(150, 255)
        red = random.randint(0, blue - 50)
        green = random.randint(0, blue - 50)
    if color == "red":
        red = random.randint(200, 255)
        green = random.randint(0, 100)
        blue = random.randint(0, 100)
    if color == "yellow":
        red = random.randint(200, 255)
        green = random.randint(200, 255)
        blue = random.randint(0, 100)
    return red, green, blue


def del_folder(folder_path):
    try:
        shutil.rmtree(folder_path)
    except OSError as e:
        print(f"while deleting '{folder_path}' error：{e.strerror}")


def count_lines_of_code(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
    return len(lines)


def read_coverage_html(file_path):
    result = {}
    start_flag = False
    end_flag = False
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
    for i, row in enumerate(lines):
        print("COVERAGE TEST LINE NO : ", i, ":", row)
        if row.find("*") >= 0 or (row.find("public") > 0 and row.find("{") > 0) or \
                row == "\n":
            continue
        if row.find("DHC") >= 0 > row.find("START") >= row.find("END"):
            print("※" * 10, ":", row)
            result[i + 1] = "ソースコメント誤り"
            continue
        if row.find("DHC") >= 0 > row.find("END") >= row.find("START"):
            print("※" * 10, ":", row)
            result[i + 1] = "ソースコメント誤り"
            continue
        if row.find("DHC") >= 0 > row.find("DEL") and row.find("START") >= 0:
            print("※" * 10, ":", row)
            start_flag = True
        if row.find("DHC") >= 0 and row.find("END") >= 0:
            print("※" * 10, ":", row)
            start_flag = False
        if start_flag is True and row.find("class=\"fc") < 0 and \
                row.find("START") < 0 and row.find("END") < 0 and row.find("//") < 0:
            print("※" * 10, ":", row)
            result[i + 1] = "カバーしない"
    return result


def read_coverage_html_from_api(file_path):
    result = {}
    start_flag = False
    end_flag = False
    with open(file_path, 'r', encoding='shift_jis', errors='ignore') as file:
        lines = file.readlines()
    for i, row in enumerate(lines):
        # print("COVERAGE TEST LINE NO : ", i, ":", row)
        if row.find("*") >= 0 or (row.find("public") > 0 and row.find("{") > 0) or \
                row == "\n":
            continue
        if row.find("DHC") >= 0 > row.find("START") >= row.find("END"):
            print("※" * 10, ":", row)
            result[i + 1] = "ソースコメント誤り"
            continue
        if row.find("DHC") >= 0 > row.find("END") >= row.find("START"):
            print("※" * 10, ":", row)
            result[i + 1] = "ソースコメント誤り"
            continue
        if row.find("DHC") >= 0 > row.find("DEL") and row.find("START") >= 0:
            print("※" * 10, ":", row)
            start_flag = True
        if row.find("DHC") >= 0 and row.find("END") >= 0:
            print("※" * 10, ":", row)
            start_flag = False
        if start_flag is True and row.find("b class") >= 0 and row.find("class=\"fc\"") < 0 and \
                row.find("START") < 0 and row.find("END") < 0 and row.find("//") < 0:
            print("※" * 10, ":", row)
            result[i + 1] = "カバーしない"
    return result


def count_max_line_from_excel(excel_path, sheet_name, col):
    wb = load_workbook(excel_path)
    try:
        sheet = wb[sheet_name]
        column = col
        max_value = None

        for row in sheet.iter_rows(min_row=2,
                                   min_col=sheet[column][0].column,
                                   max_col=sheet[column][0].column,
                                   values_only=True):
            for cell in row:
                if cell is None:
                    continue
                if max_value is None or int(cell) > int(max_value):
                    max_value = cell
    finally:
        wb.close()
        del wb
        gc.collect()

    return max_value


def return_new_lines_from_excel(excel_path, sheet_name, col):
    wb = load_workbook(excel_path)
    try:
        sheet = wb[sheet_name]
        column = col

        result_lines = []
        for row in sheet.iter_rows(min_row=3,
                                   # min_col=sheet[column][0].column,
                                   # max_col=sheet[column][0].column,
                                   values_only=True):
            if row[4] is not None and isinstance(row[4], int):
                result_lines.append(row[5])

    finally:
        wb.close()
        del wb
        gc.collect()

    return result_lines


def read_sheet_names(file_path):
    sheet_names = {}
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_names[ws['A1'].value] = sheet
        print(f"Sheet: {sheet}, A1 Value: {ws['A1'].value}")
    return sheet_names


class HyperlinkTableWidgetItem(QTableWidgetItem):
    def __init__(self, text, link):
        super().__init__(text, QTableWidgetItem.UserType)
        self.link = link

    def __lt__(self, other):
        return self.text() < other.text()


class TableWidget(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setColumnCount(4)

    def keyPressEvent(self, event):
        if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_C:
            self.copySelectedCells()
        else:
            super().keyPressEvent(event)

    def copySelectedCells(self):
        # 获取选中的单元格范围
        selected_range = self.selectedRanges()
        clipboard_text = ""

        for ranges in selected_range:
            start_row = ranges.topRow()
            end_row = ranges.bottomRow()
            start_col = ranges.leftColumn()
            end_col = ranges.rightColumn()

            for row in range(start_row, end_row + 1):
                row_text = []
                for col in range(start_col, end_col + 1):
                    item = self.item(row, col)
                    row_text.append(item.text() if item else "")

                for context in row_text:
                    clipboard_text += "\t" + context.replace("\n", "")
                clipboard_text += "\n"
                # clipboard_text += "\t|".join(row_text.replace("\n", "")) + "\n"

        clipboard = QApplication.clipboard()
        clipboard.setText(clipboard_text)


class StepProgressBar(QWidget):
    def __init__(self, steps=5):
        super().__init__()
        self.steps = steps
        self.current_step = 0
        self.setMinimumHeight(40)
        self.setMinimumWidth(300)
        self.setMouseTracking(True)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        width = self.width()
        height = self.height()
        step_width = width // self.steps
        radius = 8

        for i in range(self.steps):
            if i < self.current_step:
                brush_color = QColor(100, 149, 237)
            else:
                brush_color = QColor(211, 211, 211)

            pen = QPen(Qt.NoPen)
            painter.setPen(pen)
            painter.setBrush(QBrush(brush_color))

            painter.drawRoundedRect(QRect(i * step_width, 0, step_width - 5, height), radius, radius)

            painter.setPen(QColor(255, 255, 255))
            font = QFont("Arial", 12, QFont.Bold)
            painter.setFont(font)
            painter.drawText(QRect(i * step_width, 0, step_width - 5, height), Qt.AlignCenter, f"{i + 1}")

    def mousePressEvent(self, event):
        width = self.width()
        step_width = width // self.steps
        clicked_step = event.x() // step_width
        if clicked_step < self.steps:
            self.current_step = clicked_step + 1
            self.update()
            self.parent().tab_switched(clicked_step)

    def advance_step(self):
        if self.current_step < self.steps:
            self.current_step += 1
            self.update()
            self.parent().tab_switched(self.current_step - 1)


class SortableTable(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setColumnCount(4)
        self.setHorizontalHeaderLabels(['区分', 'ファイル', '備考', '状態'])
        self.setSortingEnabled(False)

        self.populateTable()
        self.sorting_order = Qt.AscendingOrder
        self.last_clicked_column = -1

    def populateTable(self):
        self.horizontalHeader().sectionClicked.connect(self.on_header_clicked)

    def on_header_clicked(self, logical_index):
        first_column = []

        for row in range(self.rowCount()):
            item = self.item(row, 0)
            if item is None:
                text = ""
            else:
                text = item.text()
            row_data = [self.item(row, col).text() if self.item(row, col) else "" for col in
                        range(1, self.columnCount())]
            first_column.append((text, row_data))

        first_column.sort(key=lambda x: x[1][logical_index - 1] if logical_index > 0 else x[0])

        for row, (text, row_data) in enumerate(first_column):
            self.setItem(row, 0, QTableWidgetItem(text))
            for col, data in enumerate(row_data):
                self.setItem(row, col + 1, QTableWidgetItem(data))

    def fixFirstColumn(self):
        self.verticalHeader().setVisible(False)

        new_vertical_header = QHeaderView(Qt.Orientation.Vertical)
        self.setVerticalHeader(new_vertical_header)

        for row in range(self.rowCount()):
            item = QTableWidgetItem(f'Row {row + 1}')
            self.setVerticalHeaderItem(row, item)

        self.resizeColumnsToContents()
        self.setColumnWidth(0, self.verticalHeader().width())

        self.verticalHeader().setSectionsClickable(False)
        self.verticalHeader().setSortIndicatorShown(False)


def generate_random_color(color):
    if color == "blue":
        blue = random.randint(150, 255)
        red = random.randint(0, blue - 50)
        green = random.randint(0, blue - 50)
    if color == "red":
        red = random.randint(200, 255)
        green = random.randint(0, 100)
        blue = random.randint(0, 100)
    if color == "yellow":
        red = random.randint(200, 255)
        green = random.randint(200, 255)
        blue = random.randint(0, 100)
    return red, green, blue


class BlinkingLabel(QLabel):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self._timer = QTimer(self)
        self._timer.timeout.connect(self.toggle_visibility)
        # self._timer.start(1000)
        self.base_text = ""
        self.dot_count = 0
        self.direction = 1

    def start_blinking(self):
        self._timer.start(1000)

    def stop_blinking(self):
        color = QColor(0, 0, 0)
        self.setStyleSheet("QLabel { color: %s }" % color.name())
        self._timer.stop()

    def toggle_visibility(self):
        self.base_text = self.text().replace(".", "")
        self.dot_count += self.direction

        if self.dot_count == 5:
            self.direction = -1
        elif self.dot_count == 0:
            self.direction = 1

        dots = '.' * self.dot_count
        self.setText(self.base_text + dots)

        # color = QColor(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
        red, green, blue = generate_random_color("blue")
        color = QColor(red, green, blue)
        self.setStyleSheet("QLabel { color: %s }" % color.name())
        # self.setVisible(not self.isVisible())


def unzip(file_path):
    file_name = os.path.basename(file_path)
    new_name = str(file_name.split('.')[0]) + '.zip'
    dir_path = os.path.dirname(os.path.abspath(file_path))
    new_path = os.path.join(dir_path, new_name)
    if os.path.exists(new_path):
        os.remove(new_path)
    shutil.copyfile(file_path, new_path)

    with zipfile.ZipFile(new_path, 'r') as file_zip:
        print(file_zip.namelist())
    file_zip = zipfile.ZipFile(new_path, 'r')
    zip_file_name = new_name.split('.')[0]
    zip_dir = os.path.join(dir_path, zip_file_name)
    for files in file_zip.namelist():
        file_zip.extract(files, zip_dir)
    file_zip.close()

    return zip_dir


def compare_pics(image1_path, image2_path, row_info, temp_dir, ketsugou_flag):
    # filesystem_encoding = sys.getfilesystemencoding()
    # image1_path = image1_path.encode(filesystem_encoding)
    # image2_path = image2_path.encode(filesystem_encoding)
    # imageA = cv2.imread(image1_path.decode(filesystem_encoding))
    # imageB = cv2.imread(image2_path.decode(filesystem_encoding))
    imageA = cv2.imread(image1_path)
    imageB = cv2.imread(image2_path)

    if imageA.shape[0] != imageB.shape[0] or imageA.shape[1] != imageB.shape[1]:
        return None

    grayA = cv2.cvtColor(imageA, cv2.COLOR_BGR2GRAY)
    grayB = cv2.cvtColor(imageB, cv2.COLOR_BGR2GRAY)

    template_paths = [f for f in os.listdir(os.path.join(temp_dir, 'exclude-pic')) if f.endswith('.png')]
    index_date1, index_port1 = None, None
    if "date1.png" in template_paths:
        index_date1 = template_paths.index("date1.png")
    if "port1.png" in template_paths:
        index_port1 = template_paths.index("port1.png")
    templates = [cv2.imread(os.path.join(temp_dir, 'exclude-pic', template_path), cv2.IMREAD_GRAYSCALE) for
                 template_path in template_paths]

    mask = np.ones(grayB.shape, dtype=np.uint8) * 255

    threshold = 0.9
    for i, template in enumerate(templates):
        template_w, template_h = template.shape[::-1]
        # DHC fengjm 2024/12/20 No.xxx UPD START
        # if i == index_port1:
        if index_port1 is not None and i == index_port1 and not ketsugou_flag:
            # DHC fengjm 2024/12/20 No.xxx UPD END
            result_a = cv2.matchTemplate(grayA, template, cv2.TM_CCOEFF_NORMED)
            min_val_a, max_val_a, min_loc_a, max_loc_a = cv2.minMaxLoc(result_a)
            result_b = cv2.matchTemplate(grayB, template, cv2.TM_CCOEFF_NORMED)
            min_val_b, max_val_b, min_loc_b, max_loc_b = cv2.minMaxLoc(result_b)
            # todo 一時削除
            # if max_val_a < max_val_b:
            #     return "reverse"
            if max_val_a > threshold and max_val_b > threshold \
                    and max_val_a == max_val_b:
                # and ((max_val_a - max_val_b) < 0.05 or (max_val_b - max_val_a) < 0.05):
                return "marusame"

        # 模板匹配
        result = cv2.matchTemplate(grayB, template, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        print("※" * 10, "image2_path:", image2_path)
        print("※" * 10, "template:", template_paths[i])
        print("※" * 10, "max_val:", max_val)
        if index_date1 is not None and i == index_date1:
            max_loc_list = list(max_loc)
            max_loc_list[0] = max_loc[0]
            max_loc_list[1] = max_loc[1] + 16

            max_loc = tuple(max_loc_list)
            print(max_loc)

        if max_val > threshold:
            # 更新掩膜
            cv2.rectangle(mask, max_loc, (max_loc[0] + template_w, max_loc[1] + template_h), (0, 0, 0), -1)

    # 应用掩膜到图片A和B
    maskedA = cv2.bitwise_and(grayA, grayA, mask=mask)
    maskedB = cv2.bitwise_and(grayB, grayB, mask=mask)
    # test_imageA = os.path.join("G:/image-compare", os.path.basename(image1_path))
    # cv2.imwrite(test_imageA, maskedA)
    # test_image = os.path.join("G:/image-compare", os.path.basename(image2_path))
    # cv2.imwrite(test_image, maskedB)

    try:
        (score, diff) = compare_ssim(maskedA, maskedB, full=True)
    except Exception as e:
        return None
    diff = (diff * 255).astype("uint8")
    print("SSIM: {}".format(score))

    offset = -15

    thresh = cv2.threshold(diff, 0, 255,
                           cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = imutils.grab_contours(cnts)

    nb_differences = 0
    for c in cnts:
        # if c[:, 0, 0].max() >= 1900 or c[0, 0, :].min() <= 40:
        #     continue
        # if c.shape[0] <= 2:
        #     continue
        txt = str(nb_differences + 1)
        (x, y, w, h) = cv2.boundingRect(c)
        cv2.rectangle(imageA, (x, y), (x + w, y + h), (0, 0, 255), 2)
        cv2.rectangle(imageB, (x, y), (x + w, y + h), (0, 0, 255), 2)
        cv2.putText(imageB, txt, (x, y + h + offset), cv2.FONT_HERSHEY_SCRIPT_SIMPLEX, 1, (0, 0, 0), 2)
        image_output = os.path.join(temp_dir, "evidence_check", os.path.basename(image2_path))
        cv2.imwrite(image_output, imageB)
        nb_differences += 1

    if nb_differences != 0:
        result_excel_ins(os.path.join(get_program_path(),
                                      get_str_before_first_dot(
                                          get_str_after_last_dot(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"], "\\"), ".")
                                      + "-比較結果.xlsx"),
                         image_output,
                         row_info, nb_differences)
    return nb_differences

    # cv2.imwrite(os.path.join(os.path.dirname(os.path.abspath(args['first'])), "out.png"), imageB)
    # cv2.imshow("Modified", imageB)

    # cv2.imshow("Diff", diff)
    # cv2.imshow("Thresh", thresh)
    # cv2.waitKey(0)


def result_excel_ins(excel_path, image_path, value, differences):
    try:
        print("write------------------", excel_path, value)
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            # sht = wb.active
        else:
            wb = Workbook()
            # sht = wb.active
        sht = wb.create_sheet(value.replace(" ", "").split(":")[1])
        img = ImageExcel(image_path)
        # cell = 'A' + str(int(ROW + 1))
        cell = "A1"
        sht[cell] = value
        cell_new = "A2"
        sht[cell_new] = "差異点：" + str(differences)
        # offset_img(img)
        offset_img_reset(img)
        sht.add_image(img)
        if 'Sheet' in wb.sheetnames:
            ws_for_remove = wb['Sheet']
            wb.remove(ws_for_remove)
        wb.save(excel_path)
        wb.close()
        # os.chmod(excel_path, 0o777)
    except Exception as e:
        print(f"exception '{e}' ...")
    finally:
        wb.close()
        del wb
        gc.collect()


def offset_img_reset(img):
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=1, colOff=0, row=1, rowOff=0)
    img.anchor = OneCellAnchor(_from=marker, ext=size)


def offset_img(img):
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    global COL, ROW
    marker = AnchorMarker(col=COL, colOff=0, row=ROW, rowOff=0)
    ROW = ROW + 63
    img.anchor = OneCellAnchor(_from=marker, ext=size)


def svn_lock_file(file_path, message):
    try:
        result = subprocess.run(['svn', 'lock', file_path, '-m', message], capture_output=True, text=True, check=True
                                , shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
        print("锁定成功：", result.stdout)
        return result.stdout
    except subprocess.CalledProcessError as e:
        print("锁定失败：", e.stderr)
        return e.stderr


def process_jacoco_html_report(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        # html_content = file.read()
        soup = BeautifulSoup(file, "html.parser")

    pre_tag = soup.find("pre", class_="source")
    if pre_tag:
        java_code = pre_tag.get_text()

    return java_code


class EventHandler:
    """EventHandler"""

    # DHC fengjm 2024/12/19 No.xxx UPD START
    # def __init__(self, parent):
    #     self.parent = parent
    def __init__(self, parent):
        self.parent = parent

    # DHC fengjm 2024/12/19 No.xxx UPD END

    def on_combobox_changed(self, index):
        if index == 0:
            self.parent.input_itest.setEnabled(True)

            self.parent.input_id.setEnabled(False)
            self.parent.input_type.setEnabled(False)
            self.parent.input_system.setEnabled(False)
            self.parent.input_object.setEnabled(False)
            self.parent.input_svn.setEnabled(False)
            self.parent.input_branch.setEnabled(False)
            self.parent.input_manual.setEnabled(False)
            self.parent.input_old.setEnabled(False)
            self.parent.input_new.setEnabled(False)
            self.parent.input_wbs.setEnabled(False)

            self.parent.checkbox_all.setEnabled(False)
            self.parent.checkbox_doc.setEnabled(False)
            self.parent.checkbox_context.setEnabled(False)
            self.parent.checkbox_source.setEnabled(False)
            self.parent.checkbox_modify.setEnabled(False)
            self.parent.checkbox_coverage.setEnabled(False)
            self.parent.checkbox_pic.setEnabled(False)
        else:
            self.parent.input_itest.setEnabled(False)

            self.parent.input_id.setEnabled(True)
            self.parent.input_type.setEnabled(True)
            self.parent.input_system.setEnabled(True)
            self.parent.input_object.setEnabled(True)
            self.parent.input_svn.setEnabled(True)
            self.parent.input_branch.setEnabled(True)
            self.parent.input_manual.setEnabled(True)
            self.parent.input_old.setEnabled(True)
            self.parent.input_new.setEnabled(True)
            self.parent.input_wbs.setEnabled(True)

            self.parent.checkbox_all.setEnabled(True)
            self.parent.checkbox_doc.setEnabled(False)
            self.parent.checkbox_context.setEnabled(True)
            self.parent.checkbox_source.setEnabled(True)
            self.parent.checkbox_modify.setEnabled(True)
            self.parent.checkbox_coverage.setEnabled(True)
            self.parent.checkbox_pic.setEnabled(True)

    def button_svn_click(self):
        """SVN開く"""
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            if self.parent.input_svn.text() is not None:
                open_path = self.parent.input_svn.text()
            folder_path = QFileDialog.getExistingDirectory(self.parent, "SVNパス選択",
                                                           directory=open_path,
                                                           options=options)
            if folder_path:
                print(folder_path)
                self.parent.input_svn.setText(folder_path)
        except Exception as e:
            self.parent.exec_button.setDisabled(False)
            print("An error occurred : ", e)
            raise

    def button_old_click(self):
        """元ソース開く"""
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            if self.parent.input_old.text() is not None:
                open_path = self.parent.input_old.text()
            folder_path = QFileDialog.getExistingDirectory(self.parent, "元ソースパス選択",
                                                           directory=open_path,
                                                           options=options)
            if folder_path:
                print(folder_path)
                self.parent.input_old.setText(folder_path)
        except Exception as e:
            self.parent.exec_button.setDisabled(False)
            print("An error occurred : ", e)
            raise

    def button_new_click(self):
        """新ソース開く"""
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            if self.parent.input_new.text() is not None:
                open_path = self.parent.input_new.text()
            folder_path = QFileDialog.getExistingDirectory(self.parent, "新ソースパス選択",
                                                           directory=open_path,
                                                           options=options)
            if folder_path:
                print(folder_path)
                self.parent.input_new.setText(folder_path)
        except Exception as e:
            self.parent.exec_button.setDisabled(False)
            print("An error occurred : ", e)
            raise

    def button_manual_click(self):
        """マニュアル開く"""
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        manual_file, _ = QFileDialog.getOpenFileName(None, "マニュアルを選択",
                                                     load_config_content("Paths").get('input_manual', ''),
                                                     "Excel Files (*.xlsx *.xls)", options=options)
        if os.path.exists(manual_file):
            self.parent.input_manual.setText(manual_file)
            # global input_browse
            # input_browse = evidence_file
            # global report_path
            # report_path = evidence_file.split('.xlsx')[0] + load_config_content("Output").get('output_result', '')

    def button_wbs_click(self):
        """外部WBS開く"""
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        wbs_file, _ = QFileDialog.getOpenFileName(None, "外部WBSを選択",
                                                  load_config_content("Paths").get('input_wbs', ''),
                                                  "Excel Files (*.xlsx *.xls)", options=options)
        if os.path.exists(wbs_file):
            self.parent.input_wbs.setText(wbs_file)

    # DHC fengjm 2024/12/20 No.itest ADD START
    def button_itest_click(self):
        """結合テストエビデンス開く"""
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        itest_file, _ = QFileDialog.getOpenFileName(None, "結合テストエビデンスを選択",
                                                    load_config_content("Paths").get('input_itest', ''),
                                                    "Excel Files (*.xlsx *.xls)", options=options)
        if os.path.exists(itest_file):
            self.parent.input_itest.setText(itest_file)

    # DHC fengjm 2024/12/20 No.itest ADD END

    def app_lock(self):
        """成果物をロック"""
        table_data = []
        table = self.parent.tabs.widget(0).layout().itemAt(0).widget()
        for row in range(table.rowCount()):
            row_data = []
            for column in range(table.columnCount()):
                item = table.item(row, column)
                row_data.append(item.text())

            table_data.append(row_data)

        result = ""
        for item in table_data:
            if item[2] is not None and item[2] != "":
                svn_result = svn_lock_file(item[2].replace("\n", ""),
                                           "「" + self.parent.input_id.text() + "」レビューのため、ロックします。")
                result = result + get_str_after_last_dot(item[2], "\\").replace("\n", "") + " : " + svn_result + "\n"
        set_message_box("INFO", "SVN", result)

    def app_execute(self):
        # # DHC fengjm 2024/12/20 No.radio ADD START
        # global RADIO_FLAG
        # # DHC fengjm 2024/12/20 No.radio ADD END
        #
        # # DHC fengjm 2024/12/20 No.itest ADD START
        # global ITEST_FLAG
        # # DHC fengjm 2024/12/20 No.itest ADD END
        """実行ボタン"""
        global EXCEL_TEST, EXCEL_EVIDENCE, EXCEL_LIST, EXCEL_COMPARE, \
            EXCEL_COVERAGE, EXCEL_REVIEW, EXCEL_CD_CHECKLIST, EXCEL_UT_CHECKLIST, EXCEL_TOTAL_MAP
        """tabsに全部のtableをクリア"""
        self.parent.exec_button.setDisabled(True)
        clearAllTables(self)
        self.parent.status_label.start_blinking()
        self.parent.progress_bar.setValue(0)
        check_flag, context = is_null_check(self)
        if check_flag is True:
            set_message_box("WARNING", "非空チェック", context[:len(context) - 1])
            return
        # excel_test_path = find_files(self.parent.input_svn.text(), self.parent.input_id.text(), EXCEL_TEST)
        # if excel_test_path is None:
        #     set_message_box("CRITICAL", "ファイル", "機能ID「" + self.parent.input_id.text()
        #                     + "」の成果物が見つかりませんので、チェックしてください。")
        #     return
        # else:
        #     EXCEL_TEST = excel_test_path
        if self.parent.input_test.currentIndex() == 1:
            if CHECKBOX_DOC:
                """タブ①、ドキュメントチェック"""
                self.parent.tabs.setCurrentIndex(0)
                try:
                    self.document_check()
                except Exception as e:
                    self.parent.exec_button.setDisabled(False)
                    print(f"{e}")
                    return
                self.parent.progress_bar.setValue(20)

            if CHECKBOX_CONTEXT:
                """タブ②、記入内容チェック"""
                self.parent.tabs.setCurrentIndex(1)
                # set_status_label(self, "記入内容チェック中")
                self.content_check()
                self.parent.progress_bar.setValue(40)

            if CHECKBOX_SOURCE:
                """タブ③、チェック"""
                self.parent.tabs.setCurrentIndex(2)
                self.source_check()
                self.parent.progress_bar.setValue(60)

            if CHECKBOX_MODIFY:
                """タブ④、チェック"""
                self.parent.tabs.setCurrentIndex(3)
                self.modify_check()
                self.parent.progress_bar.setValue(80)

            if CHECKBOX_COVERAGE:
                """タブ⑤、チェック"""
                self.parent.tabs.setCurrentIndex(4)
                self.coverage_check()

            if CHECKBOX_PIC:
                if self.parent.input_object.currentText() == "結合テスト":
                    EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"] = self.parent.input_itest.text()
                """タブ⑥、チェック"""
                self.parent.tabs.setCurrentIndex(5)
                try:
                    self.evidence_check()
                except Exception as e:
                    self.parent.exec_button.setDisabled(False)
                    set_message_box("CRITICAL", "エビデンス", f"エビデンスには{e}が発生しました。")
                    return
        else:
            EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"] = self.parent.input_itest.text()
            """タブ⑥、チェック"""
            self.parent.tabs.setCurrentIndex(5)
            try:
                self.evidence_check()
            except Exception as e:
                self.parent.exec_button.setDisabled(False)
                set_message_box("CRITICAL", "エビデンス", f"エビデンスには{e}が発生しました。")
                return

        #     if RADIO_FLAG == 0 or RADIO_FLAG == 1:
        #         """タブ②、記入内容チェック"""
        #         self.parent.tabs.setCurrentIndex(1)
        #         # set_status_label(self, "記入内容チェック中")
        #         self.content_check()
        #         self.parent.progress_bar.setValue(40)
        #
        #         """タブ③、チェック"""
        #         self.parent.tabs.setCurrentIndex(2)
        #         self.source_check()
        #         self.parent.progress_bar.setValue(60)
        #
        #         """タブ④、チェック"""
        #         self.parent.tabs.setCurrentIndex(3)
        #         self.modify_check()
        #         self.parent.progress_bar.setValue(80)
        #
        #         """タブ⑤、チェック"""
        #         self.parent.tabs.setCurrentIndex(4)
        #         self.coverage_check()
        #
        #     """タブ⑥、チェック"""
        #     if RADIO_FLAG == 0 or RADIO_FLAG == 2:
        #         self.parent.tabs.setCurrentIndex(5)
        #         try:
        #             self.evidence_check()
        #         except Exception as e:
        #             self.parent.exec_button.setDisabled(False)
        #             set_message_box("CRITICAL", "エビデンス", f"エビデンスには{e}が発生しました。")
        #             return
        # # DHC fengjm 2024/12/19 No.radio UPD END
        #
        # # DHC fengjm 2024/12/20 No.itest ADD START
        # elif ITEST_FLAG:
        #     EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"] = self.parent.input_itest.text()
        #     self.parent.tabs.setCurrentIndex(5)
        #     try:
        #         self.evidence_check()
        #     except Exception as e:
        #         self.parent.exec_button.setDisabled(False)
        #         set_message_box("CRITICAL", "エビデンス", f"エビデンスには{e}が発生しました。")
        #         return
        # # DHC fengjm 2024/12/20 No.itest ADD END

        set_status_label(self, "チェック完了しました。")
        self.parent.status_label.stop_blinking()
        self.parent.exec_button.setDisabled(False)
        self.parent.lock_button.setDisabled(False)
        # self.parent.save_button.setDisabled(False)
        # self.parent.exec_button.setDisabled(True)
        self.parent.progress_bar.setValue(100)
        self.parent.tabs.setCurrentIndex(5)
        for i in range(6):
            tab_table = self.parent.tabs.widget(i).layout().itemAt(0).widget()
            tab_table.scrollToTop()

    @log_and_call
    def document_check(self):
        """ドキュメントチェック"""
        global EXCEL_TEST, EXCEL_EVIDENCE, EXCEL_LIST, EXCEL_COMPARE, \
            EXCEL_COVERAGE, EXCEL_REVIEW, EXCEL_CD_CHECKLIST, EXCEL_UT_CHECKLIST, EXCEL_TOTAL_MAP
        context_list = ["エビデンスファイル", None, None, None]
        self.row_append(0, context_list, TIFFANY_BLUE)
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[0]:
            system = "cis"
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[1]:
            system = "order"
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[2]:
            system = "portal"
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[3]:
            system = "portal2"
        if self.parent.input_type.currentText() == "API":
            system = None
        for i, excel in enumerate(variables):
            if i == 0:
                """テスト仕様書"""
                if self.parent.input_object.currentText().find("結合") >= 0:
                    variables['EXCEL_TEST'] = "結合テスト仕様書"
                    variables['EXCEL_EVIDENCE'] = "結合テストエビデンス"
                    # variables[excel] = "結合テスト仕様書"
                else:
                    variables['EXCEL_TEST'] = "単体テスト仕様書"
                    variables['EXCEL_EVIDENCE'] = "単体テストエビデンス"
                excel_test_path = find_files_in_svn(self.parent.input_svn.text(), self.parent.input_id.text(),
                                                    variables[excel], True, system)
                if excel_test_path is not None:
                    svn_result = svn_check_file(excel_test_path)
                    context_list = [None, variables[excel], excel_test_path + "\n" + svn_result, "〇"]
                    EXCEL_TOTAL_MAP[excel] = excel_test_path
                    self.row_append(0, context_list, Qt.white)
                else:
                    raise
            else:
                """その他"""
                if variables[excel] == EXCEL_COVERAGE:
                    excel_coverage_path = find_coverage_in_folder(os.path.dirname(EXCEL_TOTAL_MAP['EXCEL_TEST']),
                                                                  system)
                    context_list = [None, variables[excel], excel_coverage_path, "ー"]
                    EXCEL_TOTAL_MAP[excel] = excel_coverage_path
                    self.row_append(0, context_list, TIFFANY_BLUE)
                    continue
                else:
                    excel_evidence_path = find_files_in_svn(self.parent.input_svn.text(),
                                                            self.parent.input_id.text(), variables[excel], False,
                                                            system)
                    if excel_evidence_path is None and variables[excel] == EXCEL_COMPARE:
                        EXCEL_COMPARE = "差分修正確認結果"
                        variables[excel] = "差分修正確認結果"
                        excel_evidence_path = find_files_in_svn(self.parent.input_svn.text(),
                                                                self.parent.input_id.text(), variables[excel], False,
                                                                system)
                if excel_evidence_path is not None:
                    svn_result = svn_check_file(excel_evidence_path)
                    context_list = [None, variables[excel], excel_evidence_path + "\n" + svn_result, "〇"]
                    EXCEL_TOTAL_MAP[excel] = excel_evidence_path
                    self.row_append(0, context_list, Qt.white)
                else:
                    context_list = [None, variables[excel], None, "✕"]
                    EXCEL_TOTAL_MAP[excel] = None
                    self.row_append(0, context_list, TIFFANY_BLUE)

    def row_append(self, index, context, color):
        tab_table = self.parent.tabs.widget(index).layout().itemAt(0).widget()
        current_row_count = tab_table.rowCount()
        tab_table.setRowCount(current_row_count + 1)

        # Fill the new row with data
        for col, line in enumerate(context):
            if line is None:
                line = ""
            item = QTableWidgetItem(f"{line}")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if color is not None:
                item.setBackground(color)
            if col == 4:
                font = item.font()
                font.setBold(True)
                item.setFont(font)
            tab_table.setItem(current_row_count, col, item)

        tab_table.resizeColumnsToContents()
        tab_table.resizeRowsToContents()
        QApplication.processEvents()
        # bottomRightItem = tab_table.item(current_row_count, 0)
        # tab_table.scrollToItem(bottomRightItem)
        tab_table.scrollToBottom()

        # TODO
        # for column in range(tab_table.columnCount()):
        #     # item = QTableWidgetItem(f"Row {current_row_count + 1}, Column {column + 1}")
        #     item = QTableWidgetItem(tab_table.item(current_row_count, col))
        #     item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        #     tab_table.setItem(current_row_count, column, item)

    @log_and_call
    def content_check(self):
        """記入内容チェック"""
        global EXCEL_TEST, EXCEL_EVIDENCE, EXCEL_LIST, EXCEL_COMPARE, \
            EXCEL_COVERAGE, EXCEL_REVIEW, EXCEL_CD_CHECKLIST, EXCEL_UT_CHECKLIST, EXCEL_TOTAL_MAP
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[0]:
            system = "EnabilityCis"
        else:
            system = self.parent.input_system.currentText()

        name_list = read_name_excel(self.parent.input_wbs.text(), "WBS")
        code_user_name = None
        code_user_date = None
        code_review_user_name = None
        code_review_user_date = None
        test_user_name = None
        test_user_date = None
        test_review_user_name = None
        test_review_user_date = None

        for name_content in name_list:
            if name_content[4] == self.parent.input_id.text() and \
                    name_content[3] == system and \
                    name_content[2] == self.parent.input_object.currentText():
                code_user_name = name_content[18]
                code_user_date = name_content[21]
                code_review_user_name = name_content[27]
                code_review_user_date = name_content[30]
                test_user_name = name_content[36]
                test_user_date = name_content[39]
                test_review_user_name = name_content[45]
                test_review_user_date = name_content[48]
                break

        print("※" * 10, code_user_name, "|", code_user_date, "|", test_user_name, "|", test_user_date,
              "|", code_review_user_name, "|", code_review_user_date, "|", test_review_user_name, "|",
              test_review_user_date)

        if self.parent.input_object.currentText().find("結合"):
            if test_user_name is None:
                self.row_append(1, ["", "",
                                    "この機能IDはWBSで見つかりませんでしたので、WBSを確認してください。この機能の一部はスキップします。",
                                    "✕"], CHINA_RED)
        else:
            if code_user_name is None:
                self.row_append(1, ["", "",
                                    "この機能IDはWBSで見つかりませんでしたので、WBSを確認してください。この機能の一部はスキップします。",
                                    "✕"], CHINA_RED)
                return
        """表紙チェック"""
        self.row_append(1, ["シート「表紙」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(variables):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)
                if EXCEL_TOTAL_MAP[excel] is not None:
                    if excel in ['EXCEL_LIST', 'EXCEL_COMPARE', 'EXCEL_CD_CHECKLIST']:
                        """表紙チェック"""
                        self.read_hyoushi_sheet(EXCEL_TOTAL_MAP[excel], "表紙", value, code_user_name, code_user_date)
                    else:
                        self.read_hyoushi_sheet(EXCEL_TOTAL_MAP[excel], "表紙", value, test_user_name, test_user_date)

        """②手修正作業結果"""
        self.row_append(1, ["シート「②手修正作業結果」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')

                if EXCEL_TOTAL_MAP[excel] is not None:
                    self.read_tesyusei_sheet(EXCEL_TOTAL_MAP[excel], "②手修正作業結果", value, test_user_name,
                                             test_user_date,
                                             test_review_user_name, test_review_user_date)

        """③カバレッジテストシナリオ"""
        self.row_append(1, ["シート「③カバレッジテストシナリオ」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')

                if EXCEL_TOTAL_MAP[excel] is not None:
                    self.read_scenario_sheet(EXCEL_TOTAL_MAP[excel], "③カバレッジテストシナリオ", value, test_user_name,
                                             test_user_date,
                                             test_review_user_name, test_review_user_date)

        """④画面確認結果"""
        self.row_append(1, ["シート「④画面確認結果」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')

                if EXCEL_TOTAL_MAP[excel] is not None:
                    self.read_gamenkekka_sheet(EXCEL_TOTAL_MAP[excel], "④画面確認結果", value, test_user_name,
                                               test_user_date,
                                               test_review_user_name, test_review_user_date)

        """⑤作業結果確認チェック"""
        self.row_append(1, ["シート「⑤作業結果確認」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)

                if EXCEL_TOTAL_MAP[excel] is not None:
                    if excel in ['EXCEL_LIST', 'EXCEL_COMPARE', 'EXCEL_CD_CHECKLIST']:
                        self.read_kekka_sheet(EXCEL_TOTAL_MAP[excel], "⑤作業結果確認", value, code_user_name,
                                              code_user_date,
                                              code_review_user_name, code_review_user_date)
                    else:
                        self.read_kekka_sheet(EXCEL_TOTAL_MAP[excel], "⑤作業結果確認", value, test_user_name,
                                              test_user_date,
                                              test_review_user_name, test_review_user_date)

        """成果物一覧"""
        self.row_append(1, ["シート「成果物一覧」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)

                if EXCEL_TOTAL_MAP[excel] is not None:
                    if excel in ['EXCEL_LIST', 'EXCEL_COMPARE', 'EXCEL_CD_CHECKLIST']:
                        self.read_ichiran_sheet(EXCEL_TOTAL_MAP[excel], "成果物一覧", value, code_user_name,
                                                code_user_date,
                                                code_review_user_name, code_review_user_date)
                    else:
                        self.read_ichiran_sheet(EXCEL_TOTAL_MAP[excel], "成果物一覧", value, test_user_name,
                                                test_user_date,
                                                test_review_user_name, test_review_user_date)

        """ソースファイル"""
        self.row_append(1, ["シート「ソースファイル」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)

                if EXCEL_TOTAL_MAP[excel] is not None:
                    if excel in ['EXCEL_LIST', 'EXCEL_COMPARE', 'EXCEL_CD_CHECKLIST']:
                        self.read_source_file_sheet(EXCEL_TOTAL_MAP[excel], "ソースファイル", value, code_user_name,
                                                    code_user_date,
                                                    code_review_user_name, code_review_user_date)
                    else:
                        self.read_source_file_sheet(EXCEL_TOTAL_MAP[excel], "ソースファイル", value, test_user_name,
                                                    test_user_date,
                                                    test_review_user_name, test_review_user_date)

        """補足説明"""
        self.row_append(1, ["シート「補足説明」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)

                if EXCEL_TOTAL_MAP[excel] is not None:
                    if excel in ['EXCEL_LIST', 'EXCEL_COMPARE', 'EXCEL_CD_CHECKLIST']:
                        self.read_hosoku_sheet(EXCEL_TOTAL_MAP[excel], "補足説明", value, code_user_name,
                                               code_user_date,
                                               code_review_user_name, code_review_user_date)
                    else:
                        self.read_hosoku_sheet(EXCEL_TOTAL_MAP[excel], "補足説明", value, test_user_name,
                                               test_user_date,
                                               test_review_user_name, test_review_user_date)

        """CDチェックリスト"""
        self.row_append(1, ["シート「CDチェックリスト」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)

                if EXCEL_TOTAL_MAP[excel] is not None:
                    self.read_cd_checklist_sheet(EXCEL_TOTAL_MAP[excel], "CDチェックリスト", value, code_user_name,
                                                 code_user_date,
                                                 code_review_user_name, code_review_user_date)

        """UTチェックリスト"""
        self.row_append(1, ["シート「UTチェックリスト」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)

                if EXCEL_TOTAL_MAP[excel] is not None:
                    self.read_ut_checklist_sheet(EXCEL_TOTAL_MAP[excel], "UTチェックリスト", value, test_user_name,
                                                 test_user_date,
                                                 test_review_user_name, test_review_user_date)

        """レニュー記録"""
        self.row_append(1, ["シート「レビュー記録」", None, None, None], TIFFANY_BLUE)
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)

                if EXCEL_TOTAL_MAP[excel] is not None:
                    self.read_review_sheet(EXCEL_TOTAL_MAP[excel], "レビュー記録", value, test_user_name,
                                           test_user_date,
                                           test_review_user_name, test_review_user_date)

    def read_review_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date, wbs_review_name, wbs_review_date):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return

        try:
            for row in ws.iter_rows(min_row=4,
                                    max_row=10,
                                    values_only=True):
                if row[1] is not None and row[1] != 0:
                    if row[3] != 1:
                        context_list = [None, value, f"{row[0]}の対応完了率は100％に達していない", "✕"]
                        self.row_append(1, context_list, CHINA_RED)
                    else:
                        if row[5] != 1:
                            context_list = [None, value, f"{row[0]}再確認完了率は100％に達していない", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
        finally:
            wb.close()
            del wb
            gc.collect()

    def read_ut_checklist_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date, wbs_review_name,
                                wbs_review_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return

        try:
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    file_flag = False
                    if cell.column_letter == "A" and isinstance(cell.value, int) is True:
                        file_flag = True
                    if file_flag is True and row[3].value is not None:
                        self.hyoushi_name_check(row[5].value, wbs_user, value,
                                                f"確認者「{row[5].row}, {row[5].column_letter}」")
                        self.hyoushi_name_check(row[5].value, wbs_user, value,
                                                f"レビュー者「{row[8].row}, {row[8].column_letter}」")
                        if row[4].value is None:
                            context_list = [None, value, f"チェック結果「{row[4].row}, "
                                                         f"{row[4].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        if row[6].value is None:
                            context_list = [None, value, f"確認日「{row[6].row}, "
                                                         f"{row[6].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[6].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                context_list = [None, value, f"確認日「{row[6].row}, "
                                                             f"{row[6].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if row[7].value is None:
                            context_list = [None, value, f"レビュー結果「{row[7].row}, "
                                                         f"{row[7].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        if row[9].value is None:
                            context_list = [None, value, f"レビュー日「{row[9].row}, "
                                                         f"{row[9].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[9].value, dt) is False:
                                # dt.strptime(row[9].value, '%Y/%m/%d')
                                self.parent.exec_button.setDisabled(False)
                                context_list = [None, value, f"レビュー日「{row[9].row}, "
                                                             f"{row[9].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if isinstance(row[6].value, dt) and isinstance(row[9].value, dt):
                            if row[6].value > row[9].value:
                                context_list = [None, value, f"レビュー日「{row[9].row}, "
                                                             f"{row[9].column_letter}」は確認日より前になってはいけません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        break
                    else:
                        break
        finally:
            wb.close()
            del wb
            gc.collect()

    def read_cd_checklist_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date, wbs_review_name,
                                wbs_review_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return

        try:
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    file_flag = False
                    if cell.column_letter == "A" and isinstance(cell.value, int) is True:
                        file_flag = True
                    if file_flag is True and row[1].value is not None:
                        self.hyoushi_name_check(row[3].value, wbs_user, value,
                                                f"確認者「{row[3].row}, {row[3].column_letter}」")
                        self.hyoushi_name_check(row[6].value, wbs_review_name, value,
                                                f"レビュー者「{row[6].row}, {row[6].column_letter}」")
                        if row[2].value is None:
                            context_list = [None, value, f"チェック結果「{row[2].row}, "
                                                         f"{row[2].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        if row[5].value is None:
                            context_list = [None, value, f"レビュー結果「{row[5].row}, "
                                                         f"{row[5].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)

                        if row[4].value is None:
                            context_list = [None, value, f"確認日「{row[4].row}, "
                                                         f"{row[4].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[4].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                context_list = [None, value, f"確認日「{row[4].row}, "
                                                             f"{row[4].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if row[5].value is None:
                            context_list = [None, value, f"レビュー結果「{row[5].row}, "
                                                         f"{row[5].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        if row[6].value is None:
                            context_list = [None, value, f"レビュー者「{row[6].row}, "
                                                         f"{row[6].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        if row[7].value is None:
                            context_list = [None, value, f"レビュー日「{row[7].row}, "
                                                         f"{row[7].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[7].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                context_list = [None, value, f"レビュー日「{row[7].row}, "
                                                             f"{row[7].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if isinstance(row[4].value, dt) and isinstance(row[7].value, dt):
                            if row[4].value > row[7].value:
                                context_list = [None, value, f"レビュー日「{row[7].row}, "
                                                             f"{row[7].column_letter}」は確認日より前になってはいけません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        break
                    else:
                        break
        finally:
            wb.close()
            del wb
            gc.collect()

    def read_hosoku_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date, wbs_review_name, wbs_review_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return
        try:
            func_sys = ws.cell(row=1, column=column_letter_to_number("C"))
            self.hyoushi_name_check(func_sys.value, self.parent.input_system.currentText(), value,
                                    f"システム「{func_sys.row}, {func_sys.column_letter}」")
            func_id = ws.cell(row=2, column=column_letter_to_number("C"))
            self.hyoushi_name_check(func_id.value, self.parent.input_id.text(), value,
                                    f"機能ID「{func_id.row}, {func_id.column_letter}」")
            # func_name = ws.cell(row=3, column=column_letter_to_number("C")).value
            # self.hyoushi_name_check(func_id.value, self.parent.input_id.text(), value,
            #                         f"機能ID「{func_id.row}, {func_id.column_letter}」")

            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    file_flag = False
                    if cell.column_letter == "A" and isinstance(cell.value, int) is True:
                        file_flag = True
                    if file_flag is True:
                        self.hyoushi_name_check(row[5].value, wbs_user, value,
                                                f"作成者「{row[5].row}, {row[5].column_letter}」")
                        self.hyoushi_name_check(row[7].value, wbs_review_name, value,
                                                f"確認者「{row[7].row}, {row[7].column_letter}」")
                        if row[6].value is None:
                            context_list = [None, value, f"作成日「{row[6].row}, "
                                                         f"{row[6].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_REDCHINA_RED)
                        else:
                            if isinstance(row[6].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                context_list = [None, value, f"作成日「{row[6].row}, "
                                                             f"{row[6].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if row[8].value is None:
                            context_list = [None, value, f"確認日「{row[8].row}, "
                                                         f"{row[8].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[8].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                context_list = [None, value, f"確認日「{row[8].row}, "
                                                             f"{row[8].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if isinstance(row[6].value, dt) and isinstance(row[8].value, dt):
                            if row[6].value > row[8].value:
                                context_list = [None, value, f"確認日「{row[8].row}, "
                                                             f"{row[8].column_letter}」は作成日より前になってはいけません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        break
                    else:
                        break
        finally:
            wb.close()
            del wb
            gc.collect()

    def read_source_file_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date, wbs_review_name,
                               wbs_review_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return
        try:
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    file_flag = False
                    if cell.column_letter == "B" and isinstance(cell.value, int) is True:
                        file_flag = True
                    if file_flag is True:
                        self.hyoushi_name_check(row[6].value, wbs_user, value,
                                                f"作成者「{row[6].row}, {row[6].column_letter}」")
                        self.hyoushi_name_check(row[8].value, wbs_review_name, value,
                                                f"確認者「{row[8].row}, {row[8].column_letter}」")
                        if row[7].value is None:
                            context_list = [None, value, f"作成日「{row[7].row}, "
                                                         f"{row[7].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[7].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                context_list = [None, value, f"作成日「{row[7].row}, "
                                                             f"{row[7].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if row[9].value is None:
                            context_list = [None, value, f"確認日「{row[9].row}, "
                                                         f"{row[9].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[9].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                context_list = [None, value, f"確認日「{row[9].row}, "
                                                             f"{row[9].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if isinstance(row[7].value, dt) and isinstance(row[9].value, dt):
                            if row[7].value > row[9].value:
                                context_list = [None, value, f"確認日「{row[9].row}, "
                                                             f"{row[9].column_letter}」は作成日より前になってはいけません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        break
                    else:
                        break
        finally:
            wb.close()
            del wb
            gc.collect()

    def is_consecutive(self, numbers):
        return all(numbers[i] + 1 == numbers[i + 1] for i in range(len(numbers) - 1))

    def read_ichiran_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date, wbs_review_name, wbs_review_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return
        try:
            column_values = []
            excel_values = []
            for row in range(6, ws.max_row + 1):
                cell_value = ws[f'A{row}'].value
                if cell_value is not None:
                    excel_values.append(ws[f'B{row}'].value)
                    try:
                        column_values.append(int(cell_value))
                    except ValueError:
                        context_list = [None, value, f"第「{row}」行目のＮＯ.が数字ではない", "✕"]
                        self.row_append(1, context_list, CHINA_RED)
                        column_values = []

            if len(column_values) > 0 and self.is_consecutive(column_values) is False:
                context_list = [None, value, f"ＮＯ.の順番は間違っています", "✕"]
                self.row_append(1, context_list, CHINA_RED)

            excel_flag = True
            excel_msg = []
            required_substrings = ['単体テスト仕様書',
                                   '単体テストエビデンス',
                                   'CDチェックリスト',
                                   'UTチェックリスト']
            for string in required_substrings:
                string_flag = False
                for excel in excel_values:
                    if excel.find(string) >= 0:
                        string_flag = True
                if not string_flag:
                    excel_flag = False
                    excel_msg.append(string)

            if not excel_flag:
                msg = "重要的なファイル"
                for context in excel_msg:
                    msg = msg + "「" + context + "」"
                msg = msg + "が成果物一覧に漏れてしまう"
                context_list = [None, value, msg, "✕"]
                self.row_append(1, context_list, CHINA_RED)

            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    file_flag = False
                    if cell.column_letter == "A" and isinstance(cell.value, int) is True:
                        file_flag = True
                    if file_flag is True:
                        self.hyoushi_name_check(row[4].value, wbs_user, value,
                                                f"作成者「{row[4].row}, {row[4].column_letter}」")
                        self.hyoushi_name_check(row[6].value, wbs_review_name, value,
                                                f"確認者「{row[6].row}, {row[6].column_letter}」")
                        if row[5].value is None:
                            context_list = [None, value, f"作成日「{row[5].row}, "
                                                         f"{row[5].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[5].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                context_list = [None, value, f"作成日「{row[5].row}, "
                                                             f"{row[5].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if row[7].value is None:
                            context_list = [None, value, f"確認日「{row[7].row}, "
                                                         f"{row[7].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[7].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                context_list = [None, value, f"確認日「{row[7].row}, "
                                                             f"{row[7].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if isinstance(row[5].value, dt) and isinstance(row[7].value, dt):
                            if row[5].value > row[7].value:
                                context_list = [None, value, f"確認日「{row[7].row}, "
                                                             f"{row[7].column_letter}」は作成日より前になってはいけません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        break
                    else:
                        break
        finally:
            wb.close()
            del wb
            gc.collect()

    def read_kekka_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date, wbs_review_name, wbs_review_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return
        try:
            for i in range(3):
                func_create_user = ws.cell(row=i + 1 + 7, column=column_letter_to_number("AX"))
                self.hyoushi_name_check(func_create_user.value, wbs_user, value,
                                        f"作成者「{func_create_user.row}, {func_create_user.column_letter}」")
                func_review_user = ws.cell(row=i + 1 + 7, column=column_letter_to_number("BB"))
                self.hyoushi_name_check(func_review_user.value, wbs_review_name, value,
                                        f"確認者「{func_review_user.row}, {func_review_user.column_letter}」")
                func_create_date = ws.cell(row=i + 1 + 7, column=column_letter_to_number("AZ"))
                if func_create_date.value is None:
                    context_list = [None, value, f"実施日「{func_create_date.row}, "
                                                 f"{func_create_date.column_letter}」はまだ記入していません", "✕"]
                    self.row_append(1, context_list, CHINA_RED)
                else:
                    if isinstance(func_create_date.value, dt) is False:
                        # self.parent.exec_button.setDisabled(False)
                        # context_list = [None, value, f"実施日「{func_create_date.row}, "
                        #                              f"{func_create_date.column_letter}」のフォーマットが正しくありません", "✕"]
                        # self.row_append(1, context_list, CHINA_RED)
                        pass
                    if isinstance(func_create_date.value, int) is False:
                        context_list = [None, value, f"実施日「{func_create_date.row}, "
                                                     f"{func_create_date.column_letter}」のフォーマットが正しくありません", "✕"]
                        self.row_append(1, context_list, CHINA_RED)
                func_review_date = ws.cell(row=i + 1 + 7, column=column_letter_to_number("BD"))
                if func_review_date.value is None:
                    context_list = [None, value, f"確認日「{func_review_date.row}, "
                                                 f"{func_review_date.column_letter}」はまだ記入していません", "✕"]
                    self.row_append(1, context_list, CHINA_RED)
                else:
                    if isinstance(func_review_date.value, dt) is False:
                        # self.parent.exec_button.setDisabled(False)
                        # context_list = [None, value, f"確認日「{func_review_date.row}, "
                        #                              f"{func_review_date.column_letter}」のフォーマットが正しくありません", "✕"]
                        # self.row_append(1, context_list, CHINA_RED)
                        pass
                    if isinstance(func_review_date.value, int) is False:
                        context_list = [None, value, f"確認日「{func_review_date.row}, "
                                                     f"{func_review_date.column_letter}」のフォーマットが正しくありません", "✕"]
                        self.row_append(1, context_list, CHINA_RED)
                if isinstance(func_create_date.value, int) and isinstance(func_review_date.value, int):
                    if func_create_date.value > func_review_date.value:
                        context_list = [None, value, f"確認日「{func_review_date.row}, "
                                                     f"{func_review_date.column_letter}」は作成日より前になってはいけません", "✕"]
                        self.row_append(1, context_list, CHINA_RED)
        finally:
            wb.close()
            del wb
            gc.collect()

    def read_gamenkekka_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date, wbs_review_name, wbs_review_date):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return
        try:
            func_id = ws.cell(row=3, column=column_letter_to_number("D"))
            self.hyoushi_name_check(func_id.value, self.parent.input_id.text(), value,
                                    f"機能ID「{func_id.row}, {func_id.column_letter}」")
            func_name = ws.cell(row=3, column=column_letter_to_number("F"))

            for row in ws.iter_rows(values_only=False):
                file_flag = False
                if isinstance(row[2].value, int) is True:
                    file_flag = True

                if file_flag is True:
                    if row[6].value == "〇":
                        continue
                    else:
                        context_list = [None, value, f"番号：「{row[2].value}」の現新画面比較列は 〇 ではありません", "✕"]
                        self.row_append(1, context_list, SCHENBRUNN_YELLOW)
                else:
                    continue

            wb.close()
        finally:
            wb.close()
            del wb
            gc.collect()

    def read_scenario_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date, wbs_review_name, wbs_review_date):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return
        try:
            func_id = ws.cell(row=8, column=column_letter_to_number("C"))
            self.hyoushi_name_check(func_id.value.split("/")[0], self.parent.input_id.text(), value,
                                    f"機能ID「{func_id.row}, {func_id.column_letter}」")
            # func_name = ws.cell(row=8, column=column_letter_to_number("C")).value.split("/")[1]
            # self.hyoushi_name_check(func_id.value.split("/")[1], self.parent.input_id.text(), value,
            #                         f"機能ID「{func_id.row}, {func_id.column_letter}」")
            # func_sys = ws.cell(row=2, column=column_letter_to_number("F"))
            # self.hyoushi_name_check(func_sys.value, self.parent.input_system.currentText(), value,
            #                         f"システム「{func_sys.row}, {func_sys.column_letter}」")
            # func_type = ws.cell(row=2, column=column_letter_to_number("P"))
            # self.hyoushi_name_check(func_type.value, self.parent.input_type.currentText(), value,
            #                         f"区分「{func_type.row}, {func_type.column_letter}」")
            for row in ws.iter_rows(values_only=False):
                file_flag = False
                if isinstance(row[0].value, int) is True:
                    file_flag = True

                if file_flag is True:
                    if row[43].value == "〇":
                        self.hyoushi_name_check(row[45].value, wbs_user, value,
                                                f"実施者「{row[45].row}, {row[45].column_letter}」")
                        self.hyoushi_name_check(row[49].value, wbs_review_name, value,
                                                f"確認者「{row[49].row}, {row[49].column_letter}」")
                        if row[47].value is None:
                            context_list = [None, value, f"実施日「{row[47].row}, "
                                                         f"{row[47].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[47].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                # context_list = [None, value, f"実施日「{row[47].row}, "
                                #                              f"{row[47].column_letter}」のフォーマットが正しくありません", "✕"]
                                # self.row_append(1, context_list, CHINA_RED)
                                pass
                            if isinstance(row[47].value, int) is False:
                                context_list = [None, value, f"実施日「{row[47].row}, "
                                                             f"{row[47].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if row[51].value is None:
                            context_list = [None, value, f"確認日「{row[51].row}, "
                                                         f"{row[51].column_letter}」はまだ記入していません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
                        else:
                            if isinstance(row[51].value, dt) is False:
                                # self.parent.exec_button.setDisabled(False)
                                # context_list = [None, value, f"確認日「{row[51].row}, "
                                #                              f"{row[51].column_letter}」のフォーマットが正しくありません", "✕"]
                                # self.row_append(1, context_list, CHINA_RED)
                                pass
                            if isinstance(row[51].value, int) is False:
                                context_list = [None, value, f"確認日「{row[51].row}, "
                                                             f"{row[51].column_letter}」のフォーマットが正しくありません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        if isinstance(row[47].value, int) and isinstance(row[51].value, int):
                            if row[47].value > row[51].value:
                                context_list = [None, value, f"確認日「{row[51].row}, "
                                                             f"{row[51].column_letter}」は実施日より前になってはいけません", "✕"]
                                self.row_append(1, context_list, CHINA_RED)
                        continue
                    else:
                        context_list = [None, value, f"ケースＮＯ：「{row[0].value}」の結果は 〇 ではありません", "✕"]
                        self.row_append(1, context_list, SCHENBRUNN_YELLOW)
                else:
                    continue

            wb.close()
        finally:
            wb.close()
            del wb
            gc.collect()

    def read_tesyusei_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date, wbs_review_name, wbs_review_date):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return
        try:
            # table_data = []
            # list_path = ""
            # table = self.parent.tabs.widget(0).layout().itemAt(0).widget()
            # for row in range(table.rowCount()):
            #     row_data = []
            #     for column in range(table.columnCount()):
            #         item = table.item(row, column)
            #         if item.text() == '成果物一覧':
            #             list_path = table.item(row, column + 1).text().replace("\n", "")
            #         row_data.append(item.text())
            #     table_data.append(row_data)
            wb_list = openpyxl.load_workbook(EXCEL_TOTAL_MAP["EXCEL_LIST"], data_only=True)
            try:
                ws_list = wb_list["ソースファイル"]
            except KeyError:
                self.parent.exec_button.setDisabled(False)
                return
            source_files = []
            for row in ws_list.iter_rows():
                if row[1].value is not None and isinstance(row[1].value, int):
                    source_files.append(row[2].value.split("\\")[len(row[2].value.split("\\")) - 1])

            tesyusei_list = []
            for row in ws.iter_rows(min_row=4,
                                    max_row=4,
                                    values_only=True):
                for index, cell in enumerate(row):
                    if index < 10:
                        continue
                    if cell is not None and cell != "なし":
                        tesyusei_list.append(cell)
                        if cell not in source_files:
                            context_list = [None, value, f"{cell}は成果物一覧に存在しません", "✕"]
                            self.row_append(1, context_list, CHINA_RED)
            for context in source_files:
                if context not in tesyusei_list:
                    context_list = [None, value, f"成果物一覧の{context}は「②手修正作業結果」に存在しません", "✕"]
                    self.row_append(1, context_list, CHINA_RED)

            wb.close()
        finally:
            wb.close()
            del wb
            gc.collect()

    def read_hyoushi_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return
        try:
            func_id = ws.cell(row=2, column=column_letter_to_number("V"))
            self.hyoushi_name_check(func_id.value, self.parent.input_id.text(), value,
                                    f"機能ID「{func_id.row}, {func_id.column_letter}」")
            func_name = ws.cell(row=3, column=column_letter_to_number("V"))
            func_sys = ws.cell(row=2, column=column_letter_to_number("F"))
            self.hyoushi_name_check(func_sys.value, self.parent.input_system.currentText(), value,
                                    f"システム「{func_sys.row}, {func_sys.column_letter}」")
            func_type = ws.cell(row=2, column=column_letter_to_number("P"))
            self.hyoushi_name_check(func_type.value, self.parent.input_type.currentText(), value,
                                    f"区分「{func_type.row}, {func_type.column_letter}」")
            func_create_user = ws.cell(row=2, column=column_letter_to_number("AI"))
            self.hyoushi_name_check(func_create_user.value, "DHC" + wbs_user, value,
                                    f"作成者「{func_create_user.row}, {func_create_user.column_letter}」")
            func_update_user = ws.cell(row=3, column=column_letter_to_number("AI"))
            self.hyoushi_name_check(func_update_user.value, "DHC" + wbs_user, value,
                                    f"更新者「{func_update_user.row}, {func_update_user.column_letter}」")
            func_create_date = ws.cell(row=2, column=column_letter_to_number("AP"))
            if func_create_date.value is None:
                context_list = [None, value, f"作成日「{func_create_date.row}, "
                                             f"{func_create_date.column_letter}」はまだ記入していません", "✕"]
                self.row_append(1, context_list, CHINA_RED)
            else:
                try:
                    if isinstance(func_create_date.value, str) is True:
                        dt.strptime(func_create_date.value, '%Y/%m/%d')
                except ValueError:
                    self.parent.exec_button.setDisabled(False)
                    context_list = [None, value, f"作成日「{func_create_date.row}, "
                                                 f"{func_create_date.column_letter}」のフォーマットが正しくありません", "✕"]
                    self.row_append(1, context_list, CHINA_RED)
            func_update_date = ws.cell(row=2, column=column_letter_to_number("AP"))
            if func_update_date.value is None:
                context_list = [None, value, f"更新日「{func_update_date.row}, "
                                             f"{func_update_date.column_letter}」はまだ記入していません", "✕"]
                self.row_append(1, context_list, CHINA_RED)
            else:
                try:
                    if isinstance(func_update_date.value, str) is True:
                        dt.strptime(func_update_date.value, '%Y/%m/%d')
                except ValueError:
                    self.parent.exec_button.setDisabled(False)
                    context_list = [None, value, f"更新日「{func_update_date.row}, "
                                                 f"{func_update_date.column_letter}」のフォーマットが正しくありません", "✕"]
                    self.row_append(1, context_list, CHINA_RED)
            wb.close()
        finally:
            wb.close()
            del wb
            gc.collect()

    def hyoushi_name_check(self, func_cell, cell, value, context):
        if cell is not None and func_cell != cell:
            context_list = [None, value, context, "✕"]
            self.row_append(1, context_list, CHINA_RED)

    @log_and_call
    def source_check(self):
        """ソースチェック"""
        global EXCEL_TOTAL_MAP
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[0]:
            system = "cis"
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[1]:
            system = "order"
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[2]:
            system = "portal"
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[3]:
            system = "portal2"
        moto_source_files = []
        new_source_files = []
        tab_table = self.parent.tabs.widget(2).layout().itemAt(0).widget()
        tab_table.setColumnCount(5)
        tab_table.setHorizontalHeaderLabels(['区分', 'ファイル', '備考', '状態', '手修正'])
        if EXCEL_TOTAL_MAP["EXCEL_LIST"] is not None:
            self.row_append(2, ["GIT", "GITにソース有無チェック", None, None, None], TIFFANY_BLUE)
            source_list = read_excel_list(EXCEL_TOTAL_MAP["EXCEL_LIST"], "ソースファイル")
            if source_list is not None:
                for source_file in source_list:
                    os.makedirs(os.path.join(get_program_path(), "source_check"), exist_ok=True)
                    # dir_path = os.path.join(get_program_path(), "source_check")
                    # command = 'icacls "' + dir_path + '" /grant Users:(OI)(CI)F /T'
                    # process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    # stdout, stderr = process.communicate()
                    # if process.returncode != 0:
                    #     print(f"Error modifying file: {stderr.decode('utf-8')}")
                    #     return False
                    old_file_path = find_source_files(os.path.join(self.parent.input_old.text(), system),
                                                      get_str_after_last_dot(source_file[0], "\\"))
                    if old_file_path is None:
                        item_color = QColor(255, 255, 178)
                        self.row_append(2, ["", source_file[0], "元ソースに存在しません。", "✕", source_file[1]], item_color)
                        continue
                    if download_source_from_git(old_file_path.replace(self.parent.input_old.text() + "\\", ""),
                                                os.path.join(get_program_path(), "source_check",
                                                             get_str_after_last_dot(old_file_path, "\\"))) is False:
                        self.row_append(2, ["", get_str_after_last_dot(source_file[0], "\\"), "GITにソースは存在しません。", "✕",
                                            source_file[1]],
                                        CHINA_RED)
                    else:
                        self.row_append(2, ["", get_str_after_last_dot(source_file[0], "\\"), source_file[0], "〇",
                                            source_file[1]], Qt.white)

        # if EXCEL_TOTAL_MAP["EXCEL_COMPARE"] is not None:
        #     self.row_append(2, [f"{EXCEL_COMPARE}", None, None, None], TIFFANY_BLUE)
        #     java_lists = read_sheet_names(EXCEL_TOTAL_MAP["EXCEL_COMPARE"])
        #     for java_file in java_lists:
        #         old_file_path = find_source_files(os.path.join(self.parent.input_old.text(), system),
        #                                           get_str_after_last_dot(java_file, "\\"))
        #         if old_file_path is not None:
        #             old_file_lines = count_lines_of_code(old_file_path)
        #             new_file_lines = count_max_line_from_excel(EXCEL_TOTAL_MAP["EXCEL_COMPARE"], java_file, "B")
        #             if old_file_lines != new_file_lines:
        #                 self.row_append(2, ["", java_file, str(old_file_lines) + "|" + str(new_file_lines), "✕"],
        #                                 CHINA_RED)
        #             else:
        #                 self.row_append(2, ["", java_file, "", "〇"], Qt.white)
        #         else:
        #             self.row_append(2, ["", java_file, "ソースは基準ソースに存在しません。", "✕"], CHINA_RED)

        del_folder(os.path.join(get_program_path(), "source_check"))

    def remove_leading_symbols(self, s):
        # 使用正则表达式去掉开头的所有符号和空白字符
        return re.sub(r'^[\s\u3000]+', '', s)

    @log_and_call
    def modify_check(self):
        """手修正のチェック"""
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[0]:
            system = "cis"
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[1]:
            system = "order"
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[2]:
            system = "portal"
        if self.parent.input_system.currentText() == ENABILITY_SYSTEM[3]:
            system = "portal2"
        if EXCEL_TOTAL_MAP["EXCEL_COMPARE"] is None:
            self.row_append(3, [None, None, "手修正がないため、この機能の一部はスキップします。", None], MARS_GREEN)
            return
        if EXCEL_TOTAL_MAP["EXCEL_COMPARE"] is not None:
            self.row_append(3, [f"{EXCEL_COMPARE}", "ファイル一致性", None, None], TIFFANY_BLUE)
            source_table = self.parent.tabs.widget(2).layout().itemAt(0).widget()
            source_table_data = []
            for row in range(source_table.rowCount()):
                row_data = []
                for column in range(source_table.columnCount()):
                    item = source_table.item(row, column)
                    row_data.append(item.text())
                if row_data[4] == '○':
                    source_table_data.append(row_data)

            java_lists = read_sheet_names(EXCEL_TOTAL_MAP["EXCEL_COMPARE"])
            for row in source_table_data:
                is_flag = False
                for key, value in java_lists.items():
                    if row[1].find(key) >= 0:
                        is_flag = True
                if is_flag is False:
                    self.row_append(3, ["", row[1], "手修正確認結果にはこのDIFFシートが漏れる。", "✕"], CHINA_RED)

            os.makedirs(os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_COMPARE"]), "compare"), exist_ok=True)
            source_list = self.read_java_from_excel_list(EXCEL_TOTAL_MAP["EXCEL_LIST"])
            different_lines = []
            for java_file, sheet_name in java_lists.items():
                print("java_file, sheet_name : ", java_file, sheet_name)
                different_flag = False
                full_java_path = ""
                for row_content in source_list:
                    if row_content[2].find(java_file) > 0:
                        full_java_path = row_content[2]
                if full_java_path == "":
                    self.row_append(3, ["", java_file, "成果物一覧に存在しません。", "✕"], CHINA_RED)
                    continue
                else:
                    print("full_java_path : ", full_java_path)
                    if INPUT_OBJECT == '差分取込':
                        git_file_path = full_java_path
                    else:
                        git_file_path = os.path.join(system, full_java_path)
                    download_source_from_git(git_file_path,
                                             os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]),
                                                          "compare", java_file))
                    git_file_lines = count_lines_of_code(
                        os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]),
                                     "compare", java_file))
                    new_file_lines = count_max_line_from_excel(EXCEL_TOTAL_MAP["EXCEL_COMPARE"], sheet_name, "E")
                    if git_file_lines != new_file_lines:
                        different_flag = True
                        self.row_append(3, ["", java_file,
                                            "GIT: " + str(git_file_lines) + " | " + "EXCEL: " + str(new_file_lines),
                                            "✕"],
                                        CHINA_RED)
                    else:
                        different_lines = []
                        with open(
                                os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]), "compare", java_file),
                                'r', encoding='utf-8', errors='ignore') as file:
                            git_lines = file.readlines()
                        new_lines = return_new_lines_from_excel(EXCEL_TOTAL_MAP["EXCEL_COMPARE"], sheet_name, "F")
                        for index, line in enumerate(git_lines):
                            if new_lines[index] is None:
                                new_lines[index] = ""
                            # if line[:-1] != new_lines[index].replace(" ", " "):
                            # if line[:-1].replace("\t", "        ") != new_lines[index].replace(" ", " "):
                            if self.remove_leading_symbols(line.replace("\n", "")).replace("\t", "").replace(" ", "") \
                                    != self.remove_leading_symbols(new_lines[index].replace("\n", "")).replace(" ", ""):
                                print(f"不一致行目{index + 1}", "GIT：[",
                                      self.remove_leading_symbols(line[:-1]).replace("\t", ""),
                                      "]\nEXCEL：[",
                                      self.remove_leading_symbols(new_lines[index]).replace(" ", ""),
                                      "]")
                                different_lines.append(index + 1)

                if different_flag is False:
                    if len(different_lines) > 0:
                        self.row_append(3, ["", java_file, different_lines, "✕"], CHINA_RED)
                    else:
                        self.row_append(3, ["", java_file, "", "〇"], Qt.white)

        del_folder(os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_COMPARE"]), "compare"))

        self.row_append(3, [None, "JAVAファイル以外の記入内容", None, None], TIFFANY_BLUE)
        for java_file, sheet_name in java_lists.items():
            if sheet_name.endswith(".java") is False:
                check_result = []
                wb = load_workbook(EXCEL_TOTAL_MAP["EXCEL_COMPARE"])
                try:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows():
                        # print(f"row {row[6].row}:", row[6].value)
                        if len(row) > 6 and row[6].value is not None and row[6].value.find("No.") >= 0:
                            if (len(row) >= 8 and (row[7].value is None or row[7].value == "")) \
                                    or len(row) < 8:
                                check_result.append(str(row[6].row))
                finally:
                    wb.close()
                    del wb
                    gc.collect()
                if len(check_result) > 0:
                    result_str = "記入漏れ : "
                    for check_content in check_result:
                        result_str = result_str + " " + str(check_content) + " , "
                    result_str = result_str[:-2]
                    self.row_append(3, ["", java_file, result_str, "✕"], CHINA_RED)

        self.row_append(3, [None, "手修正点数", None, None], TIFFANY_BLUE)
        for java_file, sheet_name in java_lists.items():
            pass

    @log_and_call
    def coverage_check(self):
        """カバーのチェック"""
        if os.path.exists(os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]), "カバレッジ")) is False:
            self.row_append(4, [None, None, "カバーがないのため、この機能の一部はスキップします。", None], MARS_GREEN)
        else:
            os.makedirs(os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]),
                                     "カバレッジ", "coverage"), exist_ok=True)
            extract_zip_files(os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]), "カバレッジ"),
                              os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]), "カバレッジ", "coverage"))

            if self.parent.input_system.currentText() != "API":
                file_names = fuzzy_search(os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]), "カバレッジ",
                                                       "coverage"), "*java*")
            print("※" * 10, file_names)

            source_table = self.parent.tabs.widget(2).layout().itemAt(0).widget()
            source_table_data = []
            for row in range(source_table.rowCount()):
                row_data = []
                for column in range(source_table.columnCount()):
                    item = source_table.item(row, column)
                    row_data.append(item.text())
                if row_data[4] == '○':
                    source_table_data.append(row_data)

            system = ""
            if self.parent.input_system.currentText() == ENABILITY_SYSTEM[0]:
                system = "cis"
            if self.parent.input_system.currentText() == ENABILITY_SYSTEM[1]:
                system = "order"
            if self.parent.input_system.currentText() == ENABILITY_SYSTEM[2]:
                system = "portal"
            if self.parent.input_system.currentText() == ENABILITY_SYSTEM[3]:
                system = "portal2"

            for item in file_names:
                self.find_source(source_table_data, item, system)

            self.row_append(4, [None, "成果物一覧からソース一致性", "", None], TIFFANY_BLUE)
            source_list = self.read_java_from_excel_list(EXCEL_TOTAL_MAP["EXCEL_LIST"])
            if self.parent.input_type.currentText() == "API":
                file_names = api_coverage_search(
                    os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]), "カバレッジ", "coverage"),
                    source_list)

                for java in file_names:
                    self.check_source_is_same_from_git_api(system, java.split("|=")[0],
                                                           source_list, java.split("|=")[1])

                self.row_append(4, [None, "カバー状況", None, None], TIFFANY_BLUE)
                for java in file_names:
                    print("COVERAGE TEST : ", java.split("|=")[1])
                    result_map = read_coverage_html_from_api(java.split("|=")[0])
                    print("COVERAGE TEST : ", result_map)
                    if len(result_map) > 0:
                        for key, value in result_map.items():
                            self.row_append(4, [None, java.split("|=")[1],
                                                value + "|行目(約)：" + str(key), "✕"], CHINA_RED)
                    else:
                        self.row_append(4, [None, java.split("|=")[1], "問題なし", "〇"], Qt.white)

            else:
                coverage_java = {}
                for source in source_list:
                    for file in file_names:
                        if file.endswith(source[2].split("\\")[len(source[2].split("\\")) - 1] + ".html") \
                                and ((file.find("Custom_") >= 0) == (source[2].find("Custom_") >= 0)
                                     and (file.find("_SACM") >= 0) == (source[2].find("_SACM") >= 0)
                                     and (file.find("_SUCM") >= 0) == (source[2].find("_SUCM") >= 0)):
                            coverage_java[file] = source

                for html_file, source in coverage_java.items():
                    self.check_source_is_same_from_git(system, html_file, source)

                # self.row_append(4, [None, "カバー状況", None, None], TIFFANY_BLUE)
                # for java in file_names:
                #     print("COVERAGE TEST : ", java)
                #     result_map = read_coverage_html_from_api(java)
                #     print("COVERAGE TEST : ", result_map)
                #     if len(result_map) > 0:
                #         for key, value in result_map.items():
                #             self.row_append(4, [None, java,
                #                                 value + "|行目(約)：" + str(key), "✕"], CHINA_RED)
                #     else:
                #         self.row_append(4, [None, java, "問題なし", "〇"], Qt.white)

            del_folder(os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]), "カバレッジ", "coverage"))

    def check_source_is_same_from_git(self, system, html_file, java_file):
        if INPUT_OBJECT == '差分取込':
            git_file_path = java_file[2]
        else:
            git_file_path = os.path.join(system, java_file[2])
        local_java_code_from_git = os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]), "カバレッジ",
                                                "coverage", java_file[2].split("\\")[len(java_file[2].split("\\")) - 1])

        download_source_from_git(git_file_path, local_java_code_from_git)
        if count_lines_of_code(html_file) - 1 != count_lines_of_code(local_java_code_from_git):
            self.row_append(4, [None,
                                java_file[2], "カバーレポートとGITからのソースが不一致です。", "✕"],
                            CHINA_RED)
        else:
            java_code_from_coverage = process_jacoco_html_report(html_file)
            java_code_html_for_compare = java_code_from_coverage.split("\n")
            with open(local_java_code_from_git, 'r', encoding='utf-8') as file:
                java_code_git_for_compare = file.readlines()
            # print(java_code_from_coverage)
            compare_result = []
            for line1, line2 in zip(java_code_html_for_compare, java_code_git_for_compare):
                if line1 != line2.replace("\n", ""):
                    compare_result.append(line1 + "|" + line2)

            if len(compare_result) == 0:
                self.row_append(4, [None,
                                    java_file[2], "カバーレポートとGITからのソースが一致です。", "〇"],
                                Qt.white)
            else:
                self.row_append(4, [None,
                                    java_file[2], "カバーレポートとGITからのソースが不一致です。", "✕"],
                                CHINA_RED)
                for content in compare_result:
                    self.row_append(4, [None, None, content, None], CHINA_RED)

    def check_source_is_same_from_git_api(self, system, java, source_list, java_name):
        for java_file in source_list:
            if java_file[2].find(java_name + ".java") > 0:
                if INPUT_OBJECT == '差分取込':
                    git_file_path = java_file[2]
                else:
                    git_file_path = os.path.join(system, java_file[2])
                download_source_from_git(git_file_path,
                                         os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]), "カバレッジ",
                                                      "coverage", os.path.basename(java_file[2]).replace(".html", "")))
                if self.count_lines_of_code_from_code_tag(java) != count_lines_of_code(
                        os.path.join(os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]), "カバレッジ", "coverage",
                                     os.path.basename(java_file[2]).replace(".html", ""))):
                    self.row_append(4, [None,
                                        java_file[2], "カバーレポートとGITからのソースが不一致です。", "✕"],
                                    CHINA_RED)
                else:
                    self.row_append(4, [None,
                                        java_file[2], "カバーレポートとGITからのソースが一致です。", "〇"],
                                    Qt.white)

    def count_lines_of_code_from_code_tag(self, source_html):
        with open(source_html, 'r', encoding='shift_jis', errors='ignore') as file:
            lines_content = file.readlines()
        html_content = ''.join(lines_content)
        soup = BeautifulSoup(html_content, 'html.parser')
        code_tags = soup.find_all('code')

        for i, code_tag in enumerate(code_tags):
            code_text = code_tag.get_text()  # 使用get_text()获取内容
            if code_text:  # 检查获取的文本是否非空
                lines = code_text.splitlines()

        return len(lines)

    def read_java_from_excel_list(self, excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        try:
            ws = wb["ソースファイル"]
        except KeyError:
            self.parent.exec_button.setDisabled(False)
            return

        source_list = []
        for row in ws.iter_rows():
            source_list_list = []
            if row[1].value == "NO" or row[1].value is None:
                continue
            for cell in row:
                source_list_list.append(cell.value)
            source_list.append(source_list_list)
        return source_list

    def find_source(self, source_table_data, item, system):
        file_name = os.path.basename(item)
        for row in source_table_data:
            if row[1] == file_name.replace(".html", ""):
                if INPUT_OBJECT == '差分取込':
                    git_file_path = row[2]
                else:
                    git_file_path = os.path.join(system, row[2])
                download_source_from_git(git_file_path,
                                         os.path.join(
                                             os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]),
                                             "カバレッジ", "coverage", file_name.replace(".html", "")))
                self.row_append(4, [None, file_name.replace(".html", ""), "ソース一致性", None], TIFFANY_BLUE)
                if count_lines_of_code(item) - 1 != count_lines_of_code(os.path.join(
                        os.path.dirname(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"]),
                        "カバレッジ", "coverage", file_name.replace(".html", ""))):
                    self.row_append(4, [None,
                                        file_name.replace(".html", ""), "カバーレポートとGITからのソースが不一致です。", "✕"],
                                    CHINA_RED)
                else:
                    self.row_append(4, [None,
                                        file_name.replace(".html", ""), "カバーレポートとGITからのソースが一致です。", "〇"],
                                    Qt.white)
                self.row_append(4, [None,
                                    file_name.replace(".html", ""), "カバー状況", None],
                                TIFFANY_BLUE)
                print("COVERAGE TEST : ", item)
                result_map = read_coverage_html(item)
                print("COVERAGE TEST : ", result_map)
                if len(result_map) > 0:
                    for key, value in result_map.items():
                        self.row_append(4, [None, file_name.replace(".html", ""),
                                            value + "|行目：" + str(key), "✕"], CHINA_RED)
                else:
                    self.row_append(4, [None, file_name.replace(".html", ""), "問題なし", "〇"], Qt.white)
                break

    @log_and_call
    def evidence_check(self):
        """エビデンスチェック"""
        global COL, ROW
        COL, ROW = 1, 1
        if self.parent.input_type.currentText() != ENABILITY_TYPE[0]:
            self.row_append(5, [None, None, "画面のテストだけでキャプチャーを比較するため、この機能の一部はスキップします。", None], MARS_GREEN)
            return
        tab_table = self.parent.tabs.widget(5).layout().itemAt(0).widget()
        tab_table.setColumnCount(5)
        tab_table.setHorizontalHeaderLabels(['区分', 'ファイル', '備考', '状態', '違う点数'])
        self.row_append(5, ["エビデンス", "エビデンスにキャプチャーのピクセル一致性チェック", None, None, None], TIFFANY_BLUE)
        if os.path.exists(os.path.join(get_program_path(),
                                       get_str_before_first_dot(
                                           get_str_after_last_dot(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"], "\\"), ".")
                                       + "-比較結果.xlsx")):
            os.remove(os.path.join(get_program_path(),
                                   get_str_before_first_dot(
                                       get_str_after_last_dot(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"], "\\"), ".")
                                   + "-比較結果.xlsx"))
        # temp_dir = tempfile.gettempdir()
        # print("臨時フォルダー：", temp_dir)
        temp_dir = os.path.splitdrive(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"])[0] + os.path.sep
        print("エビデンスの根フォルダー：", temp_dir)
        os.makedirs(os.path.join(temp_dir, "evidence_check"), exist_ok=True)
        tmp_file = os.path.join(temp_dir, "evidence_check",
                                str(datetime.datetime.now()).
                                replace('-', '').replace('.', '').
                                replace(' ', '').replace('-', '').
                                replace(':', '') + ".xlsx")
        shutil.copyfile(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"], tmp_file)
        zip_dir = unzip(tmp_file)
        print("!" * 10, os.path.join(zip_dir, 'xl', 'drawings', 'drawing1.xml'))
        if os.path.exists(os.path.join(zip_dir, 'xl', 'drawings', 'drawing1.xml')) is False:
            self.row_append(5, [None, None, "エビデンスにはキャプチャーがありません。", None, None], CHINA_RED)
            raise
        # shutil.copytree(zip_dir, os.path.join(temp_dir, "evidence_check"), dirs_exist_ok=True)
        copy_folder(os.path.join(get_program_path(), "exclude-pic"), os.path.join(temp_dir, "exclude-pic"))
        self.pic_no_cell = []

        workbook_path = os.path.join(zip_dir, 'xl', 'workbook.xml')
        # 解析 workbook.xml
        tree = ET.parse(workbook_path)
        root = tree.getroot()
        # 命名空间处理
        ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        sheet_names = []
        for index, sheet in enumerate(root.findall('x:sheets/x:sheet', ns)):
            sheet_names.append(sheet.get('name'))
        print("sheet_names : ", sheet_names)

        sheet_to_xml = {}
        for index, sheet_name in enumerate(sheet_names, start=1):
            sheet_xml_path = os.path.join(zip_dir, f'xl/worksheets/sheet{index}.xml')
            sheet_to_xml[sheet_name] = sheet_xml_path
        print("sheet_to_xml : ", sheet_to_xml)

        drawing_to_sheet = {}
        for sheet_name, sheet_xml_path in sheet_to_xml.items():
            rels_file_path = sheet_xml_path.replace("/worksheets/", "/worksheets/_rels/").replace(".xml", ".xml.rels")
            if os.path.exists(rels_file_path):
                tree = ET.parse(rels_file_path)
                root = tree.getroot()
                namespace = {'ns': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                target_value = root.find('.//ns:Relationship', namespace).get('Target')
                if target_value.find("drawing") >= 0:
                    drawing_to_sheet[target_value.split("drawings/")[1]] = sheet_name
                print(target_value)
        print("drawing_to_sheet : ", drawing_to_sheet)

        combined_mapping = {}
        for sheet, xml in sheet_to_xml.items():
            # 查找对应的 drawing
            drawings = [drawing for drawing, sheet_name in drawing_to_sheet.items() if sheet_name == sheet]
            combined_mapping[sheet] = {
                "xml": xml,
                "drawings": drawings
            }
        # 输出结果
        print("Combined Mapping:")
        for sheet, info in combined_mapping.items():
            print(f"{sheet}: XML -> {info['xml']}, Drawings -> {info['drawings']}")
            if sheet == "データ比較" or sheet == "ファイル比較":
                continue
            if len(info['drawings']) == 1:
                xml_name = info['drawings'][0]
                if os.path.exists(os.path.join(zip_dir, 'xl', 'drawings', xml_name)) is True:
                    self.pic_no_info_create(zip_dir, xml_name, sheet)

        sorted_list = sorted(
            self.pic_no_cell,
            key=lambda x: (str(x['xml']), int(x['row']), int(x['column']))
        )
        print("※ sorted_list : ", sorted_list)

        sorted_list_new = []
        for index, sort_item in enumerate(sorted_list):
            if sort_item['column'] == '1':
                if sort_item['anchor'] == 'one':
                    sort_item['row'] = str(int(sort_item['row']) + 1)
                sorted_list_new.append(sort_item)
                continue
            if sort_item['column'] == '15' and sorted_list[index - 1]['column'] == '1':
                if sort_item['anchor'] == 'one':
                    if sorted_list[index - 1]['anchor'] == 'one':
                        sort_item['row'] = str(int(sort_item['row']) + 1)
                        sorted_list_new.append(sort_item)
                        continue
                    if sorted_list[index - 1]['anchor'] == 'two':
                        sort_item['row'] = str(int(sort_item['row']) + 1)
                        sorted_list_new.append(sort_item)
                        continue
                if sort_item['anchor'] == 'two':
                    sorted_list_new.append(sort_item)
                    continue
                    # if sorted_list[index - 1]['anchor'] == 'one':
                    #     sort_item['row'] = str(int(sort_item['row']) - 1)
                    #     sorted_list_new.append(sort_item)
                    #     continue
                    # if sorted_list[index - 1]['anchor'] == 'two':
                    #     sorted_list_new.append(sort_item)
                    #     continue
        print("※ sorted_list_new : ", sorted_list_new)

        # 假设 sorted_list_new 已经生成
        grouped_data = {}

        for item in sorted_list_new:
            row_key = item['row']
            xml_key = item['xml']

            # 使用元组 (row_key, xml_key) 作为组合键
            combined_key = (row_key, xml_key)

            # 初始化字典
            if combined_key not in grouped_data:
                grouped_data[combined_key] = []  # 创建一个列表以存储条目

            grouped_data[combined_key].append(item)  # 将项目添加到对应的组合键下

        # 打印结果
        print("※ grouped_data : ", grouped_data)

        for key, group in grouped_data.items():
            if len(group) >= 2:
                pic1 = group[0]['pic']
                pic2 = group[1]['pic']
                col1 = group[0]['column']
                col2 = group[1]['column']
                if col1 != "1" or col2 != "15":
                    self.row_append(5, [None, "sheet「" + group[0]['xml'] + "」" + "row「" + str(
                        int(group[0]['row'])) + "」",
                                        "キャプチャーの位置が間違っています。", "✕", "ー"], QColor(139, 0, 0))
                print(f"row {key}: pic1={pic1}, pic2={pic2}, col1={col1}, col2={col2}")
                image1_path = os.path.join(temp_dir,
                                           "evidence_check",
                                           os.path.basename(tmp_file).split(".")[0],
                                           "xl",
                                           "media",
                                           pic1)
                image2_path = os.path.join(temp_dir,
                                           "evidence_check",
                                           os.path.basename(tmp_file).split(".")[0],
                                           "xl",
                                           "media",
                                           pic2)

                ketsugou_flag = False
                if self.parent.input_object.currentText() == "結合テスト":
                    ketsugou_flag = True
                difference_count = compare_pics(image1_path, image2_path,
                                                "ROW_INFO : " + group[0]['xml'] + "|" + str(int(group[0]['row'])),
                                                temp_dir, ketsugou_flag)
                print("※" * 10, "difference_count", difference_count)
                if difference_count is None:
                    self.row_append(5,
                                    [None, "sheet「" + group[0]['xml'] + "」" + "row「" + str(int(group[0]['row'])) + "」",
                                     "現新キャプチャーのピクセルは一致していませんので、チェックしてください。", "✕", None], CHINA_RED)
                elif difference_count == 'reverse':
                    self.row_append(5,
                                    [None, "sheet「" + group[0]['xml'] + "」" + "row「" + str(int(group[0]['row'])) + "」",
                                     "現新キャプチャーは位置が逆になっていますか？確認してください。", "✕", None], CHINA_RED)
                elif difference_count == 'marusame':
                    self.row_append(5,
                                    [None, "sheet「" + group[0]['xml'] + "」" + "row「" + str(int(group[0]['row'])) + "」",
                                     "現新キャプチャーのピクセルがまったく同じです。確認してください。", "✕", None], CHINA_RED)
                elif difference_count > 0:
                    if 0 <= difference_count <= 1:
                        item_color = QColor(255, 255, 178)
                    if 1 < difference_count <= 5:
                        item_color = QColor(255, 99, 99)
                    if 5 < difference_count:
                        item_color = QColor(139, 0, 0)
                    self.row_append(5,
                                    [None, "シート「" + group[0]['xml'] + "」" + "行目「" + str(int(group[0]['row'])) + "」",
                                     "キャプチャーには「" + str(difference_count) + "」処違うことがある。", "✕",
                                     str(difference_count)], item_color)
                else:
                    self.row_append(5,
                                    [None, "シート「" + group[0]['xml'] + "」" + "行目「" + str(int(group[0]['row'])) + "」",
                                     "よくできました。", "〇",
                                     str(difference_count)], Qt.white)
            else:
                self.row_append(5, [None, "シート「" + group[0]['xml'] + "」" + "行目「" + str(int(group[0]['row'])) + "」",
                                    "現新キャプチャーが不足です", "✕",
                                    None],
                                QColor(139, 0, 0))

        # image_output = os.path.join(temp_dir, "output.png")
        # if os.path.exists(image_output):
        #     os.remove(image_output)
        del_folder(os.path.join(temp_dir, "evidence_check"))
        del_folder(os.path.join(temp_dir, "exclude-pic"))
        # del_folder(os.path.join(get_program_path(), "evidence_check"))

        if os.path.exists(os.path.join(get_program_path(),
                                       get_str_before_first_dot(
                                           get_str_after_last_dot(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"], "\\"), ".")
                                       + "-比較結果.xlsx")):
            os.startfile(os.path.join(get_program_path(),
                                      get_str_before_first_dot(
                                          get_str_after_last_dot(EXCEL_TOTAL_MAP["EXCEL_EVIDENCE"], "\\"), ".")
                                      + "-比較結果.xlsx"))

    def get_target_by_id(self, zip_dir, draw_file_name, rids):
        drawing_file_path = os.path.join(zip_dir, 'xl', 'drawings', '_rels',
                                         draw_file_name.replace(".xml", ".xml.rels"))
        tree = ET.parse(drawing_file_path)
        root = tree.getroot()
        namespace = {'ns': 'http://schemas.openxmlformats.org/package/2006/relationships'}
        relationship = root.find(f".//ns:Relationship[@Id='{rids}']", namespace)
        if relationship is not None:
            return relationship.get('Target').split("/media/")[1]
        else:
            return None

    def pic_no_info_create(self, zip_dir, xml_name, sheet_name):
        dom_obj = xmldom.parse(zip_dir + os.sep + 'xl' + os.sep + 'drawings' + os.sep + xml_name)
        element = dom_obj.documentElement

        def _f(subElementObj):
            for anchor in subElementObj:
                xdr_from = anchor.getElementsByTagName('xdr:from')[0]
                pic_col = xdr_from.childNodes[0].firstChild.data
                pic_row = xdr_from.childNodes[2].firstChild.data

                if anchor.getElementsByTagName('xdr:pic'):
                    pic_id = \
                        anchor.getElementsByTagName('xdr:pic')[0].getElementsByTagName('xdr:nvPicPr')[
                            0].getElementsByTagName('xdr:cNvPr')[0].getAttribute('id')
                    embed_id = \
                        anchor.getElementsByTagName('xdr:pic')[0].getElementsByTagName('xdr:blipFill')[
                            0].getElementsByTagName(
                            'a:blip')[0].getAttribute('r:embed')
                    embed = self.get_target_by_id(zip_dir, xml_name, embed_id)
                    self.pic_no_cell.append({
                        'pic_id': pic_id,
                        'pic': embed,
                        'row': str(int(pic_row) + 1),
                        'column': pic_col,
                        'xml': sheet_name,
                        'anchor': "two"
                    })

        def __f(subElementObj):
            for anchor in subElementObj:
                xdr_from = anchor.getElementsByTagName('xdr:from')[0]
                pic_col = xdr_from.childNodes[0].firstChild.data
                pic_row = xdr_from.childNodes[2].firstChild.data

                if anchor.getElementsByTagName('xdr:pic'):
                    pic_id = \
                        anchor.getElementsByTagName('xdr:pic')[0].getElementsByTagName('xdr:nvPicPr')[
                            0].getElementsByTagName('xdr:cNvPr')[0].getAttribute('id')
                    embed_id = \
                        anchor.getElementsByTagName('xdr:pic')[0].getElementsByTagName('xdr:blipFill')[
                            0].getElementsByTagName(
                            'a:blip')[0].getAttribute('r:embed')
                    embed = self.get_target_by_id(zip_dir, xml_name, embed_id)
                    self.pic_no_cell.append({
                        'pic_id': pic_id,
                        'pic': embed,
                        'row': pic_row,
                        'column': pic_col,
                        'xml': sheet_name,
                        'anchor': "one"
                    })

        sub_twoCellAnchor = element.getElementsByTagName("xdr:twoCellAnchor")
        _f(sub_twoCellAnchor)
        sub_oneCellAnchor = element.getElementsByTagName("xdr:oneCellAnchor")
        if sub_oneCellAnchor is not None:
            __f(sub_oneCellAnchor)

    def app_save(self):
        """結果ファイルを保存する"""
        set_status_label(self, "結果ファイル保存中")
        result_excel_path = os.path.join(get_program_path(), self.parent.input_id.text() +
                                         "-整合性チェック結果-" + str(datetime.datetime.now()).
                                         replace('-', '').replace('.', '').
                                         replace(' ', '').replace('-', '').
                                         replace(':', '') + ".xlsx")
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "ドキュメントチェック"
        sheet2 = wb.create_sheet("記入内容チェック")
        sheet3 = wb.create_sheet("ソースチェック")
        sheet4 = wb.create_sheet("カバーのチェック")
        sheet5 = wb.create_sheet("エビデンスチェック")
        wb.save(result_excel_path)
        wb.close()

        # set_message_box("QUESTION", "結果ファイル", "結果ファイル生成完了しました。")
        msg_box = QMessageBox()
        msg_box.setWindowTitle("結果ファイル")
        msg_box.setText("結果ファイル生成完了しました。\n"
                        "目標フォルダを開きたいですか？")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg_box.setDefaultButton(QMessageBox.No)
        msg_box.button(QMessageBox.Yes).setText("はい(&Y)")
        msg_box.button(QMessageBox.No).setText("いいえ(&N)")
        result = msg_box.exec_()
        if result == QMessageBox.Yes:
            os.startfile(get_program_path())

        self.parent.status_label.setText("結果ファイルが生成完了しました。")
        self.parent.status_label.stop_blinking()
        # set_status_label(self, "結果ファイルが生成完了しました。")

    def app_exit(self):
        """退出ボタン"""
        msg_box = QMessageBox()
        msg_box.setWindowTitle("ツールメッセージ")
        msg_box.setText("ツールを終了したいですか。")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg_box.setDefaultButton(QMessageBox.No)
        msg_box.button(QMessageBox.Yes).setText("はい(&Y)")
        msg_box.button(QMessageBox.No).setText("いいえ(&N)")
        result = msg_box.exec_()
        if result == QMessageBox.Yes:
            save_file_paths(self)
            self.parent.close()


class MainWindow(QMainWindow):
    """メインウィンドウ"""

    window_shown = pyqtSignal()

    def __init__(self):
        super().__init__()

        # 创建并显示进度窗口
        self.loading_window = LoadingWindow()
        self.loading_window.show()

        self.top_group = QGroupBox("選択")
        self.top_layout_1 = QHBoxLayout()
        self.top_layout_2 = QHBoxLayout()
        self.top_layout_3 = QHBoxLayout()
        self.top_layout_4 = QHBoxLayout()
        self.top_layout_5 = QHBoxLayout()
        # DHC fengjm 2024/12/20 No.itest ADD START
        self.top_layout_itest = QHBoxLayout()
        # DHC fengjm 2024/12/20 No.itest ADD END
        # # DHC fengjm 2024/12/19 No.raido ADD START
        # self.top_layout_radio = QHBoxLayout()
        # # DHC fengjm 2024/12/19 No.raido ADD END
        self.top_layout_checkbox = QHBoxLayout()
        self.top_layout_6 = QHBoxLayout()
        self.form_layout = QFormLayout()
        self.top_layout = QVBoxLayout()
        self.label_width = 50
        self.label_test = QLabel('エビデンスだけ')
        self.input_test = QComboBox()
        self.label_id = QLabel('機能ID')
        self.input_id = QLineEdit()
        self.label_type = QLabel('区分')
        self.input_type = QComboBox()
        self.label_system = QLabel('システム')
        self.input_system = QComboBox()
        self.label_object = QLabel('対象区分')
        self.input_object = QComboBox()
        self.label_svn = QLabel('SVNパス')
        self.input_svn = QLineEdit()
        self.label_branch = QLabel('GITブランチ')
        self.input_branch = QLineEdit()
        self.button_svn = QPushButton('開く')
        self.label_manual = QLabel('マニュアル')
        self.input_manual = QLineEdit()
        self.button_manual = QPushButton('開く')
        self.label_old = QLabel('元ソース')
        self.input_old = QLineEdit()
        self.button_old = QPushButton('開く')
        self.label_new = QLabel('新ソース')
        self.input_new = QLineEdit()
        self.button_new = QPushButton('開く')
        self.label_wbs = QLabel('外部WBS')
        self.input_wbs = QLineEdit()
        self.button_wbs = QPushButton('開く')

        # DHC fengjm 2024/12/20 No.itest ADD START
        self.label_itest = QLabel('結合テストエビデンス')
        self.input_itest = QLineEdit()
        self.button_itest = QPushButton('開く')
        # DHC fengjm 2024/12/20 No.itest ADD END
        # self.input_itest.setEnabled(False)

        # # DHC fengjm 2024/12/19 No.radio ADD START
        # self.radio1 = QRadioButton("全プログラムのチェック")
        # self.radio2 = QRadioButton("エビデンス以外のチェック")
        # self.radio3 = QRadioButton("エビデンスのチェック")
        # self.radio1.setChecked(True)
        # self.radio1.toggled.connect(self.flag1_set)
        # self.radio2.toggled.connect(self.flag2_set)
        # self.radio3.toggled.connect(self.flag3_set)
        # # DHC fengjm 2024/12/19 No.radio ADD END

        # # DHC fengjm 2024/12/20 No.itest ADD START
        # self.radio4 = QRadioButton("結合テストエビデンスのチェック")
        # self.radio4.toggled.connect(self.flag4_set)
        # # DHC fengjm 2024/12/20 No.itest ADD END

        self.checkbox_all = QCheckBox("全て")
        self.checkbox_all.stateChanged.connect(self.on_checkbox_all_state_changed)
        # self.checkbox_reverse = QCheckBox("全てない")
        # self.checkbox_reverse.stateChanged.connect(self.on_checkbox_reverse_state_changed)
        self.checkbox_doc = QCheckBox("ドキュメント")
        self.checkbox_doc.setChecked(True)
        self.checkbox_doc.setDisabled(True)
        self.checkbox_doc.stateChanged.connect(self.on_checkbox_doc_state_changed)
        self.checkbox_context = QCheckBox("記入内容")
        self.checkbox_context.stateChanged.connect(self.on_checkbox_context_state_changed)
        self.checkbox_source = QCheckBox("ソース")
        self.checkbox_source.stateChanged.connect(self.on_checkbox_source_state_changed)
        self.checkbox_modify = QCheckBox("手修正")
        self.checkbox_modify.stateChanged.connect(self.on_checkbox_modify_state_changed)
        self.checkbox_coverage = QCheckBox("カバー")
        self.checkbox_coverage.stateChanged.connect(self.on_checkbox_coverage_state_changed)
        self.checkbox_pic = QCheckBox("エビデンス")
        self.checkbox_pic.stateChanged.connect(self.on_checkbox_pic_state_changed)

        self.bottom_right_group = QGroupBox('結果')
        self.bottom_right_layout = QVBoxLayout()
        self.tabs = QTabWidget()
        self.bottom_layout = QHBoxLayout()
        self.button_group = QGroupBox("操作")
        self.button_layout = QHBoxLayout()
        self.exec_button = QPushButton('実行')
        self.lock_button = QPushButton('ロック')
        self.save_button = QPushButton('結果保存')
        self.save_button.setDisabled(True)
        self.exit_button = QPushButton('退出')
        self.tips_group = QGroupBox("状態")
        self.tips_layout = QVBoxLayout()
        self.tips_layout_1 = QHBoxLayout()
        self.tips_layout_2 = QHBoxLayout()
        self.status_label = BlinkingLabel('画面初期化')
        self.tips_label = QLabel('')
        self.progress_bar = QProgressBar()
        self.main_layout = QVBoxLayout()
        self.event_handler = None
        self.initUI()

    def initUI(self):
        """initUI"""
        # self.label_test.setFixedWidth(self.label_width)
        self.label_test.setToolTip('類型')
        self.input_test.addItems(ENABILITY_TEST_TYPE)
        self.input_test.setCurrentIndex(1)
        self.label_id.setFixedWidth(self.label_width)
        self.label_id.setToolTip('機能ID')
        self.label_type.setFixedWidth(self.label_width)
        self.label_type.setToolTip('区分')
        self.input_type.addItems(ENABILITY_TYPE)
        # self.input_type.setEditable(False)
        self.input_type.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.input_type.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.label_system.setFixedWidth(self.label_width)
        self.label_system.setToolTip('システム')
        self.input_system.addItems(ENABILITY_SYSTEM)
        self.input_system.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.label_object.setFixedWidth(self.label_width)
        self.label_object.setToolTip('対象区分')
        self.input_object.addItems(ENABILITY_OBJECT)
        self.input_object.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.label_svn.setFixedWidth(self.label_width)
        self.label_svn.setToolTip('テスト成果物SVN')
        self.label_branch.setFixedWidth(60)
        self.label_branch.setToolTip('GITブランチ')
        # self.input_svn.setReadOnly(True)
        self.label_manual.setFixedWidth(self.label_width)
        self.label_manual.setToolTip('手修正マニュアルのパス')
        # self.input_manual.setReadOnly(True)
        self.label_old.setFixedWidth(self.label_width)
        self.label_old.setToolTip('基準ソース')
        # self.input_old.setReadOnly(True)
        self.label_new.setFixedWidth(self.label_width)
        self.label_new.setToolTip('手修正後のソース')
        # self.input_new.setReadOnly(True)
        self.label_wbs.setFixedWidth(self.label_width)
        self.label_wbs.setToolTip('外部wbsのパス')
        # self.input_wbs.setReadOnly(True)

        # DHC fengjm 2024/12/20 No.itest ADD START
        self.label_itest.setFixedWidth(self.label_width)
        self.label_itest.setToolTip('結合テストのエビデンス')
        # DHC fengjm 2024/12/20 No.itest ADD END

        self.top_layout_1.addWidget(self.label_test)
        self.top_layout_1.addWidget(self.input_test)
        self.top_layout_1.addWidget(self.label_id)
        self.top_layout_1.addWidget(self.input_id)
        self.top_layout_1.addWidget(self.label_type)
        self.top_layout_1.addWidget(self.input_type)
        self.top_layout_1.addWidget(self.label_system)
        self.top_layout_1.addWidget(self.input_system)
        self.top_layout_1.addWidget(self.label_object)
        self.top_layout_1.addWidget(self.input_object)
        self.top_layout_2.addWidget(self.label_svn)
        self.top_layout_2.addWidget(self.input_svn)
        self.top_layout_2.addWidget(self.button_svn)
        self.top_layout_2.addWidget(self.label_branch)
        self.top_layout_2.addWidget(self.input_branch)
        self.top_layout_3.addWidget(self.label_manual)
        self.top_layout_3.addWidget(self.input_manual)
        self.top_layout_3.addWidget(self.button_manual)
        self.top_layout_4.addWidget(self.label_old)
        self.top_layout_4.addWidget(self.input_old)
        self.top_layout_4.addWidget(self.button_old)
        self.top_layout_4.addWidget(self.label_new)
        self.top_layout_4.addWidget(self.input_new)
        self.top_layout_4.addWidget(self.button_new)
        self.top_layout_5.addWidget(self.label_wbs)
        self.top_layout_5.addWidget(self.input_wbs)
        self.top_layout_5.addWidget(self.button_wbs)

        # DHC fengjm 2024/12/20 No.itest ADD START
        self.top_layout_itest.addWidget(self.label_itest)
        self.top_layout_itest.addWidget(self.input_itest)
        self.top_layout_itest.addWidget(self.button_itest)
        # DHC fengjm 2024/12/20 No.itest ADD END

        # # DHC fengjm 2024/12/19 No.radio ADD START
        # self.top_layout_radio.addWidget(self.radio1)
        # self.top_layout_radio.addWidget(self.radio2)
        # self.top_layout_radio.addWidget(self.radio3)
        # # DHC fengjm 2024/12/19 No.radio ADD END
        # # DHC fengjm 2024/12/20 No.itest ADD START
        # self.top_layout_radio.addWidget(self.radio4)
        # # DHC fengjm 2024/12/20 No.itest ADD END
        self.top_layout_checkbox.addWidget(self.checkbox_all)
        # self.top_layout_checkbox.addWidget(self.checkbox_reverse)
        self.top_layout_checkbox.addWidget(self.checkbox_doc)
        self.top_layout_checkbox.addWidget(self.checkbox_context)
        self.top_layout_checkbox.addWidget(self.checkbox_source)
        self.top_layout_checkbox.addWidget(self.checkbox_modify)
        self.top_layout_checkbox.addWidget(self.checkbox_coverage)
        self.top_layout_checkbox.addWidget(self.checkbox_pic)

        self.top_layout.addLayout(self.top_layout_1)
        self.top_layout.addLayout(self.top_layout_2)
        self.top_layout.addLayout(self.top_layout_3)
        self.top_layout.addLayout(self.top_layout_4)
        self.top_layout.addLayout(self.top_layout_5)
        # DHC fengjm 2024/12/20 No.itest ADD START
        self.top_layout.addLayout(self.top_layout_itest)
        # DHC fengjm 2024/12/20 No.itest ADD END
        # DHC fengjm 2024/12/19 No.radio ADD START
        # self.top_layout.addLayout(self.top_layout_radio)
        # DHC fengjm 2024/12/19 No.radio ADD END
        self.top_layout.addLayout(self.top_layout_checkbox)
        self.top_layout.addLayout(self.top_layout_6)
        self.top_group.setLayout(self.top_layout)

        for i in range(6):
            tab = QWidget()
            tab_layout = QVBoxLayout()
            # table = QTableWidget(5, 3)
            # table = SortableTable()
            # table = QTableWidget(0, 4)
            table = TableWidget(self)
            table.setHorizontalHeaderLabels(['区分', 'ファイル', '備考', '状態'])
            # for row in range(5):
            #     for col in range(3):
            #         table.setItem(row, col, QTableWidgetItem(f"Step {i + 1} - Cell ({row + 1}, {col + 1})"))
            table_font = table.horizontalHeader().font()
            table_font.setBold(True)
            table.horizontalHeader().setFont(table_font)
            tab_layout.addWidget(table)
            tab.setLayout(tab_layout)
            self.tabs.addTab(tab, f"Step {i + 1}")
            if i == 0:
                self.tabs.addTab(tab, "ドキュメントチェック")
            if i == 1:
                self.tabs.addTab(tab, "記入内容チェック")
            if i == 2:
                self.tabs.addTab(tab, "ソースチェック")
            if i == 3:
                self.tabs.addTab(tab, "手修正チェック")
            if i == 4:
                self.tabs.addTab(tab, "カバーチェック")
            if i == 5:
                self.tabs.addTab(tab, "エビデンスチェック")
        self.tabs.setStyleSheet("""
                            QTabBar::tab {
                                color: darkgray;
                            }
                            QTabBar::tab:selected {
                                color: black;
                            }
                        """)
        # self.tabs.setTabEnabled(0, False)
        # self.tabs.setTabEnabled(1, False)
        # self.tabs.setTabEnabled(2, False)
        # self.tabs.setTabEnabled(3, False)
        # self.tabs.setTabEnabled(4, False)

        self.bottom_right_layout.addWidget(self.tabs)
        self.bottom_right_group.setLayout(self.bottom_right_layout)
        self.bottom_layout.addWidget(self.bottom_right_group)
        # self.exec_button.setDisabled(True)
        # self.exec_button.setStyleSheet("background-color: red")
        # self.save_button.setDisabled(True)
        self.exit_button.setStyleSheet("background-color: lightgray")
        self.button_layout.addWidget(self.exec_button)
        self.lock_button.setDisabled(True)
        self.button_layout.addWidget(self.lock_button)
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.exit_button)
        self.button_group.setLayout(self.button_layout)

        self.status_label.start_blinking()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.tips_layout_1.addWidget(self.status_label)
        self.tips_layout_2.addWidget(self.tips_label)
        self.tips_layout_2.addWidget(self.progress_bar)
        self.tips_layout.addLayout(self.tips_layout_1)
        self.tips_layout.addLayout(self.tips_layout_2)
        self.tips_group.setLayout(self.tips_layout)

        self.main_layout.addWidget(self.top_group, 2)
        self.main_layout.addLayout(self.bottom_layout, 5)
        self.main_layout.addWidget(self.button_group, 1)
        self.main_layout.addWidget(self.tips_group, 1)

        central_widget = QWidget()
        central_widget.setLayout(self.main_layout)
        self.setCentralWidget(central_widget)
        self.setLayout(self.main_layout)

        self.setWindowTitle(f'BIP-整合性チェック-[{VERSION_INFO}]-Powered by PyQt5')
        self.setGeometry(400, 200, 1000, 700)
        self.timer_init()
        # DHC fengjm 2024/12/19 No.xxx UPD START
        # self.event_handler = EventHandler(self)
        self.event_handler = EventHandler(self)
        # DHC fengjm 2024/12/19 No.xxx UPD END

        self.init_ui()
        init_config_content()
        if INPUT_ID is not None:
            self.input_id.setText(INPUT_ID)
        if INPUT_TYPE is not None:
            self.input_type.setCurrentText(INPUT_TYPE)
        if INPUT_SYSTEM is not None:
            self.input_system.setCurrentText(INPUT_SYSTEM)
        if INPUT_SVN is not None:
            self.input_svn.setText(INPUT_SVN)
        if INPUT_MANUAL is not None:
            self.input_manual.setText(INPUT_MANUAL)
        if INPUT_OLD is not None:
            self.input_old.setText(INPUT_OLD)
        if INPUT_NEW is not None:
            self.input_new.setText(INPUT_NEW)
        if INPUT_WBS is not None:
            self.input_wbs.setText(INPUT_WBS)
        # DHC fengjm 2024/12/20 No.itest ADD START
        if INPUT_ITEST is not None:
            self.input_itest.setText(INPUT_ITEST)
        # DHC fengjm 2024/12/20 No.itest ADD END
        if INPUT_BRANCH is not None:
            self.input_branch.setText(INPUT_BRANCH)
        if INPUT_OBJECT is not None:
            self.input_object.setCurrentText(INPUT_OBJECT)
        if CHECKBOX_DOC is not None:
            self.checkbox_doc.setChecked(CHECKBOX_DOC)
        if CHECKBOX_CONTEXT is not None:
            self.checkbox_context.setChecked(CHECKBOX_CONTEXT)
        if CHECKBOX_SOURCE is not None:
            self.checkbox_source.setChecked(CHECKBOX_SOURCE)
        if CHECKBOX_MODIFY is not None:
            self.checkbox_modify.setChecked(CHECKBOX_MODIFY)
        if CHECKBOX_COVERAGE is not None:
            self.checkbox_coverage.setChecked(CHECKBOX_COVERAGE)
        if CHECKBOX_PIC is not None:
            self.checkbox_pic.setChecked(CHECKBOX_PIC)

    def init_ui(self):
        """init_ui"""
        # self.button_svn.installEventFilter(self)
        # self.button_old.installEventFilter(self)
        # self.button_new.installEventFilter(self)
        self.button_svn.clicked.connect(self.event_handler.button_svn_click)
        self.button_old.clicked.connect(self.event_handler.button_old_click)
        self.button_new.clicked.connect(self.event_handler.button_new_click)
        self.button_manual.clicked.connect(self.event_handler.button_manual_click)
        self.button_wbs.clicked.connect(self.event_handler.button_wbs_click)
        # DHC fengjm 2024/12/20 No.itest ADD START
        self.button_itest.clicked.connect(self.event_handler.button_itest_click)
        # DHC fengjm 2024/12/20 No.itest ADD END
        self.input_test.currentIndexChanged.connect(self.event_handler.on_combobox_changed)

        self.exec_button.clicked.connect(self.event_handler.app_execute)
        self.lock_button.clicked.connect(self.event_handler.app_lock)
        self.save_button.clicked.connect(self.event_handler.app_save)
        self.exit_button.clicked.connect(self.event_handler.app_exit)
        self.window_shown.connect(self.on_window_shown)

        # self.input_test.setCurrentIndex(0)
        # self.input_test.setEnabled(False)

    def closeEvent(self, event):
        """closeEvent"""
        config = ConfigParser()
        config_path = get_config_file_path()
        if os.path.exists(config_path) is False:
            return
        config.read(config_path, encoding='utf-8')
        if self.input_id.text():
            config.set('Ids', 'input_id', self.input_id.text())
        if self.input_type.currentText():
            config.set('Ids', 'input_type', self.input_type.currentText())
        if self.input_system.currentText():
            config.set('Ids', 'input_system', self.input_system.currentText())
        if self.input_svn.text():
            config.set('Paths', 'input_svn', self.input_svn.text())
        if self.input_manual.text():
            config.set('Paths', 'input_manual', self.input_manual.text())
        if self.input_old.text():
            config.set('Paths', 'input_old', self.input_old.text())
        if self.input_new.text():
            config.set('Paths', 'input_new', self.input_new.text())
        if self.input_wbs.text():
            config.set('Paths', 'input_wbs', self.input_wbs.text())
        # DHC fengjm 2024/12/20 No.itest ADD START
        if self.input_itest.text():
            config.set('Paths', 'input_itest', self.input_itest.text())
        # DHC fengjm 2024/12/20 No.itest ADD END
        if self.input_itest.text():
            config.set('Paths', 'input_branch', self.input_branch.text())
        if self.input_object.currentText():
            config.set('Ids', 'input_object', self.input_object.currentText())
        if self.checkbox_doc:
            result = self.checkbox_doc.isChecked()
            config.set('Ids', 'checkbox_doc', str(result))
        if self.checkbox_context:
            result = self.checkbox_context.isChecked()
            config.set('Ids', 'checkbox_context', str(result))
        if self.checkbox_source:
            result = self.checkbox_source.isChecked()
            config.set('Ids', 'checkbox_source', str(result))
        if self.checkbox_modify:
            result = self.checkbox_modify.isChecked()
            config.set('Ids', 'checkbox_modify', str(result))
        if self.checkbox_coverage:
            result = self.checkbox_coverage.isChecked()
            config.set('Ids', 'checkbox_coverage', str(result))
        if self.checkbox_pic:
            result = self.checkbox_pic.isChecked()
            config.set('Ids', 'checkbox_pic', str(result))
        with open(get_config_file_path(), 'w', encoding='utf-8') as configfile:
            config.write(configfile)

    def on_checkbox_all_state_changed(self, state):
        if state == 2:
            # self.checkbox_doc.setChecked(True)
            self.checkbox_context.setChecked(True)
            self.checkbox_source.setChecked(True)
            self.checkbox_modify.setChecked(True)
            self.checkbox_coverage.setChecked(True)
            self.checkbox_pic.setChecked(True)
        else:
            # self.checkbox_doc.setChecked(False)
            self.checkbox_context.setChecked(False)
            self.checkbox_source.setChecked(False)
            self.checkbox_modify.setChecked(False)
            self.checkbox_coverage.setChecked(False)
            self.checkbox_pic.setChecked(False)

    def on_checkbox_doc_state_changed(self, state):
        global CHECKBOX_DOC, CHECKBOX_CONTEXT, CHECKBOX_SOURCE, CHECKBOX_MODIFY, CHECKBOX_COVERAGE, CHECKBOX_PIC
        if state == 2:
            CHECKBOX_DOC = True
        else:
            CHECKBOX_DOC = False
        print(" " + str(CHECKBOX_DOC) + " " + str(CHECKBOX_CONTEXT) + " " +
              str(CHECKBOX_SOURCE) + " " + str(CHECKBOX_MODIFY) + " " + str(CHECKBOX_COVERAGE) + " " + str(
            CHECKBOX_PIC))

    def on_checkbox_context_state_changed(self, state):
        global CHECKBOX_DOC, CHECKBOX_CONTEXT, CHECKBOX_SOURCE, CHECKBOX_MODIFY, CHECKBOX_COVERAGE, CHECKBOX_PIC
        if state == 2:
            CHECKBOX_CONTEXT = True
        else:
            CHECKBOX_CONTEXT = False
        print(" " + str(CHECKBOX_DOC) + " " + str(CHECKBOX_CONTEXT) + " " +
              str(CHECKBOX_SOURCE) + " " + str(CHECKBOX_MODIFY) + " " + str(CHECKBOX_COVERAGE) + " " + str(
            CHECKBOX_PIC))

    def on_checkbox_source_state_changed(self, state):
        global CHECKBOX_DOC, CHECKBOX_CONTEXT, CHECKBOX_SOURCE, CHECKBOX_MODIFY, CHECKBOX_COVERAGE, CHECKBOX_PIC
        if state == 2:
            CHECKBOX_SOURCE = True
        else:
            CHECKBOX_SOURCE = False
        print(" " + str(CHECKBOX_DOC) + " " + str(CHECKBOX_CONTEXT) + " " +
              str(CHECKBOX_SOURCE) + " " + str(CHECKBOX_MODIFY) + " " + str(CHECKBOX_COVERAGE) + " " + str(
            CHECKBOX_PIC))

    def on_checkbox_modify_state_changed(self, state):
        global CHECKBOX_DOC, CHECKBOX_CONTEXT, CHECKBOX_SOURCE, CHECKBOX_MODIFY, CHECKBOX_COVERAGE, CHECKBOX_PIC
        if state == 2:
            CHECKBOX_MODIFY = True
        else:
            CHECKBOX_MODIFY = False
        print(" " + str(CHECKBOX_DOC) + " " + str(CHECKBOX_CONTEXT) + " " +
              str(CHECKBOX_SOURCE) + " " + str(CHECKBOX_MODIFY) + " " + str(CHECKBOX_COVERAGE) + " " + str(
            CHECKBOX_PIC))

    def on_checkbox_coverage_state_changed(self, state):
        global CHECKBOX_DOC, CHECKBOX_CONTEXT, CHECKBOX_SOURCE, CHECKBOX_MODIFY, CHECKBOX_COVERAGE, CHECKBOX_PIC
        if state == 2:
            CHECKBOX_COVERAGE = True
        else:
            CHECKBOX_COVERAGE = False
        print(" " + str(CHECKBOX_DOC) + " " + str(CHECKBOX_CONTEXT) + " " +
              str(CHECKBOX_SOURCE) + " " + str(CHECKBOX_MODIFY) + " " + str(CHECKBOX_COVERAGE) + " " + str(
            CHECKBOX_PIC))

    def on_checkbox_pic_state_changed(self, state):
        global CHECKBOX_DOC, CHECKBOX_CONTEXT, CHECKBOX_SOURCE, CHECKBOX_MODIFY, CHECKBOX_COVERAGE, CHECKBOX_PIC
        if state == 2:
            CHECKBOX_PIC = True
        else:
            CHECKBOX_PIC = False
        print(" " + str(CHECKBOX_DOC) + " " + str(CHECKBOX_CONTEXT) + " " +
              str(CHECKBOX_SOURCE) + " " + str(CHECKBOX_MODIFY) + " " + str(CHECKBOX_COVERAGE) + " " + str(
            CHECKBOX_PIC))

    # def eventFilter(self, obj, event):
    #     # if obj == self.input_kinoid and event.type() == event.FocusOut:
    #     #     self.change_text_color(self.input_kinoid.text())
    #     if obj == self.button_svn:
    #         if event.type() == event.MouseButtonRelease:
    #             self.folder_open("SVNパス選択")
    #
    #     if obj == self.button_old:
    #         if event.type() == event.MouseButtonRelease:
    #             self.folder_open("現行ソースパス選択")
    #
    #     if obj == self.button_new:
    #         if event.type() == event.MouseButtonRelease:
    #             self.folder_open("新ソースパス選択")
    #
    #     return super().eventFilter(obj, event)
    #
    # def folder_open(self, folder_name):
    #     """目標開く"""
    #     try:
    #         options = QFileDialog.Options()
    #         options |= QFileDialog.DontUseNativeDialog
    #         if self.input_svn.text() is not None:
    #             open_path = self.input_svn.text()
    #         folder_path = QFileDialog.getExistingDirectory(self, folder_name,
    #                                                        directory=open_path,
    #                                                        options=options)
    #         if folder_path:
    #             print(folder_path)
    #             self.input_svn.setText(folder_path)
    #         return True
    #     except Exception as e:
    #         print("An error occurred : ", e)
    #         raise

    # # DHC fengjm 2024/12/19 No.radio ADD START
    # def flag1_set(self):
    #     global RADIO_FLAG
    #     RADIO_FLAG = 0
    #     global ITEST_FLAG
    #     ITEST_FLAG = False
    #
    # def flag2_set(self):
    #     global RADIO_FLAG
    #     RADIO_FLAG = 1
    #     global ITEST_FLAG
    #     ITEST_FLAG = False
    #
    # def flag3_set(self):
    #     global RADIO_FLAG
    #     RADIO_FLAG = 2
    #     global ITEST_FLAG
    #     ITEST_FLAG = False
    #
    # # DHC fengjm 2024/12/19 No.radio ADD END
    # # DHC fengjm 2024/12/20 No.itest ADD START
    # def flag4_set(self):
    #     global ITEST_FLAG
    #     ITEST_FLAG = True
    #
    # # DHC fengjm 2024/12/20 No.itest ADD END
    def on_button_clicked(self):
        self.step_bar.advance_step()

    def tab_switched(self, index):
        self.tabs.setCurrentIndex(index)

    def timer_init(self):
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_datetime)
        self.timer.start(1000)

    def update_datetime(self):
        self.current_datetime = QDateTime.currentDateTime().toString(Qt.ISODate)
        self.tips_label.setText(f'{self.current_datetime}')

    def showEvent(self, event):
        super().showEvent(event)
        self.window_shown.emit()  # 在窗口显示时发出信号

    def on_window_shown(self):
        print("主窗口已显示！")
        self.loading_window.close()


class LoadingWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("起動中...")
        self.setGeometry(300, 300, 300, 100)

        layout = QVBoxLayout()
        self.label = QLabel("アプリ実行中、少々お待ちください...", self)
        layout.addWidget(self.label)

        self.setLayout(layout)


class Worker(QObject):
    finished = pyqtSignal()

    def run(self, window):
        window.show()
        self.finished.emit()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Windows')  # Windows , windowsvista , Fusion

    window = MainWindow()
    window.show()

    sys.exit(app.exec_())
    # app = QApplication(sys.argv)
    # app.setStyle('Windows')  # Windows , windowsvista , Fusion
    #
    # # 创建并显示进度窗口
    # loading_window = LoadingWindow()
    # loading_window.show()
    #
    # # 创建主窗口
    # main_window = MainWindow()
    # worker = Worker()
    #
    # # 连接信号
    # worker.finished.connect(loading_window.close)
    # worker.finished.connect(main_window.show)
    #
    # # 启动主窗口初始化的线程
    # threading.Thread(target=lambda: worker.run(main_window)).start()
    # # window = MainWindow()
    # # window.show()
    # sys.exit(app.exec_())
